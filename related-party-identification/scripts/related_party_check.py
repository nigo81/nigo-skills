#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
关联方核查引擎 / Related Party Identification Engine
=====================================================
读取 cicpa-company-query 完整导出的 _files 目录（52个维度 xlsx），
对其中所有公司做八层规则比对，输出多 sheet Excel 核查报告。

规则集对齐证监会/财政部近年处罚案例（蓝山科技、天沃科技、卓朗科技、
爱康科技、合纵科技等），把"监管认定的最低核查动作"固化成自动判定。

用法:
    python3 related_party_check.py \
        --data-dir "完整维度导出_files/" \
        --target "被审计单位全称" \
        -o "关联方核查报告.xlsx"
"""
import argparse
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from itertools import combinations
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ============================================================
# 常量
# ============================================================

# 基础工商信息列映射（A:企业名称 ... W:经营范围）
BASIC_COLS = {
    "name": 1, "company_id": 2, "status": 3, "legal_person": 4,
    "capital": 5, "found_date": 6, "province": 7, "city": 8, "district": 9,
    "phone": 10, "website": 11, "email": 12, "credit_code": 13,
    "reg_no": 14, "org_code": 15, "insured": 16, "company_type": 17,
    "industry1": 18, "industry2": 19, "industry3": 20,
    "former_name": 21, "address": 22, "business_scope": 23,
}

# 政府/公共实体名（作为实控人/受益人/关键人员出现时排除，避免所有国企误报关联）
EXCLUDED_ENTITIES = {
    "国务院国有资产监督管理委员会", "国务院", "财政部", "国家发改委",
    "中华人民共和国财政部", "中华人民共和国国务院",
    "地方国资委", "省国资委", "市国资委",
    "上海市国有资产监督管理委员会", "北京市国有资产监督管理委员会",
    "广东省人民政府国有资产监督管理委员会", "浙江省人民政府国有资产监督管理委员会",
    "江苏省人民政府国有资产监督管理委员会", "山东省人民政府国有资产监督管理委员会",
    "四川省政府国有资产监督管理委员会", "湖北省人民政府国有资产监督管理委员会",
    "湖南省人民政府国有资产监督管理委员会", "河南省人民政府国有资产监督管理委员会",
    "河北省人民政府国有资产监督管理委员会", "福建省人民政府国有资产监督管理委员会",
    "安徽省人民政府国有资产监督管理委员会", "辽宁省人民政府国有资产监督管理委员会",
    "陕西省人民政府国有资产监督管理委员会", "山西省人民政府国有资产监督管理委员会",
    "重庆市国有资产监督管理委员会", "天津市人民政府国有资产监督管理委员会",
    "江西省国有资产监督管理委员会", "云南省人民政府国有资产监督管理委员会",
    "广西壮族自治区人民政府国有资产监督管理委员会", "贵州省人民政府国有资产监督管理委员会",
    "新疆维吾尔自治区人民政府国有资产监督管理委员会",
}
# 公众人物/知名企业家（作为股东/董监高出现时不作为关联信号）
KNOWN_PUBLIC_FIGURES = {
    "马云", "马化腾", "李彦宏", "雷军", "周鸿祎", "刘强东", "丁磊",
    "张一鸣", "王兴", "黄峥", "张朝阳", "史玉柱", "陈天桥", "贾跃亭",
    "董明珠", "任正非", "柳传志", "王健林", "许家印", "郭广昌",
    "沈南鹏", "张磊", "李开复", "俞敏洪", "罗永浩", "李斌", "何小鹏",
    "李想", "王小川", "周源", "唐岩", "冯鑫", "王石", "孙宏斌",
    "曹德旺", "宗庆后", "马明哲", "姚劲波", "龚宇",
    "蔡文胜", "吴世春", "朱啸虎",
    "刘永好", "马蔚华", "张近东", "梁建章",
    "江南春", "傅盛", "王小川",
}

# 公共邮箱域名（同这些域名不算关联信号）
PUBLIC_EMAIL_DOMAINS = {
    "qq.com", "163.com", "126.com", "sina.com", "sohu.com", "hotmail.com",
    "gmail.com", "outlook.com", "foxmail.com", "yeah.net", "139.com", "189.cn",
    "aliyun.com", "wo.cn", "vip.qq.com", "vip.163.com", "tom.com", "21cn.com",
    "188.com", "2980.com", "263.net", "mail.com", "yahoo.com", "live.com",
    "icloud.com", "me.com", "msn.com",
}

# 维度文件名（容错：缺文件跳过）
FILES = {
    "basic": "基础工商信息.xlsx",
    "shareholder": "股东信息.xlsx",
    "shareholder_new": "最新公示股东.xlsx",
    "actual_controller": "实际控制人.xlsx",
    "ultimate_beneficiary": "最终受益人.xlsx",
    "invest_new": "对外投资（新）.xlsx",
    "invest": "对外投资.xlsx",
    "holding": "参控股企业.xlsx",
    "branch": "分支机构.xlsx",
    "core_team": "核心团队.xlsx",
    "main_persons": "主要人员（高管）.xlsx",
    "legal_change": "法定代表人变更.xlsx",
    "customer": "客户.xlsx",
    "supplier": "供应商.xlsx",
    "guarantee": "对外担保.xlsx",
    "pledge": "股权质押.xlsx",
    "pledge2": "股权出质.xlsx",
    "mortgage": "动产抵押.xlsx",
    "change": "变更记录.xlsx",
    "trademark": "商标.xlsx",
    "software": "软件著作权.xlsx",
    "patent": "专利信息.xlsx",
    "wechat": "微信公众号.xlsx",
    "invoice": "发票信息.xlsx",
    "abnormal": "经营异常.xlsx",
}

# 风险等级
HARD = "🔴硬关联"      # high
MEDIUM = "🟡可疑红旗"   # medium
LOW = "🟢轻微异常"      # low

# Excel 颜色
FILL_HARD = PatternFill("solid", fgColor="FFC7CE")     # 红
FILL_MEDIUM = PatternFill("solid", fgColor="FFEB9C")   # 黄
FILL_LOW = PatternFill("solid", fgColor="C6EFCE")      # 绿
FILL_HEADER = PatternFill("solid", fgColor="305496")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_TITLE = Font(bold=True, size=14, color="305496")


# ============================================================
# 数据加载（容错：文件缺失/空文件返回空表）
# ============================================================

def load_workbook_safe(path):
    """加载 xlsx，返回 (rows_list, ncols)；文件缺失或异常返回 ([], 0)。"""
    if not os.path.exists(path):
        return [], 0
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
            if any(c is not None and str(c).strip() != "" for c in r):
                rows.append(r)
        return rows, ws.max_column
    except Exception as e:
        print(f"  ⚠️ 读取失败 {os.path.basename(path)}: {e}", file=sys.stderr)
        return [], 0


def build_dim_index(data_dir):
    """加载所有维度文件，返回 {key: {公司名: [行...]}} 索引。"""
    index = {}
    for key, fname in FILES.items():
        path = os.path.join(data_dir, fname)
        rows, ncols = load_workbook_safe(path)
        by_company = defaultdict(list)
        for r in rows:
            # 公司名称一般在第2列（B），基础工商在第1列（A），分支机构/担保/法人变更在第1列
            if key == "basic":
                name = r[0] if r else None
            elif key in ("branch", "guarantee", "legal_change", "invoice"):
                name = r[0] if r else None  # 这些表公司名在A列
            else:
                name = r[1] if len(r) > 1 else None
            if name and str(name).strip():
                by_company[str(name).strip()].append(r)
        index[key] = dict(by_company)
        cnt = sum(len(v) for v in by_company.values())
        companies = len(by_company)
        if companies:
            print(f"  ✓ {fname}: {companies} 家公司, {cnt} 条记录")
        else:
            print(f"  · {fname}: 无数据")
    return index


# ============================================================
# 数据清洗
# ============================================================

def strip_html(s):
    if s is None:
        return ""
    return re.sub(r"<[^>]+>", "", str(s))


def normalize_address(addr):
    """地址归一化：去HTML、去空格标点、去邮编。"""
    if not addr:
        return ""
    s = strip_html(addr).strip()
    s = s.replace(" ", "").replace("\u3000", "").replace("\t", "")
    s = re.sub(r"[，,。；;:：、（）()\[\]【】「」“”‘’\-—_~`'\"#]", "", s)
    s = re.sub(r"(省|市|自治区|特别行政区)", "", s)  # 去省市后缀干扰
    s = re.sub(r"\d{6}(?!\d)", "", s)  # 去6位邮编
    return s


def address_relationship(a1, a2):
    """比较两个地址。返回 (类型, 说明) 或 None。
    类型: exact(完全相同) / contains(前缀包含，同栋楼不同房间)"""
    n1, n2 = normalize_address(a1), normalize_address(a2)
    if not n1 or not n2 or len(n1) < 6 or len(n2) < 6:
        return None
    if n1 == n2:
        return ("exact", f"完全相同: {n1[:30]}")
    # 前缀包含（较短的地址是较长地址的前缀 → 同栋楼不同房间）
    short, long_ = (n1, n2) if len(n1) <= len(n2) else (n2, n1)
    if len(short) >= 8 and long_.startswith(short):
        return ("contains", f"同址不同室: {short[:25]}…")
    # 提取"路/街/号+楼栋"核心段比对（同栋楼）
    core1 = re.search(r"(.{0,4}(?:路|街|道|巷|弄|号|大厦|大楼|大厦|商务楼|科技园|产业园)[^室层栋楼户]+)", n1)
    core2 = re.search(r"(.{0,4}(?:路|街|道|巷|弄|号|大厦|大楼|商务楼|科技园|产业园)[^室层栋楼户]+)", n2)
    if core1 and core2:
        c1, c2 = core1.group(1), core2.group(1)
        if c1 and c2 and len(c1) >= 8 and c1 == c2:
            return ("same_building", f"疑似同栋楼: {c1[:25]}…")
    return None


def normalize_phones(phone_str):
    """拆分多值电话，归一化，返回电话集合。"""
    if not phone_str:
        return set()
    s = strip_html(phone_str)
    parts = re.split(r"[,，;；、/\s]+", s)
    result = set()
    for p in parts:
        p = p.strip()
        if not p:
            continue
        p = re.sub(r"[\-—()（）\s]", "", p)
        if p.startswith("+86"):
            p = p[3:]
        elif p.startswith("86") and len(p) > 11:
            p = p[2:]
        # 手机
        if re.fullmatch(r"1[3-9]\d{9}", p):
            result.add(p)
        # 座机 0XX-XXXXXXXX
        elif re.fullmatch(r"0\d{10,11}", p):
            result.add(p)  # 含区号
            result.add(p[3:])  # 去区号版
        elif re.fullmatch(r"\d{7,8}", p):
            result.add(p)
        # 4位区号+8位
        elif re.fullmatch(r"0\d{9,10}", p):
            result.add(p)
    # 过滤掉太短的
    return {x for x in result if len(x) >= 7}


def phone_segment_adjacent(phones_a, phones_b):
    """座机号段相邻检测：同一区号+局向的连续号码（去末1位后前缀相同）。

    监管实务中，两家公司用同一号段的座机（如蓝山案 82284751 / 康居 82284750），
    几乎必然是同一办公地点的相邻分机号，是强关联信号。
    只对座机（0开头）做检测，手机号段相邻误报率高不纳入。
    """
    for pa in phones_a:
        for pb in phones_b:
            if pa.startswith("0") and pb.startswith("0") and len(pa) >= 10 and len(pb) >= 10:
                if pa[:-1] == pb[:-1] and pa != pb:
                    return (pa, pb)
    return None


def normalize_emails(email_str):
    """返回 (完整邮箱列表, 非公共域名列表)。"""
    if not email_str:
        return [], []
    s = strip_html(email_str).lower()
    found = re.findall(r"[\w.+-]+@[\w.-]+\.\w{2,}", s)
    emails, domains = [], []
    for em in found:
        emails.append(em)
        domain = em.split("@")[-1]
        if domain not in PUBLIC_EMAIL_DOMAINS:
            domains.append(domain)
    return emails, list(set(domains))


def normalize_name(n):
    """姓名归一化（去HTML、去空格）。"""
    if not n:
        return ""
    s = strip_html(n).strip()
    return s.replace(" ", "").replace("\u3000", "")


def parse_capital(s):
    """解析注册资本字符串为元（float）。'1.5亿'→1.5e8, '2000万'→2e7。"""
    if not s:
        return None
    s2 = strip_html(str(s))
    m = re.search(r"([\d.]+)\s*亿", s2)
    if m:
        return float(m.group(1)) * 1e8
    m = re.search(r"([\d.]+)\s*万", s2)
    if m:
        return float(m.group(1)) * 1e4
    m = re.search(r"([\d.]+)", s2)
    if m:
        return float(m.group(1))
    return None


def parse_insured(s):
    """参保人数 → int。"""
    if s is None:
        return None
    s2 = strip_html(str(s))
    m = re.search(r"\d+", s2)
    return int(m.group()) if m else None


def parse_found_date(s):
    """成立日期 → 'YYYY-MM-DD' 或 'YYYY' 字符串。"""
    if not s:
        return None
    s2 = strip_html(str(s))[:10]
    m = re.match(r"(\d{4})[-/年](\d{1,2})?", s2)
    if m:
        y = m.group(1)
        mo = m.group(2)
        return f"{y}-{int(mo):02d}" if mo else y
    return None


# ============================================================
# 公司实体（指纹）
# ============================================================

@dataclass
class Company:
    name: str
    legal_person: str = ""
    phones: set = field(default_factory=set)
    emails: list = field(default_factory=list)
    email_domains: list = field(default_factory=list)
    addresses: list = field(default_factory=list)
    websites: list = field(default_factory=list)
    capital: Optional[float] = None
    found_date: Optional[str] = None
    insured: Optional[int] = None
    former_names: list = field(default_factory=list)
    business_scope: str = ""
    industry: str = ""
    status: str = ""
    # 关系数据
    shareholders: list = field(default_factory=list)        # [(股东名, 比例str, 是否机构)]
    actual_controllers: list = field(default_factory=list)  # [实控人名]
    beneficiaries: list = field(default_factory=list)       # [(受益人名, 比例)]
    core_persons: list = field(default_factory=list)        # [(姓名, 职务)]
    main_persons: list = field(default_factory=list)        # [(姓名, 职务)]
    investments: list = field(default_factory=list)         # [(被投资企业, 比例)]
    holdings: list = field(default_factory=list)            # [(参控股企业, 比例)]
    customers_marked_related: list = field(default_factory=list)  # 客户表自带关联方名
    suppliers_marked_related: list = field(default_factory=list)
    trademarks: list = field(default_factory=list)
    softwares: list = field(default_factory=list)
    wechats: list = field(default_factory=list)
    pledges_out: list = field(default_factory=list)   # [(出质人, 质权人)]
    mortgages: list = field(default_factory=list)     # [(抵押人, 抵押权人)]
    changes: list = field(default_factory=list)       # [(日期, 项目, 变更前, 变更后)]
    legal_changes: list = field(default_factory=list) # [(变更前法人, 变更后法人)]
    invoice_addr: str = ""
    invoice_phone: str = ""


def build_company(name, dim, basic_row):
    """从基础工商行 + 各维度构建 Company 指纹。"""
    c = Company(name=name)
    if basic_row:
        c.legal_person = normalize_name(basic_row[BASIC_COLS["legal_person"] - 1])
        c.phones = normalize_phones(basic_row[BASIC_COLS["phone"] - 1])
        c.emails, c.email_domains = normalize_emails(basic_row[BASIC_COLS["email"] - 1])
        addr = basic_row[BASIC_COLS["address"] - 1]
        if addr:
            c.addresses.append(strip_html(addr))
        web = basic_row[BASIC_COLS["website"] - 1]
        if web:
            c.websites.append(strip_html(str(web)))
        c.capital = parse_capital(basic_row[BASIC_COLS["capital"] - 1])
        c.found_date = parse_found_date(basic_row[BASIC_COLS["found_date"] - 1])
        c.insured = parse_insured(basic_row[BASIC_COLS["insured"] - 1])
        fn = basic_row[BASIC_COLS["former_name"] - 1]
        if fn and str(fn).strip():
            c.former_names = [normalize_name(x) for x in re.split(r"[,，;；、]", strip_html(fn)) if x.strip()]
        c.business_scope = strip_html(basic_row[BASIC_COLS["business_scope"] - 1])[:200]
        c.status = strip_html(basic_row[BASIC_COLS["status"] - 1])
        ind = basic_row[BASIC_COLS["industry2"] - 1] or basic_row[BASIC_COLS["industry1"] - 1]
        c.industry = strip_html(ind) if ind else ""

    # 股东
    for r in dim.get("shareholder", {}).get(name, []):
        if len(r) >= 7:
            sh = normalize_name(r[2])
            ratio = strip_html(str(r[4])) if r[4] else ""
            is_inst = strip_html(str(r[6])) if r[6] else ""
            if sh:
                c.shareholders.append((sh, ratio, is_inst in ("是", "True", "true", "1")))
    # 最新公示股东
    for r in dim.get("shareholder_new", {}).get(name, []):
        if len(r) >= 6:
            sh = normalize_name(r[3])
            ratio = strip_html(str(r[5])) if r[5] else ""
            if sh:
                c.shareholders.append((sh, ratio, "1" in strip_html(str(r[4] or ""))))
    # 实控人 / 最终受益人
    for r in dim.get("actual_controller", {}).get(name, []):
        if len(r) >= 3 and r[2]:
            c.actual_controllers.append(normalize_name(r[2]))
    for r in dim.get("ultimate_beneficiary", {}).get(name, []):
        if len(r) >= 6 and r[5]:
            c.beneficiaries.append((normalize_name(r[5]), strip_html(str(r[6] or ""))))
    # 核心团队 / 主要人员
    for r in dim.get("core_team", {}).get(name, []):
        if len(r) >= 5 and r[3]:
            c.core_persons.append((normalize_name(r[3]), strip_html(str(r[4] or ""))))
    for r in dim.get("main_persons", {}).get(name, []):
        if len(r) >= 4 and r[2]:
            c.main_persons.append((normalize_name(r[2]), strip_html(str(r[3] or ""))))
    # 对外投资 / 参控股
    for r in dim.get("invest_new", {}).get(name, []):
        if len(r) >= 5 and r[2]:
            c.investments.append((normalize_name(r[2]), strip_html(str(r[4] or ""))))
    for r in dim.get("holding", {}).get(name, []):
        if len(r) >= 4 and r[2]:
            c.holdings.append((normalize_name(r[2]), strip_html(str(r[3] or ""))))
    # 客户/供应商自带关联方标注
    for r in dim.get("customer", {}).get(name, []):
        if len(r) >= 8 and r[7]:
            c.customers_marked_related.append(normalize_name(r[7]))
    for r in dim.get("supplier", {}).get(name, []):
        if len(r) >= 8 and r[7]:
            c.suppliers_marked_related.append(normalize_name(r[7]))
    # 商标 / 软件 / 公众号
    for r in dim.get("trademark", {}).get(name, []):
        if len(r) >= 3 and r[2]:
            c.trademarks.append(strip_html(str(r[2])))
    for r in dim.get("software", {}).get(name, []):
        if len(r) >= 5 and r[4]:
            c.softwares.append(strip_html(str(r[4])))
    for r in dim.get("wechat", {}).get(name, []):
        if len(r) >= 4 and r[3]:
            c.wechats.append(strip_html(str(r[3])))
    # 股权质押 / 动产抵押
    for r in dim.get("pledge", {}).get(name, []):
        if len(r) >= 6:
            c.pledges_out.append((normalize_name(r[3] or ""), normalize_name(r[5] or "")))
    for r in dim.get("mortgage", {}).get(name, []):
        if len(r) >= 6:
            c.mortgages.append((normalize_name(r[4] or ""), normalize_name(r[5] or "")))
    # 变更记录 / 法人变更
    for r in dim.get("change", {}).get(name, []):
        if len(r) >= 5:
            c.changes.append((strip_html(str(r[2] or "")), strip_html(str(r[3] or "")),
                              strip_html(str(r[4] or ""))[:30], strip_html(str(r[5] or ""))[:30]))
    for r in dim.get("legal_change", {}).get(name, []):
        if len(r) >= 3:
            c.legal_changes.append((normalize_name(r[1] or ""), normalize_name(r[2] or "")))
    # 发票信息（补充地址电话）
    for r in dim.get("invoice", {}).get(name, []):
        if len(r) >= 8:
            if r[4]:
                c.invoice_addr = strip_html(str(r[4]))
            if r[6]:
                c.invoice_phone = strip_html(str(r[6]))
    return c


def all_person_names(c: Company):
    """公司所有关键人员姓名集合（法人+股东中的自然人+核心团队+主要人员+实控人+受益人）。"""
    names = set()
    if c.legal_person:
        names.add(c.legal_person)
    for sh, _, is_inst in c.shareholders:
        if not is_inst and sh and len(sh) <= 6:  # 自然人股东
            names.add(sh)
    for n, _ in c.core_persons:
        names.add(n)
    for n, _ in c.main_persons:
        names.add(n)
    for n in c.actual_controllers:
        names.add(n)
    for n, _ in c.beneficiaries:
        names.add(n)
    return {n for n in names if n and len(n) >= 2 and n not in EXCLUDED_ENTITIES and n not in KNOWN_PUBLIC_FIGURES}


def all_addresses(c: Company):
    """公司所有地址（现址+发票地址+变更历史地址）。"""
    addrs = list(c.addresses)
    if c.invoice_addr:
        addrs.append(c.invoice_addr)
    for _, _, before, _ in c.changes:
        if "地址" in (c.changes[0][1] if c.changes else ""):
            pass
    # 变更记录里地址变更的历史地址
    for _, item, before, after in c.changes:
        if "地址" in item and before:
            addrs.append(before)
    return [a for a in addrs if a]


# ============================================================
# 规则引擎（八层）
# ============================================================

def rule1_fingerprint(ca: Company, cb: Company):
    """维度1: 工商指纹重合（地址/电话/邮箱/网址）。"""
    hits = []
    # 电话
    common_phones = ca.phones & cb.phones
    if common_phones:
        hits.append(("phone", HARD, f"联系电话相同: {','.join(list(common_phones)[:3])}", "蓝山/卓朗/达志科技案"))
    else:
        # 电话号段相邻（座机同一区号+局向，末位不同 → 同一办公地点的连续号码）
        seg = phone_segment_adjacent(ca.phones, cb.phones)
        if seg:
            hits.append(("phone_segment", HARD, f"座机号段相邻(同一办公地): {seg[0]}↔{seg[1]}", "蓝山科技案(号段相邻)"))
    # 邮箱完全相同
    common_emails = set(ca.emails) & set(cb.emails)
    if common_emails:
        hits.append(("email", HARD, f"邮箱完全相同: {','.join(list(common_emails)[:2])}", "爱康科技案"))
    # 邮箱同域名（非公共）
    common_domains = set(ca.email_domains) & set(cb.email_domains)
    if common_domains:
        hits.append(("email_domain", HARD, f"企业邮箱同域名: {','.join(list(common_domains)[:2])}", "天沃/爱康/志高机械案"))
    # 地址
    for a1 in all_addresses(ca):
        for a2 in all_addresses(cb):
            rel = address_relationship(a1, a2)
            if rel:
                kind, desc = rel
                level = HARD if kind == "exact" else (MEDIUM if kind == "same_building" else MEDIUM)
                case = "天沃科技案(同楼同座)" if kind != "exact" else "达志科技案(地址相同)"
                hits.append(("address", level, desc, case))
                break
    # 网址同域名
    def web_domain(w):
        return re.sub(r"https?://", "", str(w)).split("/")[0].replace("www.", "").lower() if w else ""
    wa = {web_domain(w) for w in ca.websites if w}
    wb = {web_domain(w) for w in cb.websites if w}
    common_web = (wa & wb) - {""}
    if common_web:
        hits.append(("website", MEDIUM, f"官网同域名: {','.join(list(common_web)[:2])}", "实务红旗"))
    return hits


def rule2_personnel(ca: Company, cb: Company):
    """维度2: 关键人员重合。"""
    hits = []
    pa, pb = all_person_names(ca), all_person_names(cb)
    common = pa & pb
    # 法人重合是强信号
    if ca.legal_person and cb.legal_person and ca.legal_person == cb.legal_person and len(ca.legal_person) >= 2:
        hits.append(("legal_person", HARD, f"法定代表人同一人: {ca.legal_person}", "爱康/合纵科技案"))
    # 其他人员重合
    other_common = common - ({ca.legal_person, cb.legal_person} if ca.legal_person == cb.legal_person else set())
    if other_common:
        # 区分：是否任董监高（强）/ 仅股东（中）
        is_exec = any(n in {x[0] for x in ca.core_persons + ca.main_persons} for n in other_common)
        level = HARD if is_exec else MEDIUM
        case = "卓朗科技案(对方监事为本公司员工)" if is_exec else "人员重合"
        hits.append(("person", level, f"关键人员重合: {','.join(list(other_common)[:3])}", case))
    # 实控人/最终受益人一致（排除政府/公共实体）
    common_ctrl = (set(ca.actual_controllers) & set(cb.actual_controllers)) - EXCLUDED_ENTITIES
    if common_ctrl:
        hits.append(("controller", HARD, f"实际控制人相同: {','.join(list(common_ctrl)[:2])}", "股权控制类"))
    common_ben = ({b[0] for b in ca.beneficiaries} & {b[0] for b in cb.beneficiaries}) - EXCLUDED_ENTITIES
    if common_ben:
        hits.append(("beneficiary", HARD, f"最终受益人相同: {','.join(list(common_ben)[:2])}", "股权控制类"))
    return hits


def rule3_counterparty_profile(ca: Company, cb: Company, target_set):
    """维度3: 客商异常画像（对非审计对象群的公司做单方向画像）。

    审计对象群 = 被审计单位 + 其重要子公司（--target 逗号分隔）。
    只对「审计对象 ↔ 对手方」做画像；双方都是/都不是审计对象则跳过。
    """
    hits = []
    ca_is_target = ca.name in target_set
    cb_is_target = cb.name in target_set
    if ca_is_target and not cb_is_target:
        party = cb
    elif cb_is_target and not ca_is_target:
        party = ca
    else:
        return hits  # 同属审计对象群（集团内）或都是外部对手方，不做画像
    flags = []
    # 成立时间短（数据更新时距今不足24个月）—— 用成立年份判断
    if party.found_date:
        year = party.found_date[:4]
        if year.isdigit() and int(year) >= 2018:  # 近年成立（相对宽松）
            flags.append(f"{year}年成立(较新)")
    # 注册资本 vs 典型交易额：注册资本极小
    if party.capital is not None and party.capital < 1e6:  # <100万
        flags.append(f"注册资本仅{party.capital/1e4:.0f}万(过小)")
    # 参保人数
    if party.insured is not None and party.insured < 5:
        flags.append(f"参保人数{party.insured}(疑似空壳)")
    elif party.insured is not None and party.insured == 0:
        flags.append("参保人数0(空壳特征)")
    # 地址居民楼
    for a in party.addresses:
        if any(k in a for k in ["小区", "花园", "公寓", "苑", "村", "组"]):
            flags.append(f"注册地址疑似居民楼")
            break
    # 经营范围空或极广
    if not party.business_scope or len(party.business_scope) < 10:
        flags.append("经营范围缺失/极简")
    if flags:
        hits.append(("profile", MEDIUM if len(flags) >= 2 else LOW,
                     "；".join(flags[:4]), "专网通信/爱康案"))
    return hits


def rule4_disclosed_gap(dim, target_name, company_names):
    """维度4: 客户/供应商表自带关联方标注 vs 在场公司。"""
    hits = []
    # target 的客户/供应商表里被标注为"关联方"的公司名
    marked = set()
    for r in dim.get("customer", {}).get(target_name, []) + dim.get("supplier", {}).get(target_name, []):
        if len(r) >= 8 and r[7]:
            marked.add(normalize_name(r[7]))
    # 对手方中命中"已标注关联方"
    present = set(company_names)
    matched = marked & present
    for m in matched:
        if m != target_name:
            hits.append(("marked_related", HARD,
                         f"工商数据已标注为关联方: {m}", "洛娃集团案(客户已是关联方)"))
    return hits


def rule5_equity(ca: Company, cb: Company):
    """维度5: 股权/控制穿透。"""
    hits = []
    # A 投资了 B 或 B 投资了 A
    inv_a = {i[0] for i in ca.investments + ca.holdings}
    inv_b = {i[0] for i in cb.investments + cb.holdings}
    if cb.name in inv_a:
        hits.append(("invest", HARD, f"{ca.name} 对外投资/参控股 {cb.name}", "投资关系"))
    if ca.name in inv_b:
        hits.append(("invest", HARD, f"{cb.name} 对外投资/参控股 {ca.name}", "投资关系"))
    # 共同股东（自然人或机构同名）
    sh_a = {s[0] for s in ca.shareholders}
    sh_b = {s[0] for s in cb.shareholders}
    common_sh = (sh_a & sh_b) - {""}
    # 排除极常见机构名干扰（如纯"国有企业"等），自然人股东重合是强信号
    person_sh = {s for s in common_sh if len(s) <= 6}
    if person_sh:
        hits.append(("common_shareholder", MEDIUM,
                     f"共同股东: {','.join(list(person_sh)[:3])}", "股权穿透"))
    return hits


def rule6_historical(ca: Company, cb: Company):
    """维度6: 历史关联痕迹。"""
    hits = []
    # 曾用名命中
    cb_names = {cb.name} | set(cb.former_names)
    ca_names = {ca.name} | set(ca.former_names)
    former_hit = (set(ca.former_names) & cb_names) | (set(cb.former_names) & ca_names)
    if former_hit:
        hits.append(("former_name", MEDIUM, f"曾用名匹配: {','.join(list(former_hit)[:2])}", "历史关联"))
    # 曾任法人：A 现法人曾是 B 的法人（法人变更记录）
    ca_past_legals = {lc[0] for lc in ca.legal_changes if lc[0]} | {lc[1] for lc in ca.legal_changes if lc[1]}
    cb_past_legals = {lc[0] for lc in cb.legal_changes if lc[0]} | {lc[1] for lc in cb.legal_changes if lc[1]}
    # A 现法人 在 B 的法人变更历史里
    if ca.legal_person and ca.legal_person in cb_past_legals and ca.legal_person != cb.legal_person:
        hits.append(("past_legal", MEDIUM,
                     f"{ca.name}现法人{ca.legal_person}曾任{cb.name}法人", "华道生物案(代持痕迹)"))
    # 变更记录里的历史地址撞对方现地址
    cb_addr_set = {normalize_address(a) for a in all_addresses(cb)}
    for _, item, before, after in ca.changes:
        if "地址" in item and before:
            if normalize_address(before) in cb_addr_set and len(normalize_address(before)) >= 8:
                hits.append(("past_address", MEDIUM,
                             f"{ca.name}曾用地址与{cb.name}现地址重合", "历史关联"))
                break
    return hits


def rule7_guarantee(ca: Company, cb: Company):
    """维度7: 担保/资金链。"""
    hits = []
    cb_names = {cb.name} | set(cb.former_names)
    ca_names = {ca.name} | set(ca.former_names)
    # 股权质押：出质人或质权人是对方
    for pledgor, pledgee in ca.pledges_out:
        if pledgor in cb_names or pledgee in cb_names:
            hits.append(("pledge", MEDIUM,
                         f"股权质押关联(出质:{pledgor}/质权:{pledgee})", "康得新案"))
            break
    # 动产抵押
    for mortgagor, mortgagee in ca.mortgages:
        if mortgagor in cb_names or mortgagee in cb_names:
            hits.append(("mortgage", MEDIUM,
                         f"动产抵押关联(抵押:{mortgagor}/抵押权:{mortgagee})", "担保链"))
            break
    return hits


def rule8_intangible(ca: Company, cb: Company):
    """维度8: 无形资产/资源共用。"""
    hits = []
    # 商标同名（排除"图形"等纯商标类别标签）
    common_tm = (set(ca.trademarks) & set(cb.trademarks)) - {"图形", "文字", "字母", "数字", "颜色"}
    if common_tm:
        hits.append(("trademark", MEDIUM, f"同名商标: {','.join(list(common_tm)[:2])}", "无形资产共用"))
    # 软件著作权同名
    common_sw = set(ca.softwares) & set(cb.softwares)
    if common_sw:
        hits.append(("software", MEDIUM, f"同名软件著作权: {','.join(list(common_sw)[:2])}", "无形资产共用"))
    # 微信公众号同名
    common_wc = set(ca.wechats) & set(cb.wechats)
    if common_wc:
        hits.append(("wechat", LOW, f"同名微信公众号: {','.join(list(common_wc)[:2])}", "资源共用"))
    return hits


# ============================================================
# 主比对 & 汇总
# ============================================================

def compare_pair(ca, cb, dim, target_set):
    """对一对公司跑全部规则，返回 [hit_dict...]。target_set 为审计对象群（含子公司）。"""
    results = []
    # rule3 需要方向参数（判断谁是审计主体、谁是对手方），单独调用；其余规则统一双参数
    rules_2arg = [
        (rule1_fingerprint, "工商指纹重合"),
        (rule2_personnel, "关键人员重合"),
        (rule5_equity, "股权控制穿透"),
        (rule6_historical, "历史关联痕迹"),
        (rule7_guarantee, "担保资金链"),
        (rule8_intangible, "无形资产共用"),
    ]
    for rule_fn, dim_name in rules_2arg:
        try:
            for field_key, level, evidence, case in rule_fn(ca, cb):
                results.append({
                    "company_a": ca.name, "company_b": cb.name,
                    "dimension": dim_name, "field": field_key,
                    "level": level, "evidence": evidence, "case_ref": case,
                })
        except Exception as e:
            print(f"  ⚠️ 规则异常 {dim_name} {ca.name}↔{cb.name}: {e}", file=sys.stderr)
    # rule3: 客商异常画像（需 target_set 判定方向，只对 审计对象↔对手方 做画像）
    try:
        for field_key, level, evidence, case in rule3_counterparty_profile(ca, cb, target_set):
            results.append({
                "company_a": ca.name, "company_b": cb.name,
                "dimension": "客商异常画像", "field": field_key,
                "level": level, "evidence": evidence, "case_ref": case,
            })
    except Exception as e:
        print(f"  ⚠️ 规则异常 客商异常画像 {ca.name}↔{cb.name}: {e}", file=sys.stderr)
    return results


def level_rank(level):
    return {HARD: 3, MEDIUM: 2, LOW: 1}.get(level, 0)


def aggregate(all_hits, target_set):
    """按公司对汇总。返回 [{pair, levels, dimensions, max_level, is_related, ...}]。target_set 为审计对象群。"""
    by_pair = defaultdict(list)
    for h in all_hits:
        pair = tuple(sorted([h["company_a"], h["company_b"]]))
        by_pair[pair].append(h)
    summary = []
    for pair, hits in by_pair.items():
        levels = [h["level"] for h in hits]
        dims = sorted({h["dimension"] for h in hits})
        max_level = max(levels, key=level_rank)
        # 是否关联方判定：任一 HARD → 是；仅有 MEDIUM → 可疑；仅 LOW → 轻微
        has_hard = any(l == HARD for l in levels)
        has_medium = any(l == MEDIUM for l in levels)
        if has_hard:
            is_related = "🔴 是（建议确认）"
            suggestion = "实施函证、实地走访、资金流水核对；询问管理层并要求披露"
        elif has_medium:
            is_related = "🟡 可疑（建议关注）"
            suggestion = "结合交易背景进一步核查；关注交易商业合理性"
        else:
            is_related = "🟢 轻微异常"
            suggestion = "记录备查，必要时跟进"
        ca, cb = pair
        ca_t, cb_t = ca in target_set, cb in target_set
        if ca_t and cb_t:
            relation = "审计对象群内部(集团内)"
        elif ca_t or cb_t:
            relation = "审计对象-交易对手"
        else:
            relation = "交易对手之间(隐性关联)"
        summary.append({
            "company_a": ca, "company_b": cb, "relation_type": relation,
            "is_related": is_related, "max_level": max_level,
            "hit_count": len(hits), "dimensions": "、".join(dims),
            "evidence": " | ".join(sorted({h["evidence"] for h in hits}))[:200],
            "suggestion": suggestion,
            "case_ref": "、".join(sorted({h["case_ref"] for h in hits}))[:60],
            "hits": hits,
        })
    summary.sort(key=lambda x: (-level_rank(x["max_level"]), -x["hit_count"]))
    return summary


# ============================================================
# Excel 输出
# ============================================================

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_sheet(ws, headers, rows, level_col=None):
    ws.append(headers)
    style_header(ws, len(headers))
    for r in rows:
        ws.append(list(r))
        if level_col is not None:
            lv = r[level_col]
            fill = FILL_HARD if "硬" in str(lv) else (FILL_MEDIUM if "可疑" in str(lv) or "是" in str(lv) else (FILL_LOW if "轻微" in str(lv) else None))
            if fill:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=c).fill = fill
    for col in range(1, len(headers) + 1):
        max_len = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max(max_len * 1.8, 12), 50)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)


def write_report(out_path, summary, all_hits, companies, dim, target_display, data_completeness):
    wb = openpyxl.Workbook()
    # 汇总 sheet
    ws0 = wb.active
    ws0.title = "汇总判断"
    ws0["A1"] = f"关联方核查报告｜审计对象：{target_display}"
    ws0["A1"].font = FONT_TITLE
    ws0.merge_cells("A1:I1")
    headers = ["公司A", "公司B", "关系类型", "是否关联方", "风险等级", "命中维度", "命中数", "关键证据", "建议审计程序"]
    rows = [
        (s["company_a"], s["company_b"], s["relation_type"], s["is_related"],
         s["max_level"], s["dimensions"], s["hit_count"], s["evidence"], s["suggestion"])
        for s in summary
    ]
    ws0.append([])  # 空行
    start_row = 3
    ws0.append(headers)
    for r in rows:
        ws0.append(list(r))
    # 格式化汇总表（从第4行起）
    for c in range(1, len(headers) + 1):
        cell = ws0.cell(row=start_row + 1, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws0.row_dimensions[start_row + 1].height = 28
    for ri, s in enumerate(summary):
        row_idx = start_row + 2 + ri
        lv = s["max_level"]
        fill = FILL_HARD if "硬" in lv else (FILL_MEDIUM if "可疑" in lv else FILL_LOW)
        for c in range(1, len(headers) + 1):
            cell = ws0.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, w in zip("ABCDEFGHI", [22, 22, 16, 16, 14, 22, 8, 45, 35]):
        ws0.column_dimensions[col].width = w
    ws0.freeze_panes = "A5"

    # 各维度 sheet
    dim_sheets = {
        "工商指纹重合": ("01_工商指纹重合", ["公司A", "公司B", "子项", "风险等级", "证据", "监管出处"]),
        "关键人员重合": ("02_关键人员重合", ["公司A", "公司B", "子项", "风险等级", "证据", "监管出处"]),
        "客商异常画像": ("03_客商异常画像", ["公司A", "公司B", "子项", "风险等级", "证据", "监管出处"]),
        "股权控制穿透": ("05_股权控制穿透", ["公司A", "公司B", "子项", "风险等级", "证据", "监管出处"]),
        "历史关联痕迹": ("06_历史关联痕迹", ["公司A", "公司B", "子项", "风险等级", "证据", "监管出处"]),
        "担保资金链": ("07_担保资金链", ["公司A", "公司B", "子项", "风险等级", "证据", "监管出处"]),
        "无形资产共用": ("08_无形资产共用", ["公司A", "公司B", "子项", "风险等级", "证据", "监管出处"]),
    }
    # 已知关联方差集 单独处理
    marked_hits = [h for h in all_hits if h["dimension"] == "已知关联方差集"] if False else []
    # （rule4 结果已在 all_hits 里以 dimension="已知关联方差集" 加入，需在比对时设置）
    for dname, (sname, headers_d) in dim_sheets.items():
        ws = wb.create_sheet(sname)
        rows_d = [(h["company_a"], h["company_b"], h["field"], h["level"],
                   h["evidence"], h["case_ref"]) for h in all_hits if h["dimension"] == dname]
        write_sheet(ws, headers_d, rows_d, level_col=3)

    # 已知关联方差集 sheet（若 rule4 命中）
    marked = [h for h in all_hits if h["dimension"] == "已知关联方差集"]
    if marked:
        ws = wb.create_sheet("04_已知关联方差集")
        write_sheet(ws, ["审计对象", "关联方(工商数据已标注)", "风险等级", "监管出处"],
                    [(h["company_a"], h["evidence"].replace("工商数据已标注为关联方: ", ""), h["level"], h["case_ref"]) for h in marked],
                    level_col=2)

    # 数据概览 sheet
    ws_ov = wb.create_sheet("输入数据概览")
    ws_ov.append(["公司名称", "法人", "成立日期", "注册资本", "参保人数", "登记状态", "行业", "数据完整度"])
    style_header(ws_ov, 8)
    for name in companies:
        c = companies[name]
        comp = data_completeness.get(name, "")
        ws_ov.append([name, c.legal_person, c.found_date or "",
                      f"{c.capital/1e4:.0f}万" if c.capital else "",
                      c.insured if c.insured is not None else "", c.status, c.industry, comp])
    for col, w in zip("ABCDEFGH", [30, 12, 12, 12, 10, 10, 18, 14]):
        ws_ov.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"\n✅ 报告已生成: {out_path}")
    print(f"   比对公司对: {len(summary)} | 命中关联嫌疑: {sum(1 for s in summary if '硬' in s['max_level'] or '可疑' in s['max_level'])}")


# ============================================================
# main
# ============================================================

def main():
    ap = argparse.ArgumentParser(description="关联方核查引擎")
    ap.add_argument("--data-dir", required=True, help="cicpa 完整导出解压后的 _files 目录")
    ap.add_argument("--target", required=True, help="被审计单位全称（多个用逗号分隔，应包含其重要子公司）")
    ap.add_argument("-o", "--output", default="关联方核查报告.xlsx", help="输出 Excel 路径")
    args = ap.parse_args()

    # 审计对象群：被审计单位 + 其重要子公司（逗号分隔）。
    # 把子公司一并纳入，是因为造假常发生在子公司层面（如卓朗案造假挂在子公司卓朗发展名下，
    # 只把母公司当 target 会导致母公司与造假客商的直接关联信号偏弱）。
    target_set = {t.strip() for t in args.target.split(",") if t.strip()}
    target_display = " / ".join(sorted(target_set))

    if not os.path.isdir(args.data_dir):
        print(f"❌ 数据目录不存在: {args.data_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"📂 加载数据: {args.data_dir}")
    dim = build_dim_index(args.data_dir)

    # 公司清单（以基础工商信息为准）
    company_names = list(dim.get("basic", {}).keys())
    if not company_names:
        print("❌ 基础工商信息.xlsx 无数据或未找到，无法核查", file=sys.stderr)
        sys.exit(1)
    missing = target_set - set(company_names)
    if missing:
        print(f"⚠️ 以下审计对象不在基础工商信息表中（可能名称不完全匹配）: {missing}")
        print(f"   表内公司: {company_names[:5]}...")
    print(f"\n🏢 待核查公司 {len(company_names)} 家: {', '.join(company_names[:8])}{'...' if len(company_names)>8 else ''}")
    print(f"🎯 审计对象群({len(target_set)}家): {target_display}")

    # 构建公司指纹
    print("\n🔧 构建公司指纹...")
    companies = {}
    data_completeness = {}
    for name in company_names:
        basic_rows = dim.get("basic", {}).get(name, [])
        basic_row = basic_rows[0] if basic_rows else None
        c = build_company(name, dim, basic_row)
        companies[name] = c
        # 数据完整度
        filled = sum([bool(c.legal_person), bool(c.phones), bool(c.emails),
                      bool(c.addresses), c.capital is not None, c.insured is not None])
        data_completeness[name] = f"{filled}/6"

    # 全图两两比对
    print("\n🔍 执行八层规则比对...")
    all_hits = []
    names = list(companies.keys())
    for ca, cb in combinations(names, 2):
        all_hits.extend(compare_pair(companies[ca], companies[cb], dim, target_set))
    # 维度4: 已知关联方差集（对审计对象群里每个公司分别检查其客商表的关联方标注）
    for t in target_set:
        for field_key, level, evidence, case in rule4_disclosed_gap(dim, t, names):
            all_hits.append({
                "company_a": t, "company_b": evidence.replace("工商数据已标注为关联方: ", ""),
                "dimension": "已知关联方差集", "field": field_key,
                "level": level, "evidence": evidence, "case_ref": case,
            })

    print(f"\n📊 命中规则 {len(all_hits)} 条")
    # 汇总
    summary = aggregate(all_hits, target_set)
    hard_cnt = sum(1 for s in summary if "硬" in s["max_level"])
    med_cnt = sum(1 for s in summary if "可疑" in s["max_level"])
    print(f"   🔴 硬关联公司对: {hard_cnt}")
    print(f"   🟡 可疑红旗公司对: {med_cnt}")
    if summary:
        print(f"\n   Top 嫌疑:")
        for s in summary[:5]:
            print(f"     {s['max_level']} {s['company_a']} ↔ {s['company_b']}: {s['dimensions']}")

    # 输出 Excel
    write_report(args.output, summary, all_hits, companies, dim, target_display, data_completeness)


if __name__ == "__main__":
    main()
