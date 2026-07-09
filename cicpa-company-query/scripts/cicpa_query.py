#!/usr/bin/env python3
"""
注协系统工商信息批量查询工具
流程：检查 cookie → 有效则直接用 → 无效则打开浏览器让用户登录 → 自动抓取 cookie → 批量查询 → 下载 Excel
无需账号密码配置文件，登录完全在浏览器中完成。

修改记录（2026-06-04）:
1. ensure_session(): 解耦浏览器登录 + saved_at 兜底，避免非交互环境崩溃
2. browser_export(): 返回解压目录路径，删除死代码，使用 tempfile 修复竞争条件，添加 ZIP 路径遍历防护
3. discover_subsidiaries(): 重写为递归穿透多层子公司（max_depth 参数）
4. _check_cookie_valid(): 更严格的 cookie 验证（只检查 status_code）
5. 新增 --max-depth 命令行参数
"""

import json
import os
import sys
import time
import tempfile
import requests
import openpyxl
import openpyxl.utils
from pathlib import Path

# 配置
SCRIPT_DIR = Path(__file__).parent
BASE_DIR = Path.cwd()

# cookies: 始终存在工作目录（避免多项目 cookie 冲突）
COOKIE_FILE = BASE_DIR / ".cicpa_cookies.json"

# API 端点
CMIS_BASE = "https://cmis.cicpa.org.cn"
ZSK_BASE = "https://zsk-cmis.cicpa.org.cn"


def _save_cookies(cookies_dict, session_token=None):
    """保存 cookies 到文件"""
    data = {"cookies": cookies_dict, "session_token": session_token, "saved_at": time.time()}
    COOKIE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(COOKIE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _load_cookies():
    """从文件加载 cookies"""
    if COOKIE_FILE.exists():
        with open(COOKIE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def _build_session(cookies_dict):
    """根据 cookie 字典构建 requests session"""
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "Referer": f"{ZSK_BASE}/companylibrarynew/",
        "Origin": ZSK_BASE,
    })
    # 按域名分发 cookies
    zsk_domain_cookies = {"XSRF-TOKEN", "yuqing_whole_jsessionid", "cicpa_token",
                          "cicpa_ticket", "companyVerifyCode", "userid", "u_name"}
    for name, value in cookies_dict.items():
        domain = "zsk-cmis.cicpa.org.cn" if name in zsk_domain_cookies else ".cicpa.org.cn"
        session.cookies.set(name, value, domain=domain)

    xsrf_token = cookies_dict.get("XSRF-TOKEN", "")
    if xsrf_token:
        session.headers["X-Xsrf-Token"] = xsrf_token

    return session


def _check_cookie_valid(session):
    """用 API 请求检测 cookie 是否有效，返回 True/False"""
    try:
        resp = session.get(f"{ZSK_BASE}/open/industry_chain_api/v1/search/get_user_index", timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            # 有效 cookie 返回 status_code=0
            if data.get("status_code") == 0:
                return True
        return False
    except Exception:
        return False


def _ensure_playwright():
    """确保 playwright 已安装，未安装则自动 pip install。
    返回 True 表示可用，False 表示失败。
    """
    try:
        import playwright
        return True
    except ImportError:
        pass

    print("⚙ 正在安装 playwright...")
    import subprocess
    result = subprocess.run(
        [sys.executable, "-m", "pip", "install", "playwright", "-q"],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        print(f"❌ playwright 安装失败: {result.stderr[:200]}")
        print("   请手动运行: pip install playwright")
        return False
    print("✓ playwright 安装完成")
    return True


def _detect_browser_channel():
    """检测系统已安装的 Chromium 内核浏览器，避免下载 Chromium。
    
    Playwright channel 参数会直接使用系统已安装的浏览器。
    支持国内外常见浏览器：Chrome、Edge、360、QQ浏览器、搜狗、Brave、Chromium。
    
    返回 (channel, 显示名) 或 (None, None)。
    """
    import platform

    if platform.system() == "Darwin":
        checks = [
            ("chrome",  "Google Chrome",    "/Applications/Google Chrome.app"),
            ("msedge",  "Microsoft Edge",   "/Applications/Microsoft Edge.app"),
            ("chrome",  "360浏览器",         "/Applications/360Chrome.app"),
            ("chrome",  "QQ浏览器",          "/Applications/QQBrowser.app"),
            ("chrome",  "搜狗浏览器",        "/Applications/SogouExplorer.app"),
            ("chrome",  "Brave",            "/Applications/Brave Browser.app"),
            ("chromium","Chromium",         "/Applications/Chromium.app"),
        ]
    elif platform.system() == "Windows":
        checks = [
            ("chrome",  "Google Chrome",    r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
            ("chrome",  "Google Chrome x86",r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
            ("msedge",  "Microsoft Edge",   r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
            ("chrome",  "360浏览器",         r"C:\Program Files\360\360Chrome\Chrome\Application\360chrome.exe"),
            ("chrome",  "QQ浏览器",          r"C:\Program Files\Tencent\QQBrowser\QQBrowser.exe"),
            ("chrome",  "搜狗浏览器",        r"C:\Program Files\SogouExplorer\SogouExplorer.exe"),
        ]
    else:  # Linux
        checks = [
            ("chrome",  "Google Chrome",    "/usr/bin/google-chrome"),
            ("msedge",  "Microsoft Edge",   "/usr/bin/microsoft-edge"),
            ("chrome",  "360浏览器",         "/usr/bin/360chrome"),
            ("chromium","Chromium",         "/usr/bin/chromium-browser"),
        ]

    for channel, display_name, path in checks:
        if os.path.exists(path):
            return channel, display_name

    return None, None


def _ensure_browser():
    """确保有可用的浏览器。
    优先使用系统已安装的浏览器，没有才下载 Playwright 自带的 Chromium。
    返回 (channel, display_name, error_message)。
    """
    channel, display_name = _detect_browser_channel()
    if channel:
        print(f"✓ 检测到浏览器: {display_name}")
        return channel, display_name, None

    # 没有系统浏览器，下载 Chromium
    print("⚙ 未检测到系统浏览器，正在下载 Chromium（首次需要，约 100MB）...")
    import subprocess
    result = subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        return None, None, f"Chromium 下载失败: {result.stderr[:200]}\n请手动运行: playwright install chromium"
    print("✓ Chromium 下载完成")
    return None, "Chromium (已下载)", None


def _do_browser_login():
    """打开浏览器让用户登录，登录完成后抓取 cookies。
    
    使用 Playwright 启动独立浏览器实例，导航到登录页。
    采用 input() 阻塞等待用户确认（用户在浏览器中完成登录后按回车）。
    
    注意：此函数需要在交互式终端中运行（有 stdin）。
    如果没有交互式终端（如被 AI 子进程调用），应改用 do_login_guided()。
    """
    import threading

    # 自动安装 playwright
    if not _ensure_playwright():
        return None

    # 自动检测/下载浏览器
    channel, display_name, err = _ensure_browser()
    if err:
        print(f"❌ {err}")
        return None

    from playwright.sync_api import sync_playwright

    launch_opts = {"headless": False}
    if channel:
        launch_opts["channel"] = channel

    print()
    print("=" * 55)
    print(f"  正在启动浏览器: {display_name}")
    print("=" * 55)
    print()
    print("  请在浏览器中完成以下操作：")
    print("    1. 选择用户类型（如：注册会计师）")
    print("    2. 输入会员编号和密码")
    print("    3. 拖动滑块完成验证")
    print("    4. 点击登录")
    print("    5. 登录成功后，在主页上点击「行业执业知识库」")
    print()
    print("  ⏳ 登录完成后回到终端按回车键确认。")
    print()

    cookie_dict = None

    with sync_playwright() as p:
        browser = p.chromium.launch(**launch_opts)
        context = browser.new_context()
        page = context.new_page()

        # 直接导航到登录页
        try:
            page.goto("https://cmis.cicpa.org.cn/#/login",
                      wait_until="domcontentloaded", timeout=30000)
        except Exception as e:
            print(f"⚠ 页面加载超时，但浏览器已打开，可以继续操作")

        # 等待用户在终端按回车确认
        try:
            input(">>> 登录完成后按回车: ")
        except EOFError:
            # 没有交互式终端（被 AI 子进程调用）
            print()
            print("⚠ 检测到非交互式终端，无法等待用户输入")
            print("  请在终端中手动运行: python cicpa_query.py --login")
            try:
                browser.close()
            except Exception:
                pass
            return None

        print()
        print(">>> 正在提取 cookies...")

        # 确保在 zsk 知识库页面上
        try:
            current_url = page.url
            if "zsk-cmis" not in current_url:
                print(">>> 正在跳转到行业知识库...")
                page.goto("https://zsk-cmis.cicpa.org.cn/companylibrarynew/",
                          wait_until="domcontentloaded", timeout=15000)
                time.sleep(3)
        except Exception as e:
            print(f"⚠ 跳转失败: {e}")

        # 提取所有 cookies
        try:
            all_cookies = context.cookies()
            cookie_dict = {c["name"]: c["value"] for c in all_cookies}
            session_token = None
            try:
                session_token = page.evaluate(
                    '() => { try { return localStorage.getItem("session_token") } catch(e) { return null } }')
            except Exception:
                pass

            _save_cookies(cookie_dict, session_token)
            print(f">>> ✅ Cookies 已保存（共 {len(cookie_dict)} 个）")
        except Exception as e:
            print(f"❌ 提取 cookies 失败: {e}")

        try:
            browser.close()
        except Exception:
            pass

    return cookie_dict


def save_cookies_from_dict(cookies_dict):
    """直接保存外部传入的 cookie 字典（供 AI 通过 agent-browser 传入）
    
    Args:
        cookies_dict: {name: value} 格式的 cookie 字典
    Returns:
        bool: 保存是否成功
    """
    _save_cookies(cookies_dict)
    # 验证
    session = _build_session(cookies_dict)
    if _check_cookie_valid(session):
        print(f"✅ Cookies 已保存并验证有效（共 {len(cookies_dict)} 个）")
        return True
    else:
        print(f"⚠ Cookies 已保存但 API 验证未通过（共 {len(cookies_dict)} 个）")
        return False


def do_login():
    """独立登录命令：打开浏览器 → 用户登录 → 保存 cookies → 验证有效性。
    
    这个函数由 --login 参数触发，与查询逻辑完全解耦。
    """
    cookie_dict = _do_browser_login()
    if not cookie_dict:
        return False

    # 验证 cookies 是否有效
    session = _build_session(cookie_dict)
    if _check_cookie_valid(session):
        print()
        print("=" * 55)
        print("  ✅ 登录成功，cookies 已验证有效！")
        print(f"  Cookie 文件: {COOKIE_FILE}")
    else:
        print()
        print("⚠ Cookies 已保存但 API 验证未通过")
        print("  可能需要重新登录或在注协系统中多停留几秒")
        print(f"  Cookie 文件: {COOKIE_FILE}")

    return True


def ensure_session(force_login=False):
    """统一入口：检查 cookie → 有效返回 session → 无效打开浏览器登录

    Args:
        force_login: 强制重新登录（忽略已有 cookie）

    Returns:
        requests.Session 或 None（登录失败时）
    """
    cookie_file = COOKIE_FILE
    if cookie_file.exists():
        try:
            with open(cookie_file, "r") as f:
                cookie_data = json.load(f)
        except:
            cookie_data = {}
    else:
        cookie_data = {}

    cookies = cookie_data.get("cookies", {})

    if cookies and not force_login:
        # saved_at 存在时检查时效性
        saved_at = cookie_data.get("saved_at")
        if saved_at is not None:
            age_hours = (time.time() - saved_at) / 3600
            if age_hours >= 24:
                print(f"Cookie 已过期（{age_hours:.1f} 小时前保存）")
                cookies = {}

        if cookies:
            session = _build_session(cookies)
            if _check_cookie_valid(session):
                return session
            else:
                print("Cookie 无效，需要重新登录")

    # 需要登录但非交互环境 → 直接报错
    if not sys.stdin.isatty():
        print("❌ Cookie 无效/过期，当前为非交互环境无法自动登录")
        print("   请通过以下方式获取 cookie：")
        print("   1. 运行: python cicpa_query.py --login")
        print("   2. 或用 Playwright MCP 工具登录后，用 --save-cookies 保存")
        return None

    # 交互环境才走浏览器登录
    cookie_dict = _do_browser_login()
    if cookie_dict:
        save_cookies_from_dict(cookie_dict)
        return _build_session(cookie_dict)
    return None


def upload_companies(session, company_names):
    """
    上传企业名单进行批量查询
    返回 batch_id
    """
    # 创建临时 Excel 文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["企业名称"])
    for name in company_names:
        ws.append([name])

    tmp_fd, tmp_file = tempfile.mkstemp(suffix=".xlsx", prefix="cicpa_batch_")
    os.close(tmp_fd)
    wb.save(tmp_file)

    # 获取 XSRF token
    xsrf_token = session.cookies.get("XSRF-TOKEN", "")
    if not xsrf_token:
        # 尝试从其他 cookie 名称获取
        for cookie in session.cookies:
            if "xsrf" in cookie.name.lower() or "csrf" in cookie.name.lower():
                xsrf_token = cookie.value
                break

    url = f"{ZSK_BASE}/open/industry_chain_api/v1/batch_search/upload_local"

    with open(tmp_file, "rb") as f:
        files = {"file": ("companies.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers = {}
        if xsrf_token:
            headers["X-Xsrf-Token"] = xsrf_token

        resp = session.post(url, files=files, headers=headers)

    os.remove(tmp_file)

    if resp.status_code != 200:
        raise Exception(f"上传失败: HTTP {resp.status_code}, {resp.text}")

    result = resp.json()
    if result.get("status_code") != 0:
        raise Exception(f"上传失败: {result.get('status_msg')}")

    data = result["data"]
    print(f"  上传成功: 共 {data['all_num']} 家，匹配 {data['hit_num']} 家，未匹配 {data['not_hit_num']} 家")
    return data["batch_id"]


def query_results(session, batch_id):
    """
    查询批量搜索结果
    返回企业详细信息列表
    """
    url = f"{ZSK_BASE}/open/industry_chain_api/v1/batch_search/enter_search"

    # 获取 XSRF token
    xsrf_token = session.cookies.get("XSRF-TOKEN", "")
    
    payload = {
        "batch_id": batch_id,
        "page": 1,
        "pageSize": 10000,
    }
    
    headers = {}
    if xsrf_token:
        headers["X-Xsrf-Token"] = xsrf_token

    resp = session.post(url, json=payload, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"查询失败: HTTP {resp.status_code}, {resp.text[:200]}")

    result = resp.json()
    if result.get("status_code") != 0:
        raise Exception(f"查询失败: {result.get('status_msg')}")

    return result["data"]


def search_company(query, page_size=5, force_login=False):
    """
    轻量单企业搜索（不走 batch upload，直接调 home_search API）
    适合查 1 家公司的 org_id 等基础信息。

    :param query: 搜索关键词（企业名称/简称/代码）
    :param page_size: 返回条数（默认 5）
    :param force_login: 是否强制重新登录
    :return: 企业列表（每条含 org_id, name, legal_person 等）
    """
    session = ensure_session(force_login=force_login)
    if session is None:
        print("❌ 无法获取有效 session，请重试")
        return []

    url = f"{ZSK_BASE}/open/industry_chain_api/v1/search/home_search"

    xsrf_token = session.cookies.get("XSRF-TOKEN", "")
    headers = {"Content-Type": "application/json;charset=UTF-8"}
    if xsrf_token:
        headers["X-Xsrf-Token"] = xsrf_token

    payload = {
        "condition": [],
        "query": query,
        "from": 0,
        "page": 1,
        "page_size": page_size,
        "size": page_size,
        "type": "company",
        "ranges": [],
    }

    resp = session.post(url, json=payload, headers=headers)
    if resp.status_code != 200:
        print(f"❌ 搜索失败: HTTP {resp.status_code}")
        return []

    result = resp.json()
    if result.get("status_code") != 0:
        print(f"❌ 搜索失败: {result.get('status_msg')}")
        return []

    data = result.get("data", {})
    companies = data.get("list", [])
    return companies


def batch_query(company_names, force_login=False):
    """
    批量查询工商信息
    :param company_names: 企业名称列表
    :param force_login: 是否强制重新登录
    :return: 企业详细信息列表
    """
    session = ensure_session(force_login=force_login)
    if session is None:
        print("❌ 无法获取有效 session，请重试")
        return []

    print(f"\n开始批量查询 {len(company_names)} 家企业...")

    # 上传
    batch_id = upload_companies(session, company_names)

    # 等待处理
    print("  等待查询结果...")
    time.sleep(2)

    # 查询结果
    data = query_results(session, batch_id)
    print(f"  查询完成: 找到 {data['total']} 条结果")

    return data["list"]


def discover_subsidiaries(audit_target, threshold=50.0, max_depth=5):
    """递归发现子公司，支持多层穿透

    Args:
        audit_target: 被审计单位名称
        threshold: 持股比例阈值（百分比）
        max_depth: 最大穿透深度（默认5层）

    Returns:
        dict: {子公司名称: {"ratio": 持股比例, "depth": 穿透深度, "parent": 直接母公司}}
    """
    session = ensure_session()
    if session is None:
        print("❌ 无法获取有效 session")
        return {}

    # 搜索审计目标获取 org_id
    search_url = "https://zsk-cmis.cicpa.org.cn/open/industry_chain_api/v1/search/home_search"
    xsrf = session.cookies.get("XSRF-TOKEN", "")
    headers = {}
    if xsrf:
        headers["X-Xsrf-Token"] = xsrf

    search_data = {
        "keyword": audit_target,
        "pageNo": 1,
        "pageSize": 1,
        "searchType": "enterprise"
    }

    try:
        resp = session.post(search_url, json=search_data, headers=headers, timeout=15)
        result = resp.json()
        if result.get("status_code") != 0 or not result.get("data"):
            print(f"  未找到企业: {audit_target}")
            return {}

        items = result["data"].get("searchResultList", [])
        if not items:
            print(f"  未找到企业: {audit_target}")
            return {}

        org_id = items[0].get("org_id") or items[0].get("id")
        if not org_id:
            print(f"  未获取到 org_id: {audit_target}")
            return {}
    except Exception as e:
        print(f"  搜索失败: {e}")
        return {}

    # 递归穿透
    all_subs = {}  # {name: {"ratio": float, "depth": int, "parent": str}}
    visited = {audit_target}  # 避免环

    def _penetrate(parent_name, parent_org_id, current_depth):
        if current_depth > max_depth:
            return

        equity_url = f"https://zsk-cmis.cicpa.org.cn/open/enterprise_info_api/v3/atlas/enterprise_equity?orgid={parent_org_id}"

        try:
            resp = session.get(equity_url, headers=headers, timeout=30)
            result = resp.json()
        except Exception as e:
            print(f"  穿透 {parent_name} 失败 (深度{current_depth}): {e}")
            return

        if result.get("status_code") != 0:
            return

        # 解析对外投资
        data = result.get("data", {})
        investments = data.get("invests", {}).get("children", []) or data.get("investList", []) or []

        for inv in investments:
            name = inv.get("entName", "") or inv.get("name", "").strip()
            ratio_str = inv.get("investRatio", "") or inv.get("ratio", "") or inv.get("czbl", "")
            sub_org_id = inv.get("orgId", "") or inv.get("orgid", "") or inv.get("entId", "")

            if not name or name in visited:
                continue

            # 解析持股比例
            try:
                ratio = float(str(ratio_str).replace("%", "").strip())
            except (ValueError, AttributeError):
                ratio = 0

            if ratio >= threshold:
                visited.add(name)
                all_subs[name] = {
                    "ratio": ratio,
                    "depth": current_depth,
                    "parent": parent_name,
                    "org_id": sub_org_id
                }
                print(f"  {'  ' * current_depth}├─ {name} (持股{ratio}%, 深度{current_depth})")

                # 递归穿透子公司的子公司
                if sub_org_id:
                    _penetrate(name, sub_org_id, current_depth + 1)

    print(f"开始穿透 {audit_target} 的子公司（阈值{threshold}%, 最大深度{max_depth}层）...")
    _penetrate(audit_target, org_id, 1)

    print(f"\n共发现 {len(all_subs)} 家子公司（穿透深度{max_depth}层）")
    return all_subs


def get_company_detail(org_id, company_name=None):
    """
    查询单家企业的详细信息（基本信息、股东、高管、股权结构）

    :param org_id: 企业 org_id（必填）
    :param company_name: 企业名称（可选，仅用于显示）
    :return: dict {"basic_info": {}, "shareholders": {}, "main_persons": {}, "equity": {}}
    """
    if not org_id:
        print("❌ org_id 不能为空")
        return {}

    display_name = company_name or f"org_id={org_id}"
    print(f"\n查询企业详情: {display_name}")

    session = ensure_session()
    if session is None:
        return {}

    result = {}
    xsrf_token = session.cookies.get("XSRF-TOKEN", "")

    # 1. 基本信息查询
    print("  [1/4] 查询基本信息...")
    try:
        url = f"{ZSK_BASE}/open/enterprise_info_api/v3/find_company_basic_info?orgid={org_id}"
        headers = {}
        if xsrf_token:
            headers["X-Xsrf-Token"] = xsrf_token
        resp = session.get(url, headers=headers, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            if data.get("status_code") == 0:
                result["basic_info"] = data.get("data", {})
                print(f"    ✓ 基本信息获取成功")
            else:
                print(f"    ⚠ 基本信息查询失败: {data.get('status_msg')}")
                result["basic_info"] = {}
        else:
            print(f"    ⚠ 基本信息查询失败: HTTP {resp.status_code}")
            result["basic_info"] = {}
    except Exception as e:
        print(f"    ⚠ 基本信息查询异常: {e}")
        result["basic_info"] = {}

    # 2. 股东信息查询
    print("  [2/4] 查询股东信息...")
    try:
        url = f"{ZSK_BASE}/open/enterprise_info_api/v4/stock_holder_newest?orgid={org_id}&page=1&pagesize=50&pageSize=50&page_size=50"
        headers = {}
        if xsrf_token:
            headers["X-Xsrf-Token"] = xsrf_token
        resp = session.get(url, headers=headers, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            if data.get("status_code") == 0:
                result["shareholders"] = data.get("data", {})
                print(f"    ✓ 股东信息获取成功")
            else:
                print(f"    ⚠ 股东信息查询失败: {data.get('status_msg')}")
                result["shareholders"] = {}
        else:
            print(f"    ⚠ 股东信息查询失败: HTTP {resp.status_code}")
            result["shareholders"] = {}
    except Exception as e:
        print(f"    ⚠ 股东信息查询异常: {e}")
        result["shareholders"] = {}

    # 3. 主要人员信息查询
    print("  [3/4] 查询主要人员...")
    try:
        # 先尝试 main_person_list 接口
        url = f"{ZSK_BASE}/open/enterprise_info_api/v3/main_person_list?orgid={org_id}&page=1&pageSize=50"
        headers = {}
        if xsrf_token:
            headers["X-Xsrf-Token"] = xsrf_token
        resp = session.get(url, headers=headers, timeout=30)

        if resp.status_code == 200:
            data = resp.json()
            if data.get("status_code") == 0:
                result["main_persons"] = data.get("data", {})
                print(f"    ✓ 主要人员获取成功")
            else:
                # 如果失败，尝试 tab 接口
                print(f"    ⚠ main_person_list 失败，尝试 tab 接口...")
                url = f"{ZSK_BASE}/open/enterprise_info_api/v1/main_person_tab?orgid={org_id}"
                resp = session.get(url, headers=headers, timeout=30)
                if resp.status_code == 200:
                    data = resp.json()
                    if data.get("status_code") == 0:
                        result["main_persons"] = data.get("data", {})
                        print(f"    ✓ 主要人员获取成功（tab 接口）")
                    else:
                        print(f"    ⚠ 主要人员查询失败: {data.get('status_msg')}")
                        result["main_persons"] = {}
                else:
                    print(f"    ⚠ 主要人员查询失败: HTTP {resp.status_code}")
                    result["main_persons"] = {}
        else:
            print(f"    ⚠ 主要人员查询失败: HTTP {resp.status_code}")
            result["main_persons"] = {}
    except Exception as e:
        print(f"    ⚠ 主要人员查询异常: {e}")
        result["main_persons"] = {}

    # 4. 股权结构图（投资 + 股东）
    print("  [4/4] 查询股权结构...")
    try:
        url = f"{ZSK_BASE}/open/enterprise_info_api/v3/atlas/enterprise_equity?orgid={org_id}"
        headers = {}
        if xsrf_token:
            headers["X-Xsrf-Token"] = xsrf_token
        resp = session.get(url, headers=headers, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            if data.get("status_code") == 0:
                equity_data = data.get("data", {})
                # 提取投资（invests）和股东（holders）
                result["equity"] = {
                    "invests": equity_data.get("invests", {}),
                    "holders": equity_data.get("holders", {}),
                }
                print(f"    ✓ 股权结构获取成功")
            else:
                print(f"    ⚠ 股权结构查询失败: {data.get('status_msg')}")
                result["equity"] = {"invests": {}, "holders": {}}
        else:
            print(f"    ⚠ 股权结构查询失败: HTTP {resp.status_code}")
            result["equity"] = {"invests": {}, "holders": {}}
    except Exception as e:
        print(f"    ⚠ 股权结构查询异常: {e}")
        result["equity"] = {"invests": {}, "holders": {}}

    print(f"\n✓ 企业详情查询完成")
    return result


def format_output(results):
    """格式化输出结果"""
    if not results:
        print("未找到匹配的企业信息")
        return

    print(f"\n{'='*80}")
    print(f"查询结果: 共 {len(results)} 条")
    print(f"{'='*80}\n")

    for i, company in enumerate(results, 1):
        print(f"[{i}] {company.get('name', 'N/A')}")
        print(f"    状态: {company.get('state', 'N/A')}")
        print(f"    法定代表人: {company.get('legal_person', 'N/A')}")
        print(f"    注册资本: {company.get('capital', 'N/A')}")
        print(f"    成立日期: {company.get('establish_time', 'N/A')}")
        print(f"    地址: {company.get('address', 'N/A')}")
        print(f"    电话: {', '.join(company.get('phone_num', []))}")
        print(f"    邮箱: {', '.join(company.get('e_mail', []))}")
        print(f"    官网: {', '.join(company.get('host', []))}")
        print()


def save_to_excel(results, output_path, all_fields=False, export_by_dimension=False, select_dimensions=None):
    """保存结果到 Excel

    Args:
        results: 查询结果列表
        output_path: 输出文件路径
        all_fields: 是否导出全部字段（默认只导出基础 10 个字段）
        export_by_dimension: 是否按维度分类导出多个 Excel 文件并打包成 ZIP（模拟注协系统导出）
        select_dimensions: 选择导出的维度列表（如 ['基本信息', '经营状况', '经营风险']），None 表示全部导出
    """
    # 基础字段（默认）
    basic_headers = ["企业名称", "状态", "法定代表人", "注册资本", "成立日期", "地址", "电话", "邮箱", "官网", "统一社会信用代码"]

    # 完整维度映射（5 个一级维度，30+ 个二级维度）
    full_dimensions = {
        # 基本信息
        "最新公示股东": {"dimension_code": "S0000103", "fields": ["name", "state", "legal_person", "capital", "establish_time", "unified_social_credit_code", "address", "phone_num", "e_mail", "host"]},
        "对对外投资（新）": {"dimension_code": "S0000104", "fields": []},
        "发票信息": {"dimension_code": "S0000107", "fields": []},
        "参控股企业": {"dimension_code": "S0000105", "fields": []},
        "分支机构": {"dimension_code": "S0000001", "fields": []},
        "对对外投资": {"dimension_code": "S0000015", "fields": []},
        "基础工商信息": {"dimension_code": "S0000002", "fields": ["name", "state", "legal_person", "capital", "establish_time", "unified_social_credit_code", "address", "phone_num", "e_mail", "host"]},
        "法定代表人变更": {"dimension_code": "S0000019", "fields": []},
        "最终受益人": {"dimension_code": "S0000020", "fields": []},
        "股东信息": {"dimension_code": "S0000006", "fields": []},
        "年报社保信息": {"dimension_code": "S0000034", "fields": []},
        "实际控制人": {"dimension_code": "S0000032", "fields": []},
        "变更记录": {"dimension_code": "S0000016", "fields": []},
        "主要人员（高管）": {"dimension_code": "S0000037", "fields": []},
        
        # 经营情况
        "商标": {"dimension_code": "S0000041", "fields": []},
        "税务相关信息": {"dimension_code": "S0000053", "fields": []},
        "微信公众号": {"dimension_code": "S0000101", "fields": []},
        "行业地区代码": {"dimension_code": "S0000054", "fields": []},
        "属地信息": {"dimension_code": "S0000116", "fields": []},
        "招聘信息": {"dimension_code": "S0000111", "fields": []},
        "进出口信用": {"dimension_code": "S0000112", "fields": []},
        "金融资质": {"dimension_code": "S0000114", "fields": []},
        "财税软件适用": {"dimension_code": "S0000110", "fields": []},
        "税控软件": {"dimension_code": "S0000027", "fields": []},
        "纳税人信息": {"dimension_code": "S0000035", "fields": []},
        "抽逃检查": {"dimension_code": "S0000113", "fields": []},
        "行政许可-工商": {"dimension_code": "S0000109", "fields": []},
        "建筑资质": {"dimension_code": "S0000120", "fields": []},
        "客户": {"dimension_code": "S0000119", "fields": []},
        "供应商": {"dimension_code": "S0000118", "fields": []},
        "招投标搜索": {"dimension_code": "S0000121", "fields": []},
        
        # 企业发展
        "债券信息": {"dimension_code": "S0000126", "fields": []},
        "银行授信额度": {"dimension_code": "S0000127", "fields": []},
        "融资历史": {"dimension_code": "S0000122", "fields": []},
        "理财-产品信息": {"dimension_code": "S0000128", "fields": []},
        "核心团队": {"dimension_code": "S0000123", "fields": []},
        "企业业务": {"dimension_code": "S0000124", "fields": []},
        "产品信息": {"dimension_code": "S0000125", "fields": []},
        
        # 经营风险
        "对对外担保": {"dimension_code": "S0000024", "fields": []},
        "动产抵押": {"dimension_code": "S0000012", "fields": []},
        "股权质押": {"dimension_code": "S0000013", "fields": []},
        "欠税公告": {"dimension_code": "S0000018", "fields": []},
        "经营异常": {"dimension_code": "S0000022", "fields": []},
        "知识产权出质": {"dimension_code": "S0000023", "fields": []},
        "股权出质": {"dimension_code": "S0000028", "fields": []},
        "行政处罚（工商）": {"dimension_code": "S0000029", "fields": []},
        "行政处罚（信用中国）": {"dimension_code": "S0000031", "fields": []},
        "清算信息": {"dimension_code": "S0000030", "fields": []},
        "严重违法": {"dimension_code": "S0000043", "fields": []},
        "注销信息": {"dimension_code": "S0000106", "fields": []},
        "行政处罚融合": {"dimension_code": "S0000036", "fields": []},
        
        # 诉讼风险
        "裁判文书": {"dimension_code": "S0000004", "fields": []},
        "立案信息": {"dimension_code": "S0000005", "fields": []},
        "开庭公告": {"dimension_code": "S0000007", "fields": []},
        "法院公告": {"dimension_code": "S0000009", "fields": []},
        "法院协助": {"dimension_code": "S0000011", "fields": []},
        "被执行人信息": {"dimension_code": "S0000008", "fields": []},
        "失信信息": {"dimension_code": "S0000010", "fields": []},
        "民间借贷纠纷": {"dimension_code": "S0000021", "fields": []},
        "涉诉案件": {"dimension_code": "S0000005", "fields": []},
    }

    # 字段映射（用于基础工商信息）
    field_map = {
        "企业名称": "name",
        "状态": "state",
        "法定代表人": "legal_person",
        "注册资本": "capital",
        "成立日期": "establish_time",
        "统一社会信用代码": "unified_social_credit_code",
        "地址": "address",
        "电话": "phone_num",
        "邮箱": "e_mail",
        "官网": "host",
    }

    if export_by_dimension:
        # 按维度导出多个 Excel 文件并打包成 ZIP
        import zipfile
        from datetime import datetime

        # 创建临时文件夹
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        temp_dir = Path(output_path).parent / f"批量查询_{timestamp}"
        temp_dir.mkdir(parents=True, exist_ok=True)

        # 确定要导出的维度
        if select_dimensions:
            selected_dims = {k: v for k, v in full_dimensions.items() if k in select_dimensions}
        else:
            selected_dims = full_dimensions

        # 为每个有数据的维度创建一个 Excel 文件
        exported_count = 0
        for dimension_name, dim_info in selected_dims.items():
            fields = dim_info.get("fields", [])
            if not fields or not results:
                continue

            # 检查数据是否存在
            has_data = False
            sample = results[0]
            for field_key in fields:
                if sample.get(field_key) is not None:
                    has_data = True
                    break

            if not has_data:
                # 即使没有数据，也创建空 Excel
                pass

            # 创建 Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = dimension_name

            # 写入表头
            headers = []
            for field_key in fields:
                header = field_map.get(field_key, field_key)
                headers.append(header)
            ws.append(headers)

            # 写入数据
            for company in results:
                row = []
                for field_key in fields:
                    value = company.get(field_key, "")
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value if v is not None and v != "")
                    elif isinstance(value, dict) and not value:
                        value = ""
                    elif value is True:
                        value = "是"
                    elif value is False:
                        value = "否"
                    elif value is None:
                        value = ""
                    row.append(value)
                ws.append(row)

            # 调整列宽
            for idx, _ in enumerate(headers, 1):
                column_letter = openpyxl.utils.get_column_letter(idx)
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), min_col=idx, max_col=idx):
                    cell = row[0]
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(50, max(10, max_length + 2))
                ws.column_dimensions[column_letter].width = adjusted_width

            # 保存 Excel 文件
            excel_path = temp_dir / f"{dimension_name}.xlsx"
            wb.save(excel_path)
            exported_count += 1
            print(f"  已保存: {dimension_name}.xlsx ({len(headers)} 个字段)")

        # 打包成 ZIP
        zip_path = Path(output_path)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for file_path in temp_dir.glob("*.xlsx"):
                arcname = file_path.name
                zipf.write(file_path, arcname)

        # 打包完成后再清理临时文件
        import shutil
        excel_files = list(temp_dir.glob('*.xlsx'))
        shutil.rmtree(temp_dir)

        print(f"\n已打包: {zip_path} (包含 {exported_count} 个 Excel 文件)")

    else:
        # 单个 Excel 文件导出
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工商信息"

        if all_fields:
            # 从已有数据中动态提取所有可用字段
            headers = []
            if results:
                sample = results[0]
                for field_key, value in sample.items():
                    if value is not None and value != "":
                        header = field_map.get(field_key, field_key)
                        if header not in headers:
                            headers.append(header)
            if not headers:
                headers = basic_headers
        else:
            headers = basic_headers

        # 写入表头
        ws.append(headers)

        # 写入数据
        for company in results:
            row = []
            for header in headers:
                # 反向查找字段 key
                field_key = None
                for k, v in field_map.items():
                    if v == header:
                        field_key = k
                        break
                if not field_key:
                    field_key = header

                value = company.get(field_key, "")
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value if v is not None and v != "")
                elif isinstance(value, dict) and not value:
                    value = ""
                elif value is True:
                    value = "是"
                elif value is False:
                    value = "否"
                elif value is None:
                    value = ""
                row.append(value)
            ws.append(row)

        # 自动调整列宽
        for idx, _ in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(idx)
            max_length = 0
            for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), min_col=idx, max_col=idx):
                cell = row[0]
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(50, max(10, max_length + 2))
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        print(f"结果已保存到: {output_path} (共 {len(headers)} 个字段, {len(results)} 条记录)")


# ============================================================================
# 完整导出（纯 API 方案，无需浏览器）
# 核心发现: POST /batch_search/enter_search_out_type 触发导出
# ============================================================================

def _verify_companies(input_names, extract_dir):
    """验证下载的 ZIP 中公司名称与输入是否一致"""
    import re
    base_file = Path(extract_dir) / "基础工商信息.xlsx"
    if not base_file.exists():
        # 尝试其他文件
        for f in Path(extract_dir).glob("*.xlsx"):
            if "基础" in f.name or "工商" in f.name:
                base_file = f
                break
        if not base_file.exists():
            print("  ⚠ 未找到基础工商信息文件，跳过验证")
            return

    wb = openpyxl.load_workbook(str(base_file))
    ws = wb.active
    file_companies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            file_companies.append(str(row[0]).strip())

    # 标准化比较：全角→半角、去空格
    def normalize(s):
        return re.sub(r'\s+', '', s.replace('（', '(').replace('）', ')'))

    input_norm = {normalize(n) for n in input_names}
    file_norm = {normalize(n) for n in file_companies}

    matched = input_norm & file_norm
    missed = input_norm - file_norm
    extra = file_norm - input_norm

    print()
    print(f"  📋 公司验证: 输入 {len(input_names)} 家 → 文件 {len(file_companies)} 家")

    if matched:
        # 打印精确匹配的
        matched_original = [n for n in input_names if normalize(n) in matched]
        print(f"  ✅ 匹配成功 ({len(matched)} 家):")
        for n in matched_original:
            print(f"     {n}")

    if missed:
        # 模糊匹配：检查是否是名称相似但注册名不同
        missed_original = [n for n in input_names if normalize(n) in missed]
        print(f"  ❌ 未匹配 ({len(missed)} 家):")
        for n in missed_original:
            # 尝试模糊匹配
            n_key = normalize(n)[:4]
            fuzzy = [fc for fc in file_companies if n_key in normalize(fc)]
            if fuzzy:
                print(f"     {n}  →  可能是: {fuzzy[0]}")
            else:
                print(f"     {n}  →  系统中未找到")

    if extra:
        extra_orig = [fc for fc in file_companies if normalize(fc) in extra]
        if extra_orig:
            print(f"  ℹ️ 文件中额外的公司 ({len(extra_orig)} 家):")
            for n in extra_orig[:3]:
                print(f"     {n}")
            if len(extra_orig) > 3:
                print(f"     ... 及其他 {len(extra_orig) - 3} 家")

    if len(matched) == len(input_names):
        print(f"  🎉 全部 {len(input_names)} 家公司匹配成功！")


def _get_all_dimensions(session, base_url):
    """获取所有导出维度代码"""
    resp = session.get(f"{base_url}/open/industry_chain_api/v1/search/export/get_dimension_class",
                       params={"internal": "false"})
    if resp.status_code != 200:
        return []
    data = resp.json().get("data", [])
    codes = []
    for dim in data:
        for child in dim.get("children", []):
            code = child.get("second_dimension_code", "")
            if code:
                codes.append(code)
    return codes


def browser_export(company_names, output_path):
    """纯 API 完整导出 26+ 个文件的 ZIP 包

    流程:
    1. 检查 cookie / 登录
    2. API 上传企业名单 (upload_local)
    3. API 触发查询 (enter_search)
    4. API 获取维度列表 (get_dimension_class)
    5. API 触发导出 (enter_search_out_type)
    6. API 轮询下载中心等待新任务
    7. API 下载 ZIP 文件

    Args:
        company_names: 企业名称列表
        output_path: 输出 ZIP 文件路径
    """
    import time as _time

    print()
    print("=" * 60)
    print("  注协系统完整导出（纯 API，无需浏览器）")
    print("=" * 60)
    print()

    base_url = ZSK_BASE

    # ------------------------------------------------------------------
    # 1. 获取有效 session
    # ------------------------------------------------------------------
    session = ensure_session()
    if session is None:
        return

    # ------------------------------------------------------------------
    # 2. 上传企业名单
    # ------------------------------------------------------------------
    print("[1/6] 上传企业名单...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "企业名称"
    for i, name in enumerate(company_names, 2):
        ws.cell(row=i, column=1, value=name)
    tmp_fd, temp_xlsx = tempfile.mkstemp(suffix=".xlsx", prefix="cicpa_export_")
    os.close(tmp_fd)
    wb.save(temp_xlsx)

    with open(temp_xlsx, "rb") as f:
        files = {"file": ("companies.xlsx", f,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        resp = session.post(f"{base_url}/open/industry_chain_api/v1/batch_search/upload_local",
                            files=files)

    # 删除临时文件
    try:
        os.remove(temp_xlsx)
    except Exception:
        pass

    if resp.status_code != 200:
        print(f"  ❌ 上传失败: HTTP {resp.status_code}")
        print(f"  响应: {resp.text[:200]}")
        return

    upload_data = resp.json()
    if upload_data.get("status_code") != 0:
        print(f"  ❌ 上传失败: {upload_data.get('status_msg')}")
        return

    batch_id = upload_data["data"]["batch_id"]
    hit_num = upload_data["data"].get("hit_num", 0)
    print(f"  ✓ 上传成功，batch_id = {batch_id}")
    print(f"    匹配 {hit_num} 家 / 共 {len(company_names)} 家")

    # ------------------------------------------------------------------
    # 3. 触发查询
    # ------------------------------------------------------------------
    print()
    print("[2/6] 触发批量查询...")
    resp = session.post(f"{base_url}/open/industry_chain_api/v1/batch_search/enter_search",
                        json={"batch_id": batch_id, "page": 1, "page_size": 10, "type": "credit"})
    if resp.status_code != 200:
        print(f"  ❌ 查询失败: HTTP {resp.status_code}")
        return

    search_data = resp.json()
    result_count = len(search_data.get("data", {}).get("list", []))
    print(f"  ✓ 查询完成，返回 {result_count} 条结果")

    # ------------------------------------------------------------------
    # 4. 获取维度 & 触发导出
    # ------------------------------------------------------------------
    print()
    print("[3/6] 获取导出维度...")
    dimensions = _get_all_dimensions(session, base_url)
    if not dimensions:
        print("  ❌ 获取维度列表失败")
        return
    print(f"  ✓ 获取到 {len(dimensions)} 个维度")

    print()
    print("[4/6] 触发导出 (enter_search_out_type)...")
    resp = session.post(f"{base_url}/open/industry_chain_api/v1/batch_search/enter_search_out_type",
                        json={"batch_id": batch_id, "dimensions": dimensions})
    if resp.status_code != 200:
        print(f"  ❌ 导出触发失败: HTTP {resp.status_code}")
        print(f"  响应: {resp.text[:200]}")
        return

    export_data = resp.json()
    print(f"  ✓ 导出已触发")
    print(f"    响应: {json.dumps(export_data, ensure_ascii=False)[:200]}")

    # ------------------------------------------------------------------
    # 5. 记录当前下载中心状态
    # ------------------------------------------------------------------
    print()
    print("[5/6] 轮询下载中心（等待导出任务生成）...")

    # 记录导出触发时刻（用于检测新任务）
    export_trigger_time = _time.strftime("%Y-%m-%d %H:%M", _time.localtime())
    print(f"  导出触发时间: {export_trigger_time}")

    # 轮询等待新任务（下载中心第一个任务，状态从"刷新"变为"下载"）
    max_wait = 180  # 3 分钟（大数据量可能需要更久）
    poll_interval = 5
    new_task = None

    for elapsed in range(0, max_wait, poll_interval):
        _time.sleep(poll_interval)
        resp = session.get(f"{base_url}/open/industry_chain_api/v1/download/task_list",
                           params={"page": 1, "pageSize": 5, "fresh": str(int(_time.time() * 1000))})
        if resp.status_code != 200:
            print(f"  ⚠ 查询失败 (HTTP {resp.status_code})，继续重试...")
            continue

        current_batch = [t for t in resp.json().get("data", {}).get("list", [])
                         if t.get("type") == "批量查询"]
        # 按日期降序排列，最新的在前面
        current_batch.sort(key=lambda x: x.get("date", ""), reverse=True)

        if not current_batch:
            print(f"  等待中... ({elapsed + poll_interval}s / {max_wait}s)")
            continue

        latest = current_batch[0]
        latest_date = latest.get("date", "")

        # 找日期 >= 触发时间 且 status=1（可下载）的任务
        if latest_date >= export_trigger_time and latest.get("status") == 1:
            new_task = latest
            print(f"  ✓ 发现新任务: {latest.get('name')} (status={latest.get('status')})")
            print(f"    日期: {latest_date}")
            break
        elif latest_date >= export_trigger_time:
            print(f"  任务处理中... ({latest.get('name')} status={latest.get('status')}) ({elapsed + poll_interval}s)")
        else:
            print(f"  等待新任务生成... ({elapsed + poll_interval}s / {max_wait}s)")

    # 超时后兜底：取最新的可下载任务（不管日期）
    if not new_task:
        print()
        print("  ⚠ 轮询超时，尝试取最新的可下载任务...")
        resp = session.get(f"{base_url}/open/industry_chain_api/v1/download/task_list",
                           params={"page": 1, "pageSize": 5, "fresh": str(int(_time.time() * 1000))})
        if resp.status_code == 200:
            all_batch = [t for t in resp.json().get("data", {}).get("list", [])
                         if t.get("type") == "批量查询" and t.get("status") == 1]
            all_batch.sort(key=lambda x: x.get("date", ""), reverse=True)
            if all_batch:
                new_task = all_batch[0]
                print(f"  取最新可下载任务: {new_task.get('name')} ({new_task.get('date')})")
                print(f"  ⚠ 请手动核实此任务是否包含本次查询的企业")

    if not new_task:
        print()
        print("  ❌ 轮询超时，未发现新任务")
        print("  可能原因:")
        print("    1. 导出任务处理时间较长，请稍后手动检查下载中心")
        print("    2. 系统未生成导出任务（可能需要更长等待）")
        print()
        print("  手动下载方法:")
        print("    1. 打开 https://zsk-cmis.cicpa.org.cn/companylibrarynew/account.html#/donwnloadcenter")
        print("    2. 找到最新的「批量查询」任务")
        print("    3. 点击下载")
        return

    # 等待任务完成
    if new_task.get("status") != 1:
        task_id = new_task.get("task_id")
        for elapsed in range(0, max_wait, poll_interval):
            _time.sleep(poll_interval)
            resp = session.get(f"{base_url}/open/industry_chain_api/v1/download/task_list",
                               params={"page": 1, "pageSize": 10, "fresh": str(int(_time.time() * 1000))})
            if resp.status_code != 200:
                continue
            for t in resp.json().get("data", {}).get("list", []):
                if t.get("task_id") == task_id:
                    new_task = t
                    if t.get("status") == 1:
                        print(f"  ✓ 任务完成")
                        break
            if new_task.get("status") == 1:
                break

    if new_task.get("status") != 1:
        print(f"  ❌ 任务未完成 (status={new_task.get('status')})")
        return

    # ------------------------------------------------------------------
    # 6. 下载 ZIP
    # ------------------------------------------------------------------
    print()
    print("[6/6] 下载 ZIP 文件...")
    download_url = new_task.get("url", "")
    if not download_url:
        print("  ❌ 下载链接为空")
        return

    full_url = download_url if download_url.startswith("http") else f"{base_url}/{download_url}"
    print(f"  下载: {full_url}")

    resp = session.get(full_url, stream=True)
    if resp.status_code != 200:
        print(f"  ❌ 下载失败: HTTP {resp.status_code}")
        return

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    total_size = 0
    with open(output_path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=8192):
            f.write(chunk)
            total_size += len(chunk)

    size_mb = total_size / 1024 / 1024
    print(f"  ✓ 下载完成: {output_path} ({size_mb:.2f} MB)")

    # 验证 + 解压
    import zipfile
    if zipfile.is_zipfile(output_path):
        with zipfile.ZipFile(output_path, "r") as zf:
            names = zf.namelist()
            print(f"  ✓ ZIP 包含 {len(names)} 个文件:")
            for n in names[:5]:
                print(f"    - {n}")
            if len(names) > 5:
                print(f"    ... 及其他 {len(names) - 5} 个")

            # 解压目录：去掉 .zip 后缀，追加 _files 避免与 ZIP 文件同名冲突
            extract_dir = str(Path(output_path).with_suffix("")) + "_files"
            Path(extract_dir).mkdir(parents=True, exist_ok=True)

            for info in zf.infolist():
                # 修复中文文件名编码（ZIP 内 GBK 被误读为 cp437）
                try:
                    filename = info.filename.encode('cp437').decode('utf-8')
                except (UnicodeDecodeError, UnicodeEncodeError):
                    try:
                        filename = info.filename.encode('cp437').decode('gbk')
                    except Exception:
                        filename = info.filename

                target = os.path.join(extract_dir, filename)
                # 防止路径遍历
                real_target = os.path.realpath(target)
                real_extract = os.path.realpath(extract_dir)
                if not real_target.startswith(real_extract + os.sep) and real_target != real_extract:
                    print(f"  ⚠ 跳过可疑文件: {filename}")
                    continue

                os.makedirs(os.path.dirname(target), exist_ok=True)
                with zf.open(info) as src, open(target, 'wb') as dst:
                    dst.write(src.read())

            print(f"  ✓ 已解压到: {extract_dir}")

            # 验证公司名称一致性
            _verify_companies(company_names, extract_dir)
            return extract_dir
    else:
        print(f"  ⚠ 文件不是有效的 ZIP")

    print()
    print("=" * 60)
    print(f"  ✓ 导出完成: {output_path}")
    print("=" * 60)
    return extract_dir


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="注协系统工商信息批量查询（无需配置文件，浏览器登录即可）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 首次使用：登录（打开浏览器，手动登录后自动保存 cookies）
  python cicpa_query.py --login

  # 基础查询
  python cicpa_query.py -n "华为技术有限公司" "腾讯科技（深圳）有限公司"

  # 从文件读取企业名单
  python cicpa_query.py -f companies.xlsx

  # 完整导出（53 个 Excel → ZIP）
  python cicpa_query.py -f companies.xlsx --browser-export

  # 重新登录
  python cicpa_query.py --login
""")

    parser.add_argument("-f", "--file", help="包含企业名称的 Excel/TXT 文件")
    parser.add_argument("-n", "--names", nargs="+", help="直接指定企业名称")
    parser.add_argument("-o", "--output", help="输出文件路径")
    parser.add_argument("--login", action="store_true",
                        help="独立登录：打开浏览器登录并保存 cookies（不查询）")
    parser.add_argument("--save-cookies", metavar="JSON",
                        help="保存 AI 抓取的 cookies（JSON 格式），如：'--save-cookies \"{\\\"XSRF-TOKEN\\\":\\\"xxx\\\"}\"'")
    parser.add_argument("--check-cookies", action="store_true",
                        help="检查当前 cookies 是否有效（不查询）")
    parser.add_argument("--all-fields", action="store_true", help="导出全部字段（默认 10 个）")
    parser.add_argument("--by-dimension", action="store_true", help="按维度分类导出 → ZIP")
    parser.add_argument("--select-dimensions", nargs="+", help="选择导出维度")
    parser.add_argument("--browser-export", action="store_true", help="完整导出（61 维度 → ZIP）")
    parser.add_argument("--detail", metavar="ID_OR_NAME",
                        help="查询单家企业详情（输入 org_id 或企业名称）")
    parser.add_argument("--discover-subsidiaries", metavar="COMPANY",
                        help="发现企业子公司（输入企业名称，默认持股>=50%%）")
    parser.add_argument("--subsidiary-threshold", type=float, default=50.0,
                        help="子公司持股比例阈值（默认 50%%，配合 --discover-subsidiaries 使用）")
    parser.add_argument("--max-depth", type=int, default=5,
                        help="子公司穿透最大深度（默认5层）")
    args = parser.parse_args()

    # ── 保存 AI 抓取的 cookies ──
    if args.save_cookies:
        try:
            cookies_dict = json.loads(args.save_cookies)
        except json.JSONDecodeError as e:
            print(f"❌ JSON 解析失败: {e}")
            print("  格式: python cicpa_query.py --save-cookies '{\"XSRF-TOKEN\":\"xxx\",...}'")
            sys.exit(1)
        ok = save_cookies_from_dict(cookies_dict)
        sys.exit(0 if ok else 1)

    # ── 检查 cookies ──
    if args.check_cookies:
        cookie_data = _load_cookies()
        if not cookie_data or "cookies" not in cookie_data:
            print("❌ 无 cookies 文件，需要先登录")
            sys.exit(1)
        saved_at = cookie_data.get("saved_at", 0)
        age_hours = (time.time() - saved_at) / 3600
        cookies = cookie_data["cookies"]
        print(f"Cookie 文件: {COOKIE_FILE}")
        print(f"保存时间: {age_hours:.1f} 小时前")
        print(f"Cookie 数量: {len(cookies)} 个")
        print(f"包含: {', '.join(sorted(cookies.keys()))}")
        session = _build_session(cookies)
        if _check_cookie_valid(session):
            remaining = max(0, 24 - int(age_hours))
            print(f"✅ Cookies 有效（剩余约 {remaining} 小时）")
            sys.exit(0)
        else:
            print("❌ Cookies 无效（API 验证失败），需要重新登录")
            sys.exit(1)

    # ── 独立登录模式 ──
    if args.login and not args.names and not args.file and not args.detail and not args.discover_subsidiaries:
        ok = do_login()
        sys.exit(0 if ok else 1)

    # ── 企业详情查询模式 ──
    if args.detail:
        query = args.detail.strip()
        org_id = None
        company_name = None

        # 判断是 org_id 还是企业名称
        if query.startswith("T") and len(query) > 8:
            org_id = query
            print(f"检测到 org_id: {org_id}")
        else:
            # 先搜索获取 org_id
            print(f"搜索企业: {query}")
            results = search_company(query, page_size=1)
            if results:
                org_id = results[0].get("org_id", "")
                company_name = results[0].get("name", "")
                print(f"找到企业: {company_name} (org_id={org_id})")
            else:
                print(f"❌ 未找到企业: {query}")
                sys.exit(1)

        # 查询详情
        detail = get_company_detail(org_id, company_name)
        if detail:
            # 格式化输出
            print(f"\n{'='*80}")
            print(f"企业详情: {company_name or org_id}")
            print(f"{'='*80}\n")

            # 基本信息
            basic = detail.get("basic_info", {})
            if basic:
                print("【基本信息】")
                for key, value in basic.items():
                    if value is not None and value != "":
                        print(f"  {key}: {value}")
                print()

            # 股东信息
            shareholders = detail.get("shareholders", {})
            if shareholders:
                print("【股东信息】")
                print(f"  数据结构: {list(shareholders.keys())}")
                # 尝试打印股东列表
                if isinstance(shareholders, dict):
                    for key in shareholders.keys():
                        print(f"  - {key}")
                print()

            # 主要人员
            persons = detail.get("main_persons", {})
            if persons:
                print("【主要人员】")
                print(f"  数据结构: {list(persons.keys())}")
                print()

            # 股权结构
            equity = detail.get("equity", {})
            if equity:
                print("【股权结构】")
                invests = equity.get("invests", {})
                holders = equity.get("holders", {})
                print(f"  对外投资: {list(invests.keys())}")
                print(f"  股东结构: {list(holders.keys())}")

                # 打印投资子公司
                if isinstance(invests, dict):
                    children = invests.get("children", [])
                    if children:
                        print(f"\n  对外投资 ({len(children)} 家):")
                        for child in children[:10]:
                            name = child.get("name", "").strip()
                            ratio = child.get("czbl", "")
                            if name:
                                print(f"    - {name} {ratio}")

                # 打印股东
                if isinstance(holders, dict):
                    children = holders.get("children", [])
                    if children:
                        print(f"\n  股东 ({len(children)} 位):")
                        for holder in children[:10]:
                            name = holder.get("name", "").strip()
                            ratio = holder.get("czbl", "")
                            if name:
                                print(f"    - {name} {ratio}")
                print()

            # 保存到 JSON
            if args.output:
                output_path = args.output
            else:
                output_path = str(BASE_DIR / "output" / f"{company_name or org_id}_详情.json")
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(detail, f, ensure_ascii=False, indent=2)
            print(f"\n详情已保存到: {output_path}")
        else:
            print("❌ 查询失败")
            sys.exit(1)

        sys.exit(0)

    # ── 发现子公司模式 ──
    if args.discover_subsidiaries:
        target = args.discover_subsidiaries.strip()
        threshold = args.subsidiary_threshold
        max_depth = getattr(args, 'max_depth', 5)
        print(f"发现子公司: {target}（持股 >= {threshold}%，最大深度{max_depth}层）")
        subsidiaries = discover_subsidiaries(target, threshold, max_depth)
        if subsidiaries:
            print(f"\n子公司名称列表:")
            for sub_name, sub_info in subsidiaries.items():
                print(f"  {sub_name} (持股{sub_info['ratio']}%, 深度{sub_info['depth']}, 母公司:{sub_info['parent']})")
        else:
            print(f"\n未发现持股 >= {threshold}% 的子公司")
        sys.exit(0)

    # ── 查询模式 ──
    company_names = []
    if args.names:
        company_names = args.names
    elif args.file:
        filepath = Path(args.file)
        if filepath.suffix in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    company_names.append(str(row[0]).strip())
        else:
            with open(filepath, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        company_names.append(line)
    else:
        # 从审定明细表读取
        print("未指定企业名称，尝试从测试数据读取...")
        test_data_dir = BASE_DIR / "2-测试数据"
        if test_data_dir.exists():
            for f in test_data_dir.glob("*.xlsx"):
                if "送审" in f.name or "审定" in f.name:
                    wb = openpyxl.load_workbook(f)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            company_names.append(str(row[0]).strip())
                    print(f"  从 {f.name} 读取 {len(company_names)} 家企业")
                    break

    if not company_names:
        print("错误: 未找到企业名称")
        print("用法:")
        print("  python cicpa_query.py --login          # 先登录")
        print("  python cicpa_query.py -n '华为技术有限公司'")
        print("  python cicpa_query.py -f companies.xlsx")
        sys.exit(1)

    print(f"待查询企业: {len(company_names)} 家")
    for name in company_names[:5]:
        print(f"  - {name}")
    if len(company_names) > 5:
        print(f"  ... 及其他 {len(company_names) - 5} 家")

    # 浏览器完整导出模式
    if args.browser_export:
        out = args.output or str(BASE_DIR / "output" / "完整维度导出.zip")
        browser_export(company_names, out)
        sys.exit(0)

    # 执行查询（--login + 企业名称 = 强制重新登录后查询）
    results = batch_query(company_names, force_login=args.login)

    # 输出
    format_output(results)

    # 保存
    output_path = args.output or str(BASE_DIR / "output" / "工商信息查询结果.xlsx")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    save_to_excel(results, output_path,
                  all_fields=args.all_fields,
                  export_by_dimension=args.by_dimension,
                  select_dimensions=args.select_dimensions)


if __name__ == "__main__":
    main()

