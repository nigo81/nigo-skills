#!/usr/bin/env python3
"""
中国法律法规统一查询工具
自动先查国家法律法规数据库(flk.npc.gov.cn)，查不到再查国家规章库(gov.cn)
"""

import argparse
import json
import os
import re
import sys
import time

# Add lib to path (lib is in skill root, one level up from scripts/)
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "lib"))
import flk_api
import gov_api


def unified_search(keyword: str, exact: bool = False, content_search: bool = False, size: int = 10) -> dict:
    """统一搜索：先查 flk，再查规章库"""
    results = {"flk": [], "gov": [], "keyword": keyword}

    # 1. flk
    try:
        resp = flk_api.search(keyword, exact=exact, content_search=content_search, size=size)
        results["flk"] = resp.get("rows", [])
        results["flk_total"] = resp.get("total", 0)
    except Exception as e:
        results["flk_error"] = str(e)

    # 2. gov (only title search, gov doesn't support content search)
    try:
        resp = gov_api.search(keyword, size=size)
        results["gov"] = resp.get("rows", [])
        results["gov_total"] = resp.get("total", 0)
    except Exception as e:
        results["gov_error"] = str(e)

    return results


def unified_new_since(date_str: str, law_type: str = None, all_types: bool = True) -> dict:
    """统一查询指定日期后新发布的法规（两库联查）"""
    flk_resp = flk_api.new_since_all(date_str, law_type=law_type, all_types=all_types)
    gov_resp = gov_api.new_since(date_str, reg_type=law_type)
    return {
        "since": date_str,
        "flk": flk_resp,
        "gov": gov_resp,
        "total": flk_resp.get("total", 0) + gov_resp.get("total", 0),
    }


def new_since_to_excel(date_str: str, output_file: str,
                       law_type: str = None, all_types: bool = True) -> str:
    """查询指定日期后新发布的法规并输出Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    print(f"查询 {date_str} 以来新发布的法规...", file=sys.stderr)
    data = unified_new_since(date_str, law_type=law_type, all_types=all_types)
    flk_rows = data["flk"].get("rows", [])
    gov_rows = data["gov"].get("rows", [])
    print(f"  flk: {len(flk_rows)} 条, 规章库: {len(gov_rows)} 条", file=sys.stderr)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{date_str}以来新发布法规"

    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    bfont = Font(name="微软雅黑", size=10)
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))
    sfills = {
        "有效": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "已修改": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "尚未生效": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    }
    SXX = {1: "已废止", 2: "已修改", 3: "有效", 4: "尚未生效"}

    headers = ["序号", "法规名称", "法规类别", "公布日期", "施行日期",
               "时效性", "制定机关", "数据来源", "原文链接"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.border = hfill, hfont, bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    idx = 0
    type_stats = {}
    for r in flk_rows:
        idx += 1
        row = idx + 1
        sxx = SXX.get(r.get("sxx"), r.get("sxx_text", ""))
        flxz = r.get("flxz", "未分类")
        type_stats[flxz] = type_stats.get(flxz, 0) + 1
        bbbs = r.get("bbbs", "")
        vals = [idx, r.get("title", ""), flxz, r.get("gbrq", ""), r.get("sxrq", ""),
                sxx, r.get("zdjgName", ""), "国家法律法规数据库",
                f"https://flk.npc.gov.cn/detail?id={bbbs}" if bbbs else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font, c.border = bfont, bdr
            c.alignment = Alignment(vertical="center", wrap_text=True)
        if sxx in sfills:
            ws.cell(row=row, column=6).fill = sfills[sxx]

    for r in gov_rows:
        idx += 1
        row = idx + 1
        typ = r.get("type", "未分类")
        type_stats[typ] = type_stats.get(typ, 0) + 1
        vals = [idx, r.get("title", ""), typ, r.get("date", ""), "",
                "有效", r.get("org", ""), "国家规章库", r.get("url", "")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font, c.border = bfont, bdr
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row, column=6).fill = sfills["有效"]

    for i, w in enumerate([6, 55, 15, 12, 12, 10, 25, 20, 50], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 统计sheet
    ws2 = wb.create_sheet("统计汇总")
    ws2.cell(row=1, column=1, value="类别").font = hfont
    ws2.cell(row=1, column=1).fill = hfill
    ws2.cell(row=1, column=2, value="数量").font = hfont
    ws2.cell(row=1, column=2).fill = hfill
    for i, (t, c) in enumerate(sorted(type_stats.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=i, column=1, value=t)
        ws2.cell(row=i, column=2, value=c)
    rr = len(type_stats) + 2
    ws2.cell(row=rr, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=rr, column=2, value=idx).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    wb.save(output_file)
    print(f"✅ 已保存: {output_file} (共 {idx} 条)", file=sys.stderr)
    return output_file


def batch_check_to_excel(input_excel: str, output_excel: str,
                         sheet_name: str = None, col: str = "B",
                         start_row: int = 2, delay: float = 0.5,
                         progress_file: str = None) -> str:
    """从Excel读取法规名称，批量检查有效性，输出结果Excel

    Args:
        input_excel: 输入Excel文件路径
        output_excel: 输出Excel文件路径
        sheet_name: 工作表名称（默认第一个）
        col: 法规名称所在列（默认B）
        start_row: 数据起始行（默认2，跳过表头）
        delay: 请求间隔秒数
        progress_file: 进度文件路径（支持断点续查）
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # 读取法规名称
    wb_in = openpyxl.load_workbook(input_excel, data_only=True)
    ws_in = wb_in[sheet_name] if sheet_name else wb_in.active
    col_idx = ord(col.upper()) - ord("A")
    laws = []
    for row in ws_in.iter_rows(min_row=start_row, max_row=ws_in.max_row, values_only=True):
        name = row[col_idx] if col_idx < len(row) else None
        if name and str(name).strip():
            laws.append(str(name).strip())
    wb_in.close()
    print(f"共 {len(laws)} 条法规需要查询", file=sys.stderr)

    # 加载/创建进度
    progress = {}
    if progress_file and os.path.exists(progress_file):
        with open(progress_file, "r") as f:
            progress = json.load(f)
    already = sum(1 for l in laws if l in progress)
    print(f"已完成: {already}, 待查询: {len(laws) - already}", file=sys.stderr)

    # 分批查询
    BATCH_SIZE = 10
    todo = [l for l in laws if l not in progress]
    for batch_idx in range(0, len(todo), BATCH_SIZE):
        batch = todo[batch_idx:batch_idx + BATCH_SIZE]
        batch_num = batch_idx // BATCH_SIZE + 1
        total_batches = (len(todo) + BATCH_SIZE - 1) // BATCH_SIZE
        print(f"  批次 {batch_num}/{total_batches}", file=sys.stderr)

        results = unified_batch_check(batch, delay=delay)
        for r in results:
            name = r.get("name", "")
            if not name:
                continue
            found = r.get("found", False)
            source = "国家规章库" if r.get("source") == "gov.cn" else "国家法律法规数据库" if found else ""
            info = {
                "status": r.get("status", "未找到"),
                "type": r.get("flxz", ""),
                "date": r.get("gbrq", ""),
                "org": r.get("org", ""),
                "source": source,
                "url": r.get("url", ""),
                "note": "",
            }
            matched = r.get("title", "")
            if matched and matched != name:
                info["note"] = f"匹配: {matched}"
            progress[name] = info

        if progress_file:
            with open(progress_file, "w") as f:
                json.dump(progress, f, ensure_ascii=False, indent=2)

        if batch_idx + BATCH_SIZE < len(todo):
            time.sleep(1)

    # 生成Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "法规有效性查询结果"

    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    bfont = Font(name="微软雅黑", size=10)
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))
    sfills = {
        "有效": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "有效（规章库）": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "已修改": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "已废止": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "尚未生效": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    }

    headers = ["序号", "法规名称", "查询状态", "法规类别", "公布日期",
               "时效性", "制定机关", "数据来源", "原文链接", "备注"]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_i, value=h)
        c.fill, c.font, c.border = hfill, hfont, bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    for idx, law_name in enumerate(laws, 1):
        row = idx + 1
        info = progress.get(law_name, {})
        status = info.get("status", "未查询")
        vals = [idx, law_name, status, info.get("type", ""), info.get("date", ""),
                status, info.get("org", ""), info.get("source", ""),
                info.get("url", ""), info.get("note", "")]
        for col_i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_i, value=v)
            c.font, c.border = bfont, bdr
            c.alignment = Alignment(vertical="center", wrap_text=True)
        fill = sfills.get(status)
        if fill:
            ws.cell(row=row, column=3).fill = fill
            ws.cell(row=row, column=6).fill = fill

    for i, w in enumerate([6, 50, 10, 15, 15, 10, 20, 20, 40, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 统计sheet
    ws2 = wb.create_sheet("统计汇总")
    ws2.cell(row=1, column=1, value="状态").font = hfont
    ws2.cell(row=1, column=1).fill = hfill
    ws2.cell(row=1, column=2, value="数量").font = hfont
    ws2.cell(row=1, column=2).fill = hfill
    ws2.cell(row=1, column=3, value="占比").font = hfont
    ws2.cell(row=1, column=3).fill = hfill
    stats = {}
    for law_name in laws:
        s = progress.get(law_name, {}).get("status", "未查询")
        stats[s] = stats.get(s, 0) + 1
    total = len(laws)
    for i, (s, c) in enumerate(sorted(stats.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=i, column=1, value=s)
        ws2.cell(row=i, column=2, value=c)
        ws2.cell(row=i, column=3, value=f"{c/total*100:.1f}%")
    ws2.cell(row=len(stats) + 2, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=len(stats) + 2, column=2, value=total).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10

    wb.save(output_excel)
    print(f"✅ 结果已保存: {output_excel}", file=sys.stderr)
    return output_excel


def batch_download_from_check(check_progress: dict, output_dir: str,
                              delay: float = 0.5,
                              dl_progress_file: str = None) -> dict:
    """从 batch_check 的进度数据批量下载法规全文

    Args:
        check_progress: batch_check 的进度字典 {法规名: {status, url, source, ...}}
        output_dir: 下载目录
        delay: 请求间隔
        dl_progress_file: 下载进度文件（支持断点续下）

    Returns:
        下载进度字典 {法规名: {ok: bool, file: str, error: str}}
    """
    os.makedirs(output_dir, exist_ok=True)

    # 加载下载进度
    dl_progress = {}
    if dl_progress_file and os.path.exists(dl_progress_file):
        with open(dl_progress_file, "r") as f:
            dl_progress = json.load(f)

    # 分类
    flk_items = []
    gov_items = []
    for name, info in check_progress.items():
        status = info.get("status", "")
        if "有效" not in status and status not in ("已修改", "已废止", "尚未生效"):
            continue
        url = info.get("url", "")
        source = info.get("source", "")
        if "flk.npc.gov.cn" in url:
            bbbs_id = url.split("id=")[1].split("&")[0] if "id=" in url else ""
            if bbbs_id:
                flk_items.append({"name": name, "bbbs_id": bbbs_id})
        elif source == "国家规章库" or "gov.cn" in url:
            gov_items.append({"name": name})

    total = len(flk_items) + len(gov_items)
    already = sum(1 for v in dl_progress.values() if v.get("ok"))
    print(f"总计 {total} 条待下载 (flk: {len(flk_items)}, 规章库: {len(gov_items)})", file=sys.stderr)
    print(f"已下载: {already}", file=sys.stderr)

    # 下载 flk
    print(f"=== 下载国家法律法规数据库 ({len(flk_items)} 条) ===", file=sys.stderr)
    for i, item in enumerate(flk_items):
        name = item["name"]
        if dl_progress.get(name, {}).get("ok"):
            continue
        try:
            filepath = flk_api.download_file(item["bbbs_id"], output_dir=output_dir)
            dl_progress[name] = {"ok": True, "file": os.path.basename(filepath)}
            print(f"  ✓ [{i+1}/{len(flk_items)}] {name}", file=sys.stderr)
        except Exception as e:
            dl_progress[name] = {"ok": False, "error": str(e)}
            print(f"  ✗ [{i+1}/{len(flk_items)}] {name}: {e}", file=sys.stderr)
        if (i + 1) % 20 == 0 and dl_progress_file:
            with open(dl_progress_file, "w") as f:
                json.dump(dl_progress, f, ensure_ascii=False, indent=2)
        time.sleep(delay)

    # 下载规章库
    print(f"=== 下载国家规章库 ({len(gov_items)} 条) ===", file=sys.stderr)
    for i, item in enumerate(gov_items):
        name = item["name"]
        if dl_progress.get(name, {}).get("ok"):
            continue
        try:
            filepath = gov_api.download_as_docx(name, output_dir=output_dir)
            dl_progress[name] = {"ok": True, "file": os.path.basename(filepath)}
            print(f"  ✓ [{i+1}/{len(gov_items)}] {name}", file=sys.stderr)
        except Exception as e:
            dl_progress[name] = {"ok": False, "error": str(e)}
            print(f"  ✗ [{i+1}/{len(gov_items)}] {name}: {e}", file=sys.stderr)
        if (i + 1) % 10 == 0 and dl_progress_file:
            with open(dl_progress_file, "w") as f:
                json.dump(dl_progress, f, ensure_ascii=False, indent=2)
        time.sleep(delay)

    if dl_progress_file:
        with open(dl_progress_file, "w") as f:
            json.dump(dl_progress, f, ensure_ascii=False, indent=2)

    ok = sum(1 for v in dl_progress.values() if v.get("ok"))
    fail = sum(1 for v in dl_progress.values() if not v.get("ok"))
    print(f"✅ 下载完成: 成功 {ok}, 失败 {fail}", file=sys.stderr)
    return dl_progress


def unified_batch_check(names: list, delay: float = 0.5) -> list:
    """统一批量检查：先查 flk，查不到的再查规章库"""
    # Step 1: flk batch check
    print("  [1/2] 查询国家法律法规数据库...", file=sys.stderr)
    flk_results = flk_api.batch_check(names, delay=delay)

    # Step 2: collect not-found names for gov check
    not_found_indices = []
    for i, r in enumerate(flk_results):
        if not r.get("found"):
            not_found_indices.append(i)

    if not_found_indices:
        not_found_names = [flk_results[i]["name"] for i in not_found_indices]
        print(f"  [2/2] {len(not_found_names)} 条未找到，查询国家规章库...", file=sys.stderr)
        gov_results = gov_api.batch_check(not_found_names, delay=delay)

        # Merge gov results back
        for idx, gov_r in zip(not_found_indices, gov_results):
            if gov_r.get("found"):
                flk_results[idx] = {
                    "name": gov_r["name"],
                    "found": True,
                    "title": gov_r.get("title", ""),
                    "status": "有效（规章库）",
                    "sxx": "gov_valid",
                    "gbrq": gov_r.get("date", ""),
                    "sxrq": "",
                    "bbbs": "",
                    "flxz": gov_r.get("type", ""),
                    "source": "gov.cn",
                    "org": gov_r.get("org", ""),
                    "url": gov_r.get("url", ""),
                }
    else:
        print("  [2/2] 全部找到，跳过规章库查询", file=sys.stderr)

    return flk_results


# ── CLI ──

def _fmt(r: dict) -> str:
    src = r.get("source", "flk")
    tag = "规章库" if src == "gov.cn" else r.get("flxz", "")
    return (f"  [{tag}] {r.get('title', r['name'])} | "
            f"状态:{r.get('status','?')} | 公布:{r.get('gbrq', r.get('date',''))}")


def cmd_search(args):
    if args.all_pages:
        results = {"keyword": args.keyword}
        try:
            results["flk"] = flk_api.search_all_pages(
                args.keyword, content_search=args.content, page_size=args.size).get("rows", [])
            results["flk_total"] = len(results["flk"])
        except Exception as e:
            results["flk"] = []
            results["flk_error"] = str(e)
        try:
            results["gov"] = gov_api.search_all_pages(
                args.keyword, content_search=args.content, page_size=args.size).get("rows", [])
            results["gov_total"] = len(results["gov"])
        except Exception as e:
            results["gov"] = []
            results["gov_error"] = str(e)
    else:
        results = unified_search(args.keyword, exact=args.exact,
                                  content_search=args.content, size=args.size)
    flk = results.get("flk", [])
    gov = results.get("gov", [])
    if flk:
        print(f"国家法律法规数据库 ({results.get('flk_total',0)} 条)：")
        for r in flk:
            print(f"  [{r.get('flxz','')}] {r.get('title','')} | "
                  f"状态:{r.get('sxx_text','?')} | 公布:{r.get('gbrq','')}")
    if gov:
        print(f"国家规章库 ({results.get('gov_total',0)} 条)：")
        for r in gov:
            print(f"  [{r.get('type','')}] {r.get('title','')} | "
                  f"公布:{r.get('date','')} | {r.get('org','')}")
    if not flk and not gov:
        print("两个数据库均未找到结果")


def cmd_batch_check(args):
    with open(args.file, "r", encoding="utf-8") as f:
        names = [line.strip() for line in f if line.strip()]
    print(f"共 {len(names)} 条法规待检查", file=sys.stderr)
    results = unified_batch_check(names, delay=args.delay)
    if args.json:
        print(json.dumps(results, ensure_ascii=False, indent=2))
    else:
        for r in results:
            if r.get("found"):
                src = " [规章库]" if r.get("source") == "gov.cn" else ""
                print(f"  ✓ {r['name']} → {r['status']}{src}")
            else:
                print(f"  ✗ {r['name']} → {r.get('status', '未找到')}")
        # Summary
        found_flk = sum(1 for r in results if r.get("found") and r.get("source") != "gov.cn")
        found_gov = sum(1 for r in results if r.get("source") == "gov.cn")
        not_found = sum(1 for r in results if not r.get("found"))
        print(f"\n汇总: 法律法规库找到 {found_flk}, 规章库找到 {found_gov}, 未找到 {not_found}")


def cmd_detail(args):
    d = flk_api.detail(args.id)
    print(json.dumps(d, ensure_ascii=False, indent=2))


def cmd_new_since(args):
    print(f"查询 {args.date} 以来新发布的法规...", file=sys.stderr)

    # 1. flk
    print("  [1/2] 国家法律法规数据库...", file=sys.stderr)
    if args.all_pages:
        flk_resp = flk_api.new_since_all(args.date, law_type=args.type,
                                          all_types=args.all_types)
    else:
        flk_resp = flk_api.new_since(args.date, law_type=args.type,
                                      all_types=args.all_types, size=args.size)
    flk_rows = flk_resp.get("rows", [])

    # 2. gov
    print("  [2/2] 国家规章库...", file=sys.stderr)
    gov_resp = gov_api.new_since(args.date, reg_type=args.type)
    gov_rows = gov_resp.get("rows", [])

    # Output
    if args.json:
        output = {
            "since": args.date,
            "flk": [{"title": r.get("title",""), "flxz": r.get("flxz",""),
                      "gbrq": r.get("gbrq",""), "sxrq": r.get("sxrq",""),
                      "sxx_text": r.get("sxx_text",""), "zdjgName": r.get("zdjgName",""),
                      "bbbs": r.get("bbbs",""),
                      "url": f"https://flk.npc.gov.cn/detail?id={r.get('bbbs','')}" if r.get("bbbs") else "",
                      "source": "flk"} for r in flk_rows],
            "gov": [{"title": r.get("title",""), "type": r.get("type",""),
                      "date": r.get("date",""), "org": r.get("org",""),
                      "url": r.get("url",""), "source": "gov.cn"} for r in gov_rows],
        }
        print(json.dumps(output, ensure_ascii=False, indent=2))
    else:
        if flk_rows:
            print(f"\n国家法律法规数据库 ({flk_resp.get('total',len(flk_rows))} 条)：")
            for r in flk_rows:
                print(f"  [{r.get('flxz','')}] {r.get('title','')} | "
                      f"状态:{r.get('sxx_text','?')} | 公布:{r.get('gbrq','')} | "
                      f"制定机关:{r.get('zdjgName','')}")
        if gov_rows:
            print(f"\n国家规章库 ({len(gov_rows)} 条)：")
            for r in gov_rows:
                print(f"  [{r.get('type','')}] {r.get('title','')} | "
                      f"公布:{r.get('date','')} | {r.get('org','')}")
        if not flk_rows and not gov_rows:
            print("两个数据库均无新发布法规")
        else:
            print(f"\n合计: 法律法规库 {len(flk_rows)} 条, 规章库 {len(gov_rows)} 条")


def cmd_download(args):
    os.makedirs(args.output, exist_ok=True)
    filepath = flk_api.download_file(args.id, output_dir=args.output, fmt=args.format)
    print(f"已下载: {filepath}")


def cmd_batch_download(args):
    os.makedirs(args.output, exist_ok=True)
    with open(args.file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Split into flk (has bbbs) and gov (source=gov.cn, no bbbs)
    flk_items = [r for r in data if r.get("found") and r.get("bbbs")]
    gov_items = [r for r in data if r.get("found") and r.get("source") == "gov.cn" and not r.get("bbbs")]

    ok, fail = 0, 0
    total = len(flk_items) + len(gov_items)
    print(f"共 {total} 条待下载 (法律法规库:{len(flk_items)}, 规章库:{len(gov_items)})", file=sys.stderr)

    # Download from flk
    for i, r in enumerate(flk_items):
        try:
            fp = flk_api.download_file(r["bbbs"], output_dir=args.output, fmt=args.format)
            print(f"  ✓ {r['name']} → {os.path.basename(fp)}")
            ok += 1
        except Exception as e:
            print(f"  ✗ {r['name']} → {e}")
            fail += 1
        if i < len(flk_items) - 1:
            time.sleep(args.delay)
        if (i + 1) % 20 == 0:
            print(f"  进度: {i+1}/{total}", file=sys.stderr)

    # Download from gov (convert content to docx)
    for i, r in enumerate(gov_items):
        try:
            fp = gov_api.download_as_docx(r["name"], output_dir=args.output)
            print(f"  ✓ {r['name']} → {os.path.basename(fp)} [规章库]")
            ok += 1
        except Exception as e:
            print(f"  ✗ {r['name']} → {e}")
            fail += 1
        if i < len(gov_items) - 1:
            time.sleep(args.delay)
        if (len(flk_items) + i + 1) % 20 == 0:
            print(f"  进度: {len(flk_items)+i+1}/{total}", file=sys.stderr)

    print(f"\n完成: 成功 {ok}, 失败 {fail}")


def cmd_new_since_excel(args):
    new_since_to_excel(args.date, args.output, law_type=args.type,
                       all_types=args.all_types)


def cmd_check_excel(args):
    batch_check_to_excel(args.input, args.output, sheet_name=args.sheet,
                         col=args.col, start_row=args.start_row,
                         delay=args.delay, progress_file=args.progress)


def cmd_download_from_check(args):
    with open(args.progress_file, "r") as f:
        check_progress = json.load(f)
    batch_download_from_check(check_progress, args.output,
                              delay=args.delay,
                              dl_progress_file=args.dl_progress)


# ── 合规性对照检查（附件3生成） ──

def _init_ima_client(mcp_config_path=None, kb_ids=None):
    """初始化 IMA Copilot 客户端（用于合规性对照检查）"""
    import importlib
    mcp_config_path = mcp_config_path or os.path.expanduser("~/.kiro/settings/mcp.json")
    with open(mcp_config_path) as f:
        cfg = json.load(f)
    env = cfg["mcpServers"]["ima-copilot"]["env"]
    os.environ["IMA_KNOWLEDGE_BASE_IDS"] = kb_ids or env.get("IMA_KNOWLEDGE_BASE_IDS", "")
    os.environ["IMA_X_IMA_COOKIE"] = env["IMA_X_IMA_COOKIE"]
    os.environ["IMA_X_IMA_BKN"] = env["IMA_X_IMA_BKN"]

    ima_src = os.path.expanduser("~/github/tencent-ima-copilot-mcp/src")
    if ima_src not in sys.path:
        sys.path.insert(0, ima_src)
    from config import get_config
    from ima_client import IMAAPIClient
    return IMAAPIClient(get_config())


def _extract_context_refs(messages) -> list:
    """从 IMA messages 中提取 context_ref 原文片段"""
    contexts = []
    for msg in messages:
        if msg.type.value == "system" and msg.raw:
            try:
                data = json.loads(msg.raw) if isinstance(msg.raw, str) else msg.raw
                if isinstance(data, dict):
                    for key, val_list in data.items():
                        if not key.isdigit():
                            continue
                        if isinstance(val_list, list):
                            for item in val_list:
                                if isinstance(item, dict) and "context" in item:
                                    ctx = item["context"].strip()
                                    if ctx and len(ctx) > 5:
                                        contexts.append(ctx)
            except (json.JSONDecodeError, TypeError):
                pass
    return contexts


async def _ask_ima_with_folder(client, question, kb_id, folder_id=None):
    """用指定文件夹查询 IMA 知识库，返回 (text, contexts)"""
    import asyncio as _asyncio
    from models import InitSessionRequest, EnvInfo, KnowledgeBaseInfoWithFolder

    await client.ensure_valid_token()
    session = await client._get_session()

    folder_ids = [folder_id] if folder_id else []
    init_request = InitSessionRequest(
        envInfo=EnvInfo(robotType=client.config.robot_type, interactType=0),
        relatedUrl=kb_id,
        sceneType=client.config.scene_type,
        msgsLimit=10,
        forbidAutoAddToHistoryList=False,
        knowledgeBaseInfoWithFolder=KnowledgeBaseInfoWithFolder(
            knowledge_base_id=kb_id,
            folder_ids=folder_ids
        )
    )

    url = f"{client.base_url}{client.init_session_endpoint}"
    headers = client._build_headers(for_init_session=True)
    request_json = init_request.model_dump(by_alias=True, exclude_none=True)

    async with session.post(url, json=request_json, headers=headers) as resp:
        result = await resp.json()
        data = result.get("data", result)
        session_id = data.get("sessionId", "") or data.get("session_id", "")

    if not session_id:
        return "", []

    messages = []
    gen = client.ask_question(question, session_id=session_id)
    async for msg in gen:
        messages.append(msg)

    text = client._extract_text_content(messages)
    contexts = _extract_context_refs(messages)
    return text or "", contexts


def _parse_compliance_result(text, contexts) -> dict:
    """解析 IMA 返回结果，提取制度信息"""
    output = {
        "制度名称": "/", "制度文号": "/", "制度修订情况": "/",
        "制度条款": "/", "合规性评价结果": "/", "具体情况": "/"
    }

    if not text:
        return output

    # 检查是否明确没找到
    no_match_kw = ["无对标制度", "没有找到", "未找到", "无法找到", "不存在",
                   "没有直接对应", "未能找到", "没有内容直接对应"]
    if any(kw in text for kw in no_match_kw):
        return output

    # 提取制度名称
    for m in re.findall(r'《([^》]{4,80})》', text):
        if "长庆井下" in m or "川庆" in m:
            output["制度名称"] = m
            break

    # 提取文号
    doc_matches = re.findall(r'(川庆长井[制党发]*〔\d{4}〕\d+号)', text)
    if doc_matches:
        output["制度文号"] = doc_matches[0]

    # 制度条款原文（优先用 context_ref）
    if contexts:
        seen = set()
        all_ctx = []
        for c in contexts:
            if c not in seen:
                seen.add(c)
                all_ctx.append(c)
        full = "\n".join(all_ctx)
        output["制度条款"] = full[:800] + "..." if len(full) > 800 else full

    # 合规性判断
    if output["制度名称"] == "/":
        pass
    elif "不符合" in text and "没有" not in text and "未找到" not in text:
        output["合规性评价结果"] = "不符合"
    elif "建议" in text and ("完善" in text or "补充" in text):
        output["合规性评价结果"] = "建议项"
    else:
        output["合规性评价结果"] = "符合"

    # 具体情况
    if output["合规性评价结果"] in ("不符合", "建议项"):
        cleaned = re.sub(r'\[\d+\]\(@context-ref\?id=\d+\)', '', text)
        cleaned = re.sub(r'\*\*', '', cleaned).strip()
        output["具体情况"] = cleaned[:500] + "..." if len(cleaned) > 500 else cleaned

    return output


async def _gen_attachment3_async(input_excel, output_excel, progress_file,
                                 kb_id, folder_id, max_rows=0, delay=1.5):
    """异步生成附件3"""
    import openpyxl as _openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    print("初始化 IMA 客户端...", file=sys.stderr, flush=True)
    client = _init_ima_client(kb_ids=f"7365585843274444,{kb_id}")

    # 读取附件2
    wb2 = _openpyxl.load_workbook(input_excel, data_only=True)
    ws2 = wb2.active
    rows_data = []
    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, values_only=True):
        if not row[2] and not row[5]:
            continue
        rows_data.append({
            "序号": row[0], "大类": str(row[1] or ""),
            "法规名称": str(row[2] or ""), "发布单位": str(row[3] or ""),
            "实施日期": str(row[4] or ""), "适用条款": str(row[5] or ""),
            "适用部门": str(row[6] or ""),
        })

    total = len(rows_data)
    if max_rows > 0:
        rows_data = rows_data[:max_rows]
    print(f"附件2共 {total} 条，本次处理 {len(rows_data)} 条", file=sys.stderr)

    # 加载进度
    progress = {}
    if progress_file and os.path.exists(progress_file):
        with open(progress_file, "r") as f:
            progress = json.load(f)
    print(f"已完成: {len(progress)}", file=sys.stderr)

    # 逐条查询
    for i, row in enumerate(rows_data):
        key = f"{row['序号']}_{row['法规名称'][:20]}_{row['适用条款'][:20]}"
        if key in progress:
            continue

        条款 = row["适用条款"]
        if len(条款) < 10:
            progress[key] = {"制度名称": "/", "制度文号": "/", "制度修订情况": "/",
                             "制度条款": "/", "合规性评价结果": "/", "具体情况": "/"}
            continue

        question = f"""请在长庆井下技术作业公司的制度中，找到与以下法规条款对应的公司内部制度条款。

法规：《{row['法规名称']}》
条款内容：{条款[:500]}

请按以下格式回答：
- 如果找到对应制度：给出制度名称、文号、逐字引用对应的制度条款原文
- 如果没有找到对应制度：回答"无对标制度"

注意：只找长庆井下技术作业公司自己的制度（文号以"川庆长井"开头），要求制度内容与法规条款管理的是同一件事。"""

        print(f"  [{i+1}/{len(rows_data)}] {row['法规名称'][:25]} ...",
              end=" ", file=sys.stderr, flush=True)

        text, contexts = await _ask_ima_with_folder(client, question, kb_id, folder_id)
        parsed = _parse_compliance_result(text, contexts)
        progress[key] = parsed

        制度 = parsed["制度名称"][:25] if parsed["制度名称"] != "/" else "无"
        print(f"→ {parsed['合规性评价结果']} | {制度}", file=sys.stderr)

        if (i + 1) % 5 == 0 and progress_file:
            with open(progress_file, "w") as f:
                json.dump(progress, f, ensure_ascii=False, indent=2)

        await asyncio.sleep(delay)

    if progress_file:
        with open(progress_file, "w") as f:
            json.dump(progress, f, ensure_ascii=False, indent=2)

    # 生成 Excel
    print("生成Excel...", file=sys.stderr)
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026"

    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    bfont = Font(name="微软雅黑", size=10)
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))
    wrap = Alignment(vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:N1")
    ws.cell(row=1, column=1, value="附件3-规章制度合法性对照检查表").font = Font(name="微软雅黑", size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 15):
        for attr in (hfill, hfont, bdr, center):
            pass
        ws.cell(row=2, column=col).fill = hfill
        ws.cell(row=2, column=col).font = hfont
        ws.cell(row=2, column=col).border = bdr
    ws.cell(row=2, column=1, value="序号")
    ws.merge_cells("B2:H2"); ws.cell(row=2, column=2, value="法律法规")
    ws.merge_cells("I2:L2"); ws.cell(row=2, column=9, value="制度对标情况")
    ws.merge_cells("M2:N2"); ws.cell(row=2, column=13, value="检查情况")

    col_names = ["", "大类", "法律法规名称", "发布单位", "实施日期",
                 "修订情况", "适用条款", "适用部门",
                 "对标制度名称", "对标制度文号", "制度修订情况", "对标制度条款",
                 "合规性评价结果", "具体情况"]
    for col, name in enumerate(col_names, 1):
        c = ws.cell(row=3, column=col, value=name)
        c.font, c.fill, c.border, c.alignment = hfont, hfill, bdr, center

    sfills = {
        "符合": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "不符合": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "建议项": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    for i, rd in enumerate(rows_data):
        r = i + 4
        key = f"{rd['序号']}_{rd['法规名称'][:20]}_{rd['适用条款'][:20]}"
        m = progress.get(key, {})
        vals = [rd["序号"], rd["大类"], rd["法规名称"], rd["发布单位"], rd["实施日期"],
                "保留", rd["适用条款"], rd["适用部门"],
                m.get("制度名称", "/"), m.get("制度文号", "/"), m.get("制度修订情况", "/"),
                m.get("制度条款", "/"), m.get("合规性评价结果", "/"), m.get("具体情况", "/")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.font, c.border, c.alignment = bfont, bdr, wrap
        result = m.get("合规性评价结果", "/")
        if result in sfills:
            ws.cell(row=r, column=13).fill = sfills[result]

    for i, w in enumerate([6, 15, 25, 20, 12, 10, 50, 15, 25, 18, 10, 50, 10, 15], 1):
        ws.column_dimensions[_openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    wb.save(output_excel)

    stats = {}
    for v in progress.values():
        s = v.get("合规性评价结果", "/")
        stats[s] = stats.get(s, 0) + 1
    print(f"✅ 已保存: {output_excel}", file=sys.stderr)
    for s, c in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c}", file=sys.stderr)

    await client.close()
    return output_excel


def gen_attachment3(input_excel, output_excel, kb_id="7453273606867451",
                    folder_id="folder_7453273833360160",
                    progress_file=None, max_rows=0, delay=1.5):
    """从附件2生成附件3（合规性对照检查表）

    Args:
        input_excel: 附件2 Excel 文件路径
        output_excel: 输出附件3 Excel 文件路径
        kb_id: IMA 知识库 Copilot ID
        folder_id: 限定搜索的文件夹 ID（目标公司制度）
        progress_file: 进度文件路径（支持断点续做）
        max_rows: 最大处理行数（0=全部）
        delay: 请求间隔秒数
    """
    import asyncio as _asyncio
    return _asyncio.run(_gen_attachment3_async(
        input_excel, output_excel, progress_file,
        kb_id, folder_id, max_rows, delay
    ))


import asyncio

def cmd_gen_attachment3(args):
    gen_attachment3(
        input_excel=args.input,
        output_excel=args.output,
        kb_id=args.kb_id,
        folder_id=args.folder_id,
        progress_file=args.progress,
        max_rows=args.max_rows,
        delay=args.delay,
    )


def main():
    p = argparse.ArgumentParser(
        description="中国法律法规统一查询（自动查 flk.npc.gov.cn + gov.cn 规章库）")
    sub = p.add_subparsers(dest="command", required=True)

    s = sub.add_parser("search", help="统一搜索（两库同时查）")
    s.add_argument("keyword")
    s.add_argument("--exact", action="store_true")
    s.add_argument("--content", action="store_true", help="按正文内容搜索（默认按标题）")
    s.add_argument("--all-pages", action="store_true", help="获取所有页结果")
    s.add_argument("--size", type=int, default=10)
    s.set_defaults(func=cmd_search)

    s = sub.add_parser("batch-check", help="统一批量检查（flk 优先，fallback 规章库）")
    s.add_argument("file", help="法规名称列表文件")
    s.add_argument("--delay", type=float, default=0.5)
    s.add_argument("--json", action="store_true")
    s.set_defaults(func=cmd_batch_check)

    s = sub.add_parser("detail", help="查看法规详情（flk）")
    s.add_argument("id")
    s.set_defaults(func=cmd_detail)

    s = sub.add_parser("new-since", help="查找指定日期后新发布的法规（两库联查）")
    s.add_argument("date", help="起始日期 (yyyy-MM-dd)")
    s.add_argument("--type", help="法规类型")
    s.add_argument("--all-types", action="store_true", help="查所有类型（含地方法规）")
    s.add_argument("--all-pages", action="store_true", help="获取所有页结果")
    s.add_argument("--json", action="store_true", help="输出 JSON（含完整字段，可用于生成目录）")
    s.add_argument("--size", type=int, default=100)
    s.set_defaults(func=cmd_new_since)

    s = sub.add_parser("download", help="下载法规文件（flk）")
    s.add_argument("id")
    s.add_argument("--format", default="docx", choices=["docx", "pdf"])
    s.add_argument("--output", default=".")
    s.set_defaults(func=cmd_download)

    s = sub.add_parser("batch-download", help="批量下载（从 batch-check JSON）")
    s.add_argument("file")
    s.add_argument("--format", default="docx", choices=["docx", "pdf"])
    s.add_argument("--output", default=".")
    s.add_argument("--delay", type=float, default=1.0)
    s.set_defaults(func=cmd_batch_download)

    s = sub.add_parser("new-since-excel", help="查询新发布法规并输出Excel")
    s.add_argument("date", help="起始日期 (yyyy-MM-dd)")
    s.add_argument("--output", "-o", required=True, help="输出Excel文件路径")
    s.add_argument("--type", help="法规类型")
    s.add_argument("--all-types", action="store_true", default=True)
    s.set_defaults(func=cmd_new_since_excel)

    s = sub.add_parser("check-excel", help="从Excel读取法规名称，批量检查有效性并输出Excel")
    s.add_argument("input", help="输入Excel文件路径")
    s.add_argument("--output", "-o", required=True, help="输出Excel文件路径")
    s.add_argument("--sheet", help="工作表名称")
    s.add_argument("--col", default="B", help="法规名称所在列（默认B）")
    s.add_argument("--start-row", type=int, default=2, help="数据起始行（默认2）")
    s.add_argument("--delay", type=float, default=0.5)
    s.add_argument("--progress", help="进度文件路径（支持断点续查）")
    s.set_defaults(func=cmd_check_excel)

    s = sub.add_parser("download-from-check", help="从检查进度文件批量下载法规全文")
    s.add_argument("progress_file", help="check-excel 的进度JSON文件")
    s.add_argument("--output", "-o", default="./downloads", help="下载目录")
    s.add_argument("--delay", type=float, default=0.5)
    s.add_argument("--dl-progress", help="下载进度文件（支持断点续下）")
    s.set_defaults(func=cmd_download_from_check)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
