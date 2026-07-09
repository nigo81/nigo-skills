#!/usr/bin/env python3
"""
export_report.py - 审计报告检查结果导出工具

将 run_check.py 产出的 results.json 转换为：
1. 7 sheet Excel 报告（摘要 / 报表内勾稽 / 表注勾稽 / 附注内勾稽 / 横加竖加 / 文本格式 / 检查项）
2. Markdown 复核报告（按问题/存疑分章）

设计依据：
- openspec/changes/output-quality-and-precision/design.md D1（审计师视角7 sheet + severity二元化）
- openspec/changes/output-quality-and-precision/specs/structured-report-output/spec.md

audit-only 约束：只读 results.json，不修改源审计报告文件。

用法：
    python3 export_report.py <results.json> [-o <输出目录或xlsx路径>]

输出：
- 默认输出到 results.json 同目录
- <source>_检查报告.xlsx
- <source>_复核报告.md
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 样式常量
# ============================================================

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
NORMAL_FONT = Font(name="微软雅黑", size=10)
NOTE_FONT = Font(name="微软雅黑", size=9, color="595959")

# severity 着色（二元：error=问题，warning=存疑）
SEVERITY_FILLS: dict[str, PatternFill] = {
    "error": PatternFill("solid", fgColor="FFC7CE"),
    "warning": PatternFill("solid", fgColor="FFEB9C"),
}
SEVERITY_FONTS: dict[str, Font] = {
    "error": Font(name="微软雅黑", size=10, color="9C0006"),
    "warning": Font(name="微软雅黑", size=10, color="9C6500"),
}
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# severity 中文标签（二元化）
SEVERITY_LABEL: dict[str, str] = {
    "error": "问题",
    "warning": "存疑",
}

# severity 说明（用于摘要 sheet）
SEVERITY_NOTES: list[tuple[str, str, str]] = [
    ("问题", "确认错误需修改（Agent复核确认）", "FFC7CE"),
    ("存疑", "可疑需人工核实（Agent复核存疑）", "FFEB9C"),
]

# check_type 中文显示
CHECK_TYPE_LABEL: dict[str, str] = {
    "报表间": "报表间勾稽",
    "表注": "表注勾稽",
    "附注内": "附注内勾稽",
    "横加": "横加验算",
    "竖加": "竖加验算",
    "文本": "文本检查",
    "格式": "格式检查",
}


# ============================================================
# 数据加载与规范化
# ============================================================


def load_results(path: Path) -> dict[str, Any]:
    """读取 results.json，容错处理缺失字段。"""
    try:
        text = path.read_text(encoding="utf-8")
    except FileNotFoundError:
        raise SystemExit(f"错误：找不到结果文件 {path}")
    except OSError as e:
        raise SystemExit(f"错误：读取 {path} 失败：{e}")

    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise SystemExit(f"错误：{path} 不是有效的 JSON：{e}")

    if not isinstance(data, dict):
        raise SystemExit(f"错误：{path} 顶层结构应为对象")

    # 容错：meta 和 results 缺失时给空值
    meta = data.get("meta") or {}
    if not isinstance(meta, dict):
        meta = {}
    results = data.get("results") or []
    if not isinstance(results, list):
        results = []
    data["meta"] = meta
    data["results"] = results
    return data


def field(item: dict[str, Any], name: str, default: str = "") -> str:
    """从 item 提取字段，缺失返回默认空字符串（容错）。"""
    value = item.get(name, default)
    if value is None:
        return default
    if isinstance(value, (int, float)):
        # 数字统一转字符串，保留可读形式
        if isinstance(value, float):
            return f"{value:.2f}"
        return str(value)
    return str(value)


def effective_severity(item: dict[str, Any]) -> str | None:
    """计算有效的严重程度（仅返回 error 或 warning）。

    passed=True → None（通过项不进详细sheet）
    否则取 severity 字段（error/warning），info 返回 None
    """
    if item.get("passed") is True:
        return None
    sev = str(item.get("severity", "")).lower().strip()
    if sev in {"error", "warning"}:
        return sev
    # info 返回 None，通过的项不在详细 sheet 显示
    return None


def is_financial_statement_vertical(item: dict[str, Any]) -> bool:
    """判断是否为四大财务报表的竖加（用于分类到报表内勾稽）。

    四大表：资产负债表、利润表、现金流量表、权益变动表
    """
    source_loc = field(item, "source_location")
    check_type = field(item, "check_type")

    # 必须是竖加，且 source_location 包含四大表名
    if check_type != "竖加":
        return False

    fs_keywords = [
        "资产负债表",
        "利润表",
        "现金流量表",
        "权益变动表",
        "所有者权益变动表",
    ]
    for keyword in fs_keywords:
        if keyword in source_loc:
            return True
    return False


def format_check_location(item: dict[str, Any]) -> str:
    """格式化检查位置（人话，去掉表格ID）。

    从 source_location、target_location 和 context 拼接。
    如："合并资产负债表 - 流动资产合计" 或 "项目 - 合计"
    """
    source = field(item, "source_location")
    target = field(item, "target_location")
    context = field(item, "context")

    # 去掉 "表格ID XXX" 前缀
    source = re.sub(r'^表格ID\s*\d+\s*[:：]?\s*', '', source).strip()
    target = re.sub(r'^表格ID\s*\d+\s*[:：]?\s*', '', target).strip()

    # 如果 source/target 为空，使用 context（针对横加/竖加的情况）
    if not source and not target and context:
        return context

    if source and target:
        return f"{source} → {target}"
    elif source:
        return source
    elif target:
        return target
    elif context:
        return context
    else:
        return ""


def format_check_content(item: dict[str, Any]) -> str:
    """格式化检查内容（人话）。

    从 rule_name 和 description 拼接，去掉表格ID等无意义信息。
    """
    rule = field(item, "rule_name")
    desc = field(item, "description")

    # 去掉 "id=N " 前缀
    rule = re.sub(r'^id=\d+\s*', '', rule).strip()

    # 如果 description 以 "验算：" 开头，去掉验算部分，rule 和 desc 合并
    if desc.startswith("验算："):
        desc = desc[3:].strip()
    elif "：" in desc and len(desc.split("：")[0]) < 20:
        # description 开头可能有简短提示，保留
        pass

    # 如果 rule 和 desc 很像（rule 是 desc 的前缀），只保留 desc
    if desc.startswith(rule):
        return desc
    elif rule and desc:
        return f"{rule}：{desc}"
    elif desc:
        return desc
    else:
        return rule


def format_page_hyperlink(item: dict[str, Any], results_dir: Path) -> str:
    """生成页码的超链接（绝对路径）。

    从 source_file 和 page 生成：
    =HYPERLINK("file:///绝对路径#page=N","第N页")

    如果 source_file 不是绝对路径，则相对于 results_dir 解析。
    """
    source_file = field(item, "source_file")
    page = field(item, "page")

    if not source_file or not page:
        chapter = field(item, "chapter")
        return chapter or page  # Word 无 page 用 chapter 章节定位（附注五-N 科目）

    # 解析路径
    source_path = Path(source_file)
    if not source_path.is_absolute():
        # 相对于 results_dir 解析
        source_path = results_dir / source_file

    # 转为绝对路径字符串
    abs_path = str(source_path.absolute())

    # 生成超链接公式
    return f'=HYPERLINK("file://{abs_path}#page={page}","第{page}页")'


def format_numeric_value(item: dict[str, Any], field_name: str) -> str:
    """格式化数值字段（expected/actual/difference）。

    对于非数值检查，返回空字符串。
    """
    value = item.get(field_name)
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        # 数字格式化为带千分位的字符串
        if isinstance(value, float):
            return f"{value:,.2f}"
        return f"{value:,}"
    return str(value)


def format_review_conclusion(item: dict[str, Any]) -> str:
    """格式化复核结论（从 evidence 提取）。

    优先提取 "Agent复核：XXX"（Claude 终审结论）；
    warning（存疑）若无 Agent 复核标记，按检查类型生成"为什么存疑"默认说明，
    避免审计师看到技术性 evidence（如 ai_worker_错别字）而看不懂存疑原因。
    """
    evidence = field(item, "evidence")
    # 1. 优先 Agent 复核标记（Claude Step5 终审产出）
    if evidence:
        match = re.search(r'Agent复核[：:]\s*(.*?)(?:\[|$)', evidence)
        if match:
            return match.group(1).strip()

    # 2. warning（存疑）无 Agent 复核 → 按 check_type 生成"为什么存疑"说明
    if effective_severity(item) == "warning":
        ct = field(item, "check_type")
        why = {
            "表注": "报表数与附注数不一致或取数失败，需核对附注披露口径/人工提数复核",
            "附注内": "变动表期初+增-减≠期末，需核对是否有重分类或取数误差",
            "横加": "列间加减不平，需核对是否百分比列/子项等特殊结构",
            "竖加": "明细之和≠合计，需核对是否百分比列/子项/取数误差",
            "文本": "疑似错别字或格式问题，需核对原文确认（可能 PDF 提取误差）",
            "格式": "页码/公司名/单位等格式可疑，需人工核实",
            "报表间": "报表间勾稽不平衡，需核对报表数据",
        }.get(ct, "需人工核实")
        return f"存疑：{why}"

    # 3. error（问题）或其他：返回原始 evidence
    return evidence


# ============================================================
# Excel 写出
# ============================================================


def _set_col_widths(ws, widths: list[int]) -> None:
    """按列序号设置列宽（字符数）。"""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """写表头行（深蓝填充 + 白字加粗 + 居中 + 边框）。"""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def _write_data_row(
    ws,
    row: int,
    values: list[str],
    severity: str | None = None,
    zebra: bool = False,
    center_cols: set[int] | None = None,
) -> None:
    """写一行数据。severity 非 None 时按 severity 着色；zebra=True 时叠加斑马底色。"""
    center_set = center_cols or set()
    fill = SEVERITY_FILLS.get(severity) if severity else None
    if fill is None and zebra:
        fill = ZEBRA_FILL
    font = SEVERITY_FONTS.get(severity) if severity else NORMAL_FONT
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN if col in center_set else LEFT_ALIGN
        if fill is not None:
            cell.fill = fill


def _freeze_header(ws) -> None:
    """冻结首行。"""
    ws.freeze_panes = "A2"


# ---------- Sheet 1: 摘要 ----------


def write_summary_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    meta = data["meta"]
    results = data["results"]
    company = field(meta, "company", "（未填写公司）")
    scope = field(meta, "scope", "standard")
    period = field(meta, "period", "")
    checked_at = field(meta, "checked_at", "")
    source = field(meta, "source", "")

    ws = wb.create_sheet("摘要")
    NCOL = 6  # 摘要 sheet 总列数

    # 标题行
    ws.cell(row=1, column=1, value=f"{company} 审计报告检查摘要").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # 元信息
    meta_lines: list[tuple[str, str]] = [
        ("检查期间", period),
        ("检查范围", scope),
        ("源文件", source),
        ("检查时间", checked_at),
    ]
    row = 2
    for label, value in meta_lines:
        ws.cell(row=row, column=1, value=label).font = SECTION_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOL)
        row += 1

    row += 1  # 空行

    # 总计表（5列：总检查项，通过，问题，存疑，问题率）
    ws.cell(row=row, column=1, value="一、总计").font = SECTION_FONT
    row += 1
    total_headers = ["总检查项", "通过", "问题", "存疑", "问题率"]
    _write_header_row(ws, row, total_headers)
    row += 1

    n_total = len(results)
    n_pass = sum(1 for r in results if r.get("passed") is True)
    n_error = sum(1 for r in results if effective_severity(r) == "error")
    n_warn = sum(1 for r in results if effective_severity(r) == "warning")
    issue_rate = f"{(n_error + n_warn) / n_total * 100:.1f}%" if n_total else "—"
    _write_data_row(
        ws, row,
        [str(n_total), str(n_pass), str(n_error), str(n_warn), issue_rate],
        center_cols=range(1, 6),
    )
    row += 2

    # 按 sheet 分类统计（不含通过项）
    ws.cell(row=row, column=1, value="二、按类别统计").font = SECTION_FONT
    row += 1
    type_headers = ["检查类别", "问题", "存疑", "合计"]
    _write_header_row(ws, row, type_headers)
    row += 1

    # 定义各 sheet 的筛选逻辑
    def filter_report_internal(item):
        ct = field(item, "check_type")
        sev = effective_severity(item)
        if sev not in {"error", "warning"}:
            return False
        if ct == "报表间":
            return True
        if ct == "竖加" and is_financial_statement_vertical(item):
            return True
        return False

    def filter_note_ref(item):
        ct = field(item, "check_type")
        sev = effective_severity(item)
        return ct == "表注" and sev in {"error", "warning"}

    def filter_notes_internal(item):
        ct = field(item, "check_type")
        sev = effective_severity(item)
        return ct == "附注内" and sev in {"error", "warning"}

    def filter_horizontal_vertical(item):
        ct = field(item, "check_type")
        sev = effective_severity(item)
        if sev not in {"error", "warning"}:
            return False
        if ct in {"横加", "竖加"}:
            # 排除四表竖加（已归报表内勾稽）
            if ct == "竖加" and is_financial_statement_vertical(item):
                return False
            return True
        return False

    def filter_text_format(item):
        ct = field(item, "check_type")
        sev = effective_severity(item)
        return ct in {"文本", "格式"} and sev in {"error", "warning"}

    categories = [
        ("报表内勾稽", filter_report_internal),
        ("表注勾稽", filter_note_ref),
        ("附注内勾稽", filter_notes_internal),
        ("横加竖加", filter_horizontal_vertical),
        ("文本格式", filter_text_format),
    ]

    for cat_name, filter_func in categories:
        items = [r for r in results if filter_func(r)]
        cat_error = sum(1 for r in items if effective_severity(r) == "error")
        cat_warn = sum(1 for r in items if effective_severity(r) == "warning")
        cat_total = len(items)
        _write_data_row(
            ws, row,
            [cat_name, str(cat_error), str(cat_warn), str(cat_total)],
            center_cols=(2, 3, 4),
        )
        row += 1

    row += 1

    # severity 说明
    ws.cell(row=row, column=1, value="三、严重程度说明").font = SECTION_FONT
    row += 1
    sev_headers = ["严重程度", "说明"]
    _write_header_row(ws, row, sev_headers)
    # 合并第2列到第 NCOL 列展示说明
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOL)
    row += 1

    sev_label_to_key = {"问题": "error", "存疑": "warning"}
    for label, desc, _color in SEVERITY_NOTES:
        sev_key = sev_label_to_key[label]
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = SEVERITY_FONTS[sev_key]
        cell.fill = SEVERITY_FILLS[sev_key]
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        desc_cell = ws.cell(row=row, column=2, value=desc)
        desc_cell.font = NORMAL_FONT
        desc_cell.alignment = LEFT_ALIGN
        desc_cell.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOL)
        row += 1

    # 列宽
    _set_col_widths(ws, [22, 12, 12, 12, 12, 12])


# ---------- 详细 sheet 通用结构 ----------

DETAIL_HEADERS = [
    "序号", "严重程度", "页码", "检查位置", "检查内容",
    "应为", "实际", "差异", "复核结论",
]
DETAIL_WIDTHS = [6, 10, 12, 30, 40, 16, 16, 14, 40]


def _write_detail_sheet(
    wb: Workbook,
    sheet_name: str,
    results: list[dict[str, Any]],
    results_dir: Path,
    empty_hint: str,
) -> None:
    """通用：写详细问题 sheet（所有详细sheet共结构）。"""
    ws = wb.create_sheet(sheet_name)

    if not results:
        ws.cell(row=1, column=1, value=empty_hint).font = NOTE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DETAIL_HEADERS))
        _set_col_widths(ws, DETAIL_WIDTHS)
        return

    _write_header_row(ws, 1, DETAIL_HEADERS)
    _freeze_header(ws)

    for idx, item in enumerate(results, start=1):
        sev = effective_severity(item)
        values = [
            str(idx),
            SEVERITY_LABEL.get(sev, sev),
            format_page_hyperlink(item, results_dir),
            format_check_location(item),
            format_check_content(item),
            format_numeric_value(item, "expected"),
            format_numeric_value(item, "actual"),
            format_numeric_value(item, "difference"),
            format_review_conclusion(item),
        ]
        _write_data_row(
            ws, idx + 1, values,
            severity=sev,
            zebra=(idx % 2 == 0),
            center_cols=(1, 2),
        )

    _set_col_widths(ws, DETAIL_WIDTHS)


# ---------- Sheet 2: 报表内勾稽 ----------


def write_report_internal_sheet(wb: Workbook, results: list[dict[str, Any]], results_dir: Path) -> None:
    """Sheet 2: 报表内勾稽 - 报表间 + 四表竖加，仅 error/warning。"""
    filtered = [
        r for r in results
        if (
            (field(r, "check_type") == "报表间" and effective_severity(r) in {"error", "warning"})
            or (field(r, "check_type") == "竖加" and is_financial_statement_vertical(r) and effective_severity(r) in {"error", "warning"})
        )
    ]
    _write_detail_sheet(
        wb, "报表内勾稽", filtered, results_dir,
        "未发现报表内勾稽类问题（报表间/四表竖加）",
    )


# ---------- Sheet 3: 表注勾稽 ----------


def write_note_ref_sheet(wb: Workbook, results: list[dict[str, Any]], results_dir: Path) -> None:
    """Sheet 3: 表注勾稽 - 表注，仅 error/warning。"""
    filtered = [
        r for r in results
        if field(r, "check_type") == "表注" and effective_severity(r) in {"error", "warning"}
    ]
    _write_detail_sheet(
        wb, "表注勾稽", filtered, results_dir,
        "未发现表注勾稽类问题",
    )


# ---------- Sheet 4: 附注内勾稽 ----------


def write_notes_internal_sheet(wb: Workbook, results: list[dict[str, Any]], results_dir: Path) -> None:
    """Sheet 4: 附注内勾稽 - 附注内，仅 error/warning。"""
    filtered = [
        r for r in results
        if field(r, "check_type") == "附注内" and effective_severity(r) in {"error", "warning"}
    ]
    _write_detail_sheet(
        wb, "附注内勾稽", filtered, results_dir,
        "未发现附注内勾稽类问题",
    )


# ---------- Sheet 5: 横加竖加 ----------


def write_horizontal_vertical_sheet(wb: Workbook, results: list[dict[str, Any]], results_dir: Path) -> None:
    """Sheet 5: 横加竖加 - 横加+竖加（排除四表竖加），仅 error/warning。"""
    filtered = [
        r for r in results
        if (
            field(r, "check_type") in {"横加", "竖加"}
            and not is_financial_statement_vertical(r)
            and effective_severity(r) in {"error", "warning"}
        )
    ]
    _write_detail_sheet(
        wb, "横加竖加", filtered, results_dir,
        "未发现横加竖加类问题（附注明细表）",
    )


def write_check_items_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    """Sheet 7: 检查项 - 列示所有执行的检查规则（含通过项）。"""
    results = data["results"]

    if not results:
        ws = wb.create_sheet("检查项")
        ws.cell(row=1, column=1, value="无检查项").font = NOTE_FONT
        return

    ws = wb.create_sheet("检查项")

    # 表头
    headers = ["检查类别", "检查规则", "结果", "检查内容", "涉及位置"]
    _write_header_row(ws, 1, headers)
    _freeze_header(ws)

    # 定义结果着色（通过绿、问题红、存疑黄）
    RESULT_FILLS = {
        "通过": PatternFill("solid", fgColor="C6E0B4"),
        "问题": PatternFill("solid", fgColor="FFC7CE"),
        "存疑": PatternFill("solid", fgColor="FFEB9C"),
    }
    RESULT_FONTS = {
        "通过": Font(name="微软雅黑", size=10, color="006100"),
        "问题": Font(name="微软雅黑", size=10, color="9C0006"),
        "存疑": Font(name="微软雅黑", size=10, color="9C6500"),
    }

    # 辅助函数：将 check_type 映射到检查类别
    def map_check_category(item: dict[str, Any]) -> str:
        """将原始 check_type 映射到检查类别。"""
        ct = field(item, "check_type")
        loc = field(item, "source_location")

        # 报表间勾稽：报表间 + 四表竖加
        if ct == "报表间":
            return "报表内勾稽"
        elif ct == "竖加" and is_financial_statement_vertical(item):
            return "报表内勾稽"
        # 表注勾稽
        elif ct == "表注":
            return "表注勾稽"
        # 附注内勾稽
        elif ct == "附注内":
            return "附注内勾稽"
        # 横加竖加：横加 + 竖加（非四表）
        elif ct in {"横加", "竖加"}:
            return "横加竖加"
        # 文本格式
        elif ct in {"文本", "格式"}:
            return "文本格式"
        else:
            return ct

    # 辅助函数：确定结果标签
    def get_result_label(item: dict[str, Any]) -> str:
        """根据 passed 和 severity 返回结果标签。"""
        if item.get("passed") is True:
            return "通过"
        sev = field(item, "severity").lower()
        if sev == "error":
            return "问题"
        elif sev == "warning":
            return "存疑"
        else:
            return "通过"

    # 辅助函数：格式化检查内容（人话）
    def format_check_content_human(item: dict[str, Any]) -> str:
        """格式化检查内容为易懂的人话。"""
        desc = field(item, "description")
        rule = field(item, "rule_name")

        # 如果 description 以 "XXX竖加:" 开头，简化为 "验证竖加"
        if "竖加:" in desc and "明细行求和" in desc:
            return f"验证竖加（明细之和=合计）"

        # 如果 description 以 "XXX横加:" 开头，简化为 "验证横加"
        if "横加:" in desc and ("余额" in desc or "期初" in desc):
            return f"验证横加（列间勾稽关系）"

        # 如果是 L1 报表间勾稽，简化描述
        if "资产合计" in rule and "流动" in desc:
            return "流动资产合计+非流动资产合计=资产总计"
        elif "负债合计" in rule and "流动" in desc:
            return "流动负债合计+非流动负债合计=负债合计"
        elif "权益平衡" in rule:
            return "负债合计+所有者权益合计=负债和所有者权益总计"
        elif "资产负债平衡" in rule:
            return "资产总计=负债和所有者权益总计"
        elif "营业利润推导" in rule:
            return "营业利润公式推导验证"
        elif "利润总额推导" in rule:
            return "利润总额=营业利润+营业外收入-营业外支出"
        elif "净利润推导" in rule:
            return "净利润=利润总额-所得税费用"
        elif "现金流量净增加额" in rule:
            return "经营+投资+筹资活动净额=净增加额"
        elif "现金余额勾稽" in rule:
            return "期初余额+净增加额=期末余额"
        elif "权益变动期末余额" in rule:
            return "权益变动表期末余额验证"

        # 其他情况返回 description（去掉 id=N 前缀）
        desc_clean = re.sub(r'^id=\d+\s*', '', desc).strip()
        return desc_clean

    # 辅助函数：格式化涉及位置
    def format_location_human(item: dict[str, Any]) -> str:
        """格式化涉及位置为易懂的人话。"""
        ct = field(item, "check_type")
        loc = field(item, "source_location")
        page = field(item, "page")

        # 去掉 "表格ID XXX" 前缀
        loc_clean = re.sub(r'^表格ID\s*\d+\s*[:：]?\s*', '', loc).strip()

        # 根据检查类型返回不同格式
        if ct == "报表间":
            return f"{loc_clean}（第{page}页）"
        elif ct in {"竖加", "横加"}:
            # 如果 context 存在，用 context
            ctx = field(item, "context")
            if ctx and ctx != loc_clean:
                return f"{ctx}（第{page}页）"
            return f"{loc_clean}（第{page}页）"
        elif ct in {"文本", "格式"}:
            if page:
                return f"第{page}页"
            return "全文检查"
        else:
            if page:
                return f"{loc_clean}（第{page}页）"
            return loc_clean

    # 写入所有结果（按检查类别分组）
    # 先按检查类别分组
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in results:
        category = map_check_category(item)
        if category not in grouped:
            grouped[category] = []
        grouped[category].append(item)

    # 按检查类别顺序写入（报表内勾稽 → 表注勾稽 → 附注内勾稽 → 横加竖加 → 文本格式）
    category_order = ["报表内勾稽", "表注勾稽", "附注内勾稽", "横加竖加", "文本格式"]
    row = 2

    for category in category_order:
        if category not in grouped:
            continue

        items = grouped[category]
        for item in items:
            result_label = get_result_label(item)
            values = [
                category,
                field(item, "rule_name"),
                result_label,
                format_check_content_human(item),
                format_location_human(item),
            ]

            # 写入行，按结果着色
            fill = RESULT_FILLS.get(result_label)
            font = RESULT_FONTS.get(result_label, NORMAL_FONT)
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = font
                cell.border = THIN_BORDER
                cell.alignment = LEFT_ALIGN
                if fill is not None:
                    cell.fill = fill

            row += 1

    # 列宽
    _set_col_widths(ws, [18, 25, 10, 35, 25])


# ---------- Sheet 6: 文本格式 ----------


def write_text_format_sheet(wb: Workbook, results: list[dict[str, Any]], results_dir: Path) -> None:
    """Sheet 6: 文本格式 - 文本+格式，仅 error/warning。"""
    filtered = [
        r for r in results
        if field(r, "check_type") in {"文本", "格式"} and effective_severity(r) in {"error", "warning"}
    ]
    _write_detail_sheet(
        wb, "文本格式", filtered, results_dir,
        "未发现文本格式类问题",
    )


# ---------- Excel 入口 ----------


def write_excel(data: dict[str, Any], results_path: Path, out_path: Path) -> None:
    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    results_dir = results_path.parent

    write_summary_sheet(wb, data)
    write_report_internal_sheet(wb, data["results"], results_dir)
    write_note_ref_sheet(wb, data["results"], results_dir)
    write_notes_internal_sheet(wb, data["results"], results_dir)
    write_horizontal_vertical_sheet(wb, data["results"], results_dir)
    write_text_format_sheet(wb, data["results"], results_dir)
    write_check_items_sheet(wb, data)  # 新增：检查项 sheet

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ============================================================
# Markdown 复核报告
# ============================================================


def write_markdown(data: dict[str, Any], out_path: Path) -> None:
    meta = data["meta"]
    results = data["results"]
    company = field(meta, "company", "（未填写公司）")
    scope = field(meta, "scope", "standard")
    period = field(meta, "period", "")
    checked_at = field(meta, "checked_at", "")

    n_total = len(results)
    n_pass = sum(1 for r in results if r.get("passed") is True)
    n_error = sum(1 for r in results if effective_severity(r) == "error")
    n_warn = sum(1 for r in results if effective_severity(r) == "warning")

    lines: list[str] = []
    lines.append(f"# {company} 审计报告核查报告")
    lines.append("")
    if period:
        lines.append(f"> 检查期间：{period}")
    if checked_at:
        lines.append(f"> 检查时间：{checked_at}")
    if period or checked_at:
        lines.append("")

    # 摘要
    lines.append("## 检查摘要")
    lines.append("")
    lines.append(f"- 检查范围：{scope}")
    lines.append(
        f"- 总检查项：{n_total} | 通过：{n_pass} | "
        f"问题：{n_error} | 存疑：{n_warn}"
    )
    lines.append("")

    if n_error + n_warn == 0:
        lines.append("> 未发现问题和存疑项，报告检查通过。")
        lines.append("")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text("\n".join(lines), encoding="utf-8")
        return

    def render_issue(idx: int, item: dict[str, Any]) -> list[str]:
        location = format_check_location(item)
        content = format_check_content(item)
        expected = format_numeric_value(item, "expected")
        actual = format_numeric_value(item, "actual")
        diff = format_numeric_value(item, "difference")
        evidence = format_review_conclusion(item)
        ct = CHECK_TYPE_LABEL.get(field(item, "check_type"), field(item, "check_type"))
        page = field(item, "page")

        block: list[str] = []
        block.append(f"{idx}. **{content}**")
        if location:
            block.append(f"   - 位置：{location}")
        if page:
            block.append(f"   - 页码：第{page}页")
        if ct:
            block.append(f"   - 类别：{ct}")
        val_parts = []
        if expected:
            val_parts.append(f"应为：{expected}")
        if actual:
            val_parts.append(f"实际：{actual}")
        if diff:
            val_parts.append(f"差异：{diff}")
        if val_parts:
            block.append(f"   - {' | '.join(val_parts)}")
        if evidence:
            block.append(f"   - 复核结论：{evidence}")
        # 去掉空行
        return [b for b in block if b.strip()]

    # 问题
    lines.append("## 问题（需修改）")
    lines.append("")
    errors = [r for r in results if effective_severity(r) == "error"]
    if errors:
        lines.append(
            f"> 共 {len(errors)} 项确认错误，需修改报告。"
        )
        lines.append("")
        for i, item in enumerate(errors, start=1):
            lines.extend(render_issue(i, item))
            lines.append("")
    else:
        lines.append("> 未发现确认错误。")
        lines.append("")

    # 存疑
    lines.append("## 存疑（需核实）")
    lines.append("")
    warnings = [r for r in results if effective_severity(r) == "warning"]
    if warnings:
        lines.append(f"> 共 {len(warnings)} 项存疑，建议人工复核。")
        lines.append("")
        for i, item in enumerate(warnings, start=1):
            lines.extend(render_issue(i, item))
            lines.append("")
    else:
        lines.append("> 未发现存疑项。")
        lines.append("")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")


# ============================================================
# CLI
# ============================================================


def derive_output_paths(results_path: Path, meta: dict[str, Any], output_arg: str | None) -> tuple[Path, Path]:
    """根据 -o 参数和 meta.source 推导 xlsx / md 输出路径。"""
    source = field(meta, "source")
    if source:
        source_stem = Path(source).stem
    else:
        source_stem = results_path.stem

    safe_stem = source_stem or "审计报告"
    for ch in '/\\:*?"<>|':
        safe_stem = safe_stem.replace(ch, "_")

    default_dir = results_path.parent

    if output_arg is None:
        xlsx_path = default_dir / f"{safe_stem}_检查报告.xlsx"
        md_path = default_dir / f"{safe_stem}_复核报告.md"
        return xlsx_path, md_path

    out = Path(output_arg).expanduser()
    if out.suffix.lower() == ".xlsx":
        xlsx_path = out
        md_path = out.parent / f"{out.stem}_复核报告.md"
        return xlsx_path, md_path

    xlsx_path = out / f"{safe_stem}_检查报告.xlsx"
    md_path = out / f"{safe_stem}_复核报告.md"
    return xlsx_path, md_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="将审计报告检查结果 results.json 导出为 7 sheet Excel + Markdown 复核报告",
        usage="python3 export_report.py <results.json> [-o <输出目录或xlsx路径>]",
    )
    parser.add_argument("results", help="results.json 路径")
    parser.add_argument(
        "-o", "--output",
        help="输出目录或 .xlsx 文件路径（默认 results.json 同目录）",
    )
    args = parser.parse_args(argv)

    results_path = Path(args.results).expanduser().resolve()
    if not results_path.is_file():
        raise SystemExit(f"错误：{results_path} 不是文件")

    data = load_results(results_path)
    xlsx_path, md_path = derive_output_paths(
        results_path, data["meta"], args.output
    )

    write_excel(data, results_path, xlsx_path)
    write_markdown(data, md_path)

    print(f"已生成 Excel 报告：{xlsx_path}")
    print(f"已生成 Markdown 复核报告：{md_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())