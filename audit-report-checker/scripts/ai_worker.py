#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ai_worker.py — DeepSeek 并发加速层

职责：
- 封装 DeepSeek API 调用（OpenAI 兼容）
- 并发执行 4 个场景：结构标注、文本错别字、二次复核、表注定位
- 降级处理：无 key/失效时不崩溃

设计原则（D1）：
- DeepSeek 做批量并发工作（可并行，每项独立）
- Claude 做全局语义工作（单次，需全局上下文）

安全（D3）：
- API key 不写进 skill 目录
- 检查顺序：--api-key > DEEPSEEK_API_KEY 环境变量 > ~/.deepseek/config.json
"""

import json
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Set
from urllib.parse import urlparse

try:
    import openai
except ImportError:
    raise RuntimeError("请先安装 openai 库: pip install openai")

import openpyxl


# ──────────────────────────────────────────────────────────────────
# DeepSeek Client 封装
# ──────────────────────────────────────────────────────────────────
# 默认配置（model 可被 ~/.deepseek/config.json、--model 参数或运行时选择覆盖）
# DEFAULT_MODEL 优先 flash 类（非 reasoning，单请求延迟低，适合大批量并发标注）
DEFAULT_MODEL = "deepseek-chat"
DEFAULT_BASE_URL = "https://api.deepseek.com"


class DeepSeekClient:
    """DeepSeek API 客户端，OpenAI 兼容接口。"""

    def __init__(self, api_key: str, model: str, base_url: str = "https://api.deepseek.com"):
        self.api_key = api_key
        self.model = model
        self.base_url = base_url

        # 初始化 OpenAI 客户端
        self.client = openai.OpenAI(
            api_key=api_key,
            base_url=base_url,
            timeout=120.0,  # 单请求 2 分钟上限，避免单请求挂起过久（原 600s 会让卡死请求占满线程）
        )

    def chat(self, messages: List[Dict[str, str]], temperature: float = 0.0, max_retries: int = 3) -> str:
        """发送聊天请求，返回文本内容。带重试 + finish_reason 诊断。

        不限制输出长度（不传 max_tokens），让模型完整输出。
        空返回时重试。记录 finish_reason 便于诊断。

        重试策略：
        - 可重试：openai APIConnectionError / APITimeoutError / RateLimitError /
          服务端 5xx / 空返回。
        - 不可重试（直接抛出）：AuthenticationError(401)/PermissionError(403)/
          BadRequestError(400)/其他客户端错误，避免无效退避浪费。
        """
        import time as _time
        from openai import (
            APIConnectionError,
            APITimeoutError,
            RateLimitError,
            AuthenticationError,
            PermissionDeniedError,
            BadRequestError,
        )

        last_reason = ""
        last_exc = None
        for attempt in range(max_retries):
            try:
                kwargs = dict(model=self.model, messages=messages, temperature=temperature)
                response = self.client.chat.completions.create(**kwargs)
                choice = response.choices[0]
                content = choice.message.content or ""
                last_reason = getattr(choice, "finish_reason", "") or ""
                if content.strip():
                    return content
                print(f"[WARN] 空返回(finish_reason={last_reason})，重试 {attempt+1}/{max_retries}", file=sys.stderr)
            except (APIConnectionError, APITimeoutError, RateLimitError) as e:
                # 网络/限流/服务端问题，可重试
                last_exc = e
                print(f"[WARN] API可重试异常({type(e).__name__}: {e})，重试 {attempt+1}/{max_retries}", file=sys.stderr)
            except (AuthenticationError, PermissionDeniedError, BadRequestError) as e:
                # 401/403/400 为永久性客户端错误，不再重试
                raise RuntimeError(
                    f"API 请求被拒绝({type(e).__name__}: {e})，请检查 API key、模型名称和请求参数"
                ) from e
            except Exception as e:
                # 其他未知异常，保守重试一次后抛出
                last_exc = e
                print(f"[WARN] API异常({type(e).__name__}: {e})，重试 {attempt+1}/{max_retries}", file=sys.stderr)
            # 指数退避：503 服务繁忙/限流时避免快速重试撞墙
            if attempt < max_retries - 1:
                _time.sleep(2 ** attempt)
        # 耗尽重试
        raise RuntimeError(f"API 请求多次失败（重试 {max_retries} 次）: {last_exc}")

    def chat_json(self, messages: List[Dict[str, str]], temperature: float = 0.0) -> Dict[str, Any]:
        """发送聊天请求，返回解析后的 JSON。"""
        content = self.chat(messages, temperature)

        # JSON 解析容错：剥离 ```json 包裹
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()

        # 尝试解析 JSON
        try:
            return json.loads(content)
        except json.JSONDecodeError as e:
            # 尝试用 json_repair 修复（如果安装了）
            try:
                import json_repair
                content = json_repair.repair_json(content)
                return json.loads(content)
            except ImportError:
                # json_repair 不可用，返回原始错误
                raise ValueError(f"JSON 解析失败: {e}\n原始内容: {content[:500]}...")
            except Exception as e2:
                raise ValueError(f"JSON 修复失败: {e2}\n原始内容: {content[:500]}...")

    def list_models(self, retries: int = 3) -> List[str]:
        """列出可用模型（flash 优先排序）。

        带重试 + 指数退避：DeepSeek 服务繁忙(503)/网络抖动时容错。
        """
        import time as _time
        last_err = None
        for attempt in range(retries):
            try:
                models = self.client.models.list()
                model_names = [m.id for m in models.data if m.id]

                # flash 类优先排序
                flash_models = sorted([m for m in model_names if "flash" in m.lower()])
                other_models = sorted([m for m in model_names if "flash" not in m.lower()])

                return flash_models + other_models
            except Exception as e:
                last_err = e
                if attempt < retries - 1:
                    _time.sleep(2 ** attempt)
        raise RuntimeError(f"查询模型列表失败（重试 {retries} 次）: {last_err}")


# ──────────────────────────────────────────────────────────────────
# API Key 检查
# ──────────────────────────────────────────────────────────────────
def check_api_key(api_key: Optional[str] = None) -> Tuple[bool, Optional[Dict[str, str]]]:
    """检查 API key 是否存在且有效。

    检查顺序：
    1. api_key 参数
    2. DEEPSEEK_API_KEY 环境变量
    3. ~/.deepseek/config.json

    返回：
        (has_key, config_dict): 有 key 且有效返回 True 和配置，否则返回 False 和 None
    """
    # 1. 检查参数
    if api_key:
        config = {"api_key": api_key, "model": DEFAULT_MODEL, "base_url": DEFAULT_BASE_URL}
    # 2. 检查环境变量
    elif "DEEPSEEK_API_KEY" in os.environ:
        config = {
            "api_key": os.environ["DEEPSEEK_API_KEY"],
            "model": DEFAULT_MODEL,
            "base_url": DEFAULT_BASE_URL,
        }
    # 3. 检查配置文件
    else:
        config_path = Path.home() / ".deepseek" / "config.json"
        if config_path.exists():
            try:
                config = json.loads(config_path.read_text(encoding="utf-8"))
                # 确保有必需字段（model/base_url 缺失时用默认值，不写死具体型号）
                if "api_key" not in config:
                    return False, None
                config.setdefault("model", DEFAULT_MODEL)
                config.setdefault("base_url", DEFAULT_BASE_URL)
            except Exception:
                return False, None
        else:
            return False, None

    # 验证 key 有效性（轻量测试调用）
    try:
        client = DeepSeekClient(
            api_key=config["api_key"],
            model=config["model"],
            base_url=config["base_url"],
        )
        # 调用 list_models 验证
        client.list_models()
        return True, config
    except Exception as e:
        # 区分认证失败(401/403)与服务端暂时不可用(503/超时/连接)：
        # 503/限流是 DeepSeek 服务端算力问题，不应误判 key 失效而禁用 AI
        msg = str(e).lower()
        auth_fail = any(k in msg for k in ("401", "403", "authentication", "unauthorized", "invalid api key"))
        if auth_fail:
            return False, None
        print(f"[WARN] key 验证时服务暂时不可用({e})，假定 key 有效继续：{config.get('base_url')}", file=sys.stderr)
        return True, config


# ──────────────────────────────────────────────────────────────────
# 模型列表查询
# ──────────────────────────────────────────────────────────────────
def list_models(client: DeepSeekClient) -> List[str]:
    """查询可用模型列表（flash 优先排序）。"""
    return client.list_models()


# ──────────────────────────────────────────────────────────────────
# 并发执行器
# ──────────────────────────────────────────────────────────────────
def run_concurrent(
    tasks: List[Tuple[Callable, tuple, dict]],
    max_workers: int = 40,
    timeout: int = 120,
) -> List[Any]:
    """并发执行任务，每个任务 try/except，失败返回 None 不阻塞。

    Args:
        tasks: 任务列表，每个任务是 (func, args, kwargs) 元组
        max_workers: 最大并发数，默认 10
        timeout: 单个任务超时时间（秒），默认 60

    Returns:
        结果列表，与 tasks 顺序对应，失败的任务返回 None
    """
    results = [None] * len(tasks)
    total = len(tasks)
    completed = 0
    succeeded = 0
    failed = 0
    import time as _time
    start_time = _time.time()

    def run_task(index: int, func: Callable, args: tuple, kwargs: dict) -> Any:
        try:
            return func(*args, **kwargs)
        except Exception as e:
            # 失败不阻塞，记录错误但不抛出
            print(f"[WARN] 任务 {index} 失败: {e}", file=sys.stderr)
            return None

    # 手动管理 executor（不用 with）：超时/异常时取消未完成 future 并 shutdown(wait=False)，
    # 避免 ThreadPoolExecutor 退出时 join 卡死底层请求导致进程挂起（孤儿线程由 client timeout 兜底自然结束）
    executor = ThreadPoolExecutor(max_workers=max_workers)
    futures = {
        executor.submit(run_task, idx, func, args, kwargs): idx
        for idx, (func, args, kwargs) in enumerate(tasks)
    }
    try:
        for future in as_completed(futures):
            idx = futures[future]
            ok = False
            try:
                result = future.result(timeout=timeout)
                results[idx] = result
                if result is not None:
                    ok = True
            except TimeoutError:
                print(f"[WARN] 任务 {idx} 超时", file=sys.stderr)
                results[idx] = None
            except Exception as e:
                print(f"[WARN] 任务 {idx} 异常: {e}", file=sys.stderr)
                results[idx] = None

            completed += 1
            if ok:
                succeeded += 1
            else:
                failed += 1
            elapsed = round(_time.time() - start_time, 1)
            # 实时进度（每完成一个打印，大批次时降低频率避免刷屏）
            if total <= 30 or completed % max(1, total // 20) == 0 or completed == total:
                rate = round(completed / elapsed, 1) if elapsed > 0 else 0
                eta = round((total - completed) / rate, 1) if rate > 0 else 0
                print(f"[进度] {completed}/{total} 完成（成功{succeeded} 失败{failed}）| 已用{elapsed}s 速率{rate}/s 预计剩{eta}s", file=sys.stderr, flush=True)
    finally:
        # 取消所有尚未开始的 future，立即返回（不等底层 HTTP 请求；已在线程内的请求由 client timeout 兜底）
        for f in futures:
            f.cancel()
        executor.shutdown(wait=False)

    return results


# ──────────────────────────────────────────────────────────────────
# 场景1：附注表结构标注
# ──────────────────────────────────────────────────────────────────
def _load_structure_annotation_prompt() -> str:
    """加载结构标注 prompt。"""
    script_dir = Path(__file__).parent.parent
    prompt_path = script_dir / "references" / "structure_annotation.md"
    if not prompt_path.exists():
        raise FileNotFoundError(f"结构标注 prompt 文件不存在: {prompt_path}")
    return prompt_path.read_text(encoding="utf-8")


def _table_to_text(table: Dict[str, Any]) -> str:
    """将表格转换为文本格式（用于 prompt）。"""
    headers = table.get("headers", [])
    rows = table.get("rows", [])

    # 表头行
    lines = []
    if headers:
        lines.append(" ".join(str(h) for h in headers))

    # 数据行（带行号）
    for idx, row in enumerate(rows):
        line = f"{idx}: " + " ".join(str(cell) for cell in row)
        lines.append(line)

    return "\n".join(lines)


def _is_useful_table(table: Dict[str, Any]) -> bool:
    """判断表格是否值得标注（D3.2）。

    跳过：
    - 标题表（rows 多为空/纯文字）
    - 纯文字段（无金额列）

    只标注"附注明细表"（有金额列+有合计行）。

    Args:
        table: 表格字典

    Returns:
        是否值得标注
    """
    rows = table.get("rows", [])
    headers = table.get("headers", [])

    if not rows or not headers:
        return False

    # 检查是否有金额列（headers 含 金额/余额/价值/合计/小计 等）
    has_amount_col = any(
        any(kw in str(h) for kw in ["金额", "余额", "价值", "合计", "小计", "元", "万元"])
        for h in headers
    )
    if not has_amount_col:
        return False

    # 检查是否有合计行（行标签含 合计/总计/小计）
    has_total_row = any(
        any(kw in str(row[0]) if row else False for kw in ["合计", "总计", "小计"])
        for row in rows
    )

    # 跳过多币种明细表（明细按币种拆分，合计是人民币合计，标注会误报）
    currencies = ["美元", "人民币", "欧元", "日元", "英镑", "港币", "林吉特", "瑞士法郎",
                  "加元", "澳元", "韩元", "新台币", "新加坡元", "泰铢", "越南盾", "卢布"]
    table_text = " ".join(str(h) for h in headers) + " " + " ".join(
        " ".join(str(c) for c in row) for row in rows[:30]
    )
    found_currencies = [c for c in currencies if c in table_text]
    if len(found_currencies) >= 2:
        return False

    # 检查是否为标题表（大部分行为空或纯文字）
    non_empty_rows = 0
    for row in rows:
        if row:
            # 检查是否有数字（排除纯文字行）
            has_number = any(
                re.search(r"\d", str(cell)) for cell in row
            )
            if has_number:
                non_empty_rows += 1

    # 如果有数字的行数小于总行数的 30%，可能是标题表
    is_title_table = non_empty_rows < len(rows) * 0.3 if rows else True

    return has_amount_col and (has_total_row or not is_title_table)


def annotate_tables(candidates: List[Dict[str, Any]], client: DeepSeekClient) -> List[Dict[str, Any]]:
    """并发标注候选表结构（场景1）。

    改进（D2.3 + D3）：
    1. 跳过无用表：标题表/纯文字段不送标注
    2. 批量化：一次 prompt 处理 5 张表
    3. prompt 强化百分比列区分（已在 structure_annotation.md 中）

    Args:
        candidates: 候选表列表，每项包含 id/page/headers/rows
        client: DeepSeek 客户端

    Returns:
        标注结果列表，每项包含 table_id 和 annotation（rows/columns/horizontal）
    """
    if not candidates:
        return []

    # 行数过滤：只标注 ≤20 行的表（明细表主力），>20 行多为报表本身（已有 L1 覆盖）且 AI 标注易空返回拖慢
    MAX_ROWS = 20
    to_annotate = [c for c in candidates if len(c.get("rows", [])) <= MAX_ROWS]
    skipped_large = [c for c in candidates if len(c.get("rows", [])) > MAX_ROWS]
    if skipped_large:
        print(f"[INFO] 跳过 {len(skipped_large)} 张大表(>{MAX_ROWS}行，由 L1/code_vertical 覆盖)", file=sys.stderr)

    # 跳过无用表（标题表/纯文字段）
    useful_tables = [c for c in to_annotate if _is_useful_table(c)]
    skipped_useless = [c for c in to_annotate if not _is_useful_table(c)]
    if skipped_useless:
        print(f"[INFO] 跳过 {len(skipped_useless)} 张无用表（标题表/纯文字段），标注 {len(useful_tables)} 张", file=sys.stderr)

    candidates = useful_tables
    if not candidates:
        return []

    # 加载 prompt
    prompt_template = _load_structure_annotation_prompt()
    # 提取核心 prompt（第九节）
    core_prompt_match = re.search(r"### 9\.1.*```(.*)```", prompt_template, re.DOTALL)
    if core_prompt_match:
        core_prompt = core_prompt_match.group(1).strip()
    else:
        # 备用：使用全文
        core_prompt = prompt_template

    # 批量大小（默认1张表1个请求，充分利用并发；可用 AUDIT_ANNOTATION_BATCH_SIZE 调整）
    BATCH_SIZE = int(os.environ.get("AUDIT_ANNOTATION_BATCH_SIZE", "1"))

    def annotate_batch(tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """标注一批表格（最多5张）。"""
        # 构建多表文本，用分隔符区分
        separator = "\n\n=== 表格分隔符 ===\n\n"
        table_texts = []
        for idx, table in enumerate(tables):
            table_id = table.get("id", f"table_{idx}")
            table_text = _table_to_text(table)
            table_texts.append(f"表格ID: {table_id}\n{table_text}")

        combined_text = separator.join(table_texts)

        # 构建 prompt（批量化版本，要求返回多表标注）
        batch_prompt = f"""{core_prompt}

**重要**：当前输入包含多张表格（用 "=== 表格分隔符 ===" 分隔），请为每张表返回独立的标注结果。
返回格式：
{{
  "tables": [
    {{"table_id": "表格ID1", "annotation": {{"rows": [...], "columns": [...], "horizontal": [...]}}}},
    {{"table_id": "表格ID2", "annotation": {{"rows": [...], "columns": [...], "horizontal": [...]}}}}
  ]
}}

表格：
{combined_text}
"""

        try:
            result = client.chat_json(
                [{"role": "user", "content": batch_prompt}],
                temperature=0.0,
            )
            return result.get("tables", [])
        except Exception as e:
            print(f"[WARN] 批量标注失败: {e}", file=sys.stderr)
            return []

    # 按批次并发标注（原串行 N 批，单批 API 延迟叠加会超时；改为并发，max_workers 限流避免 DeepSeek 503）
    batches = [candidates[i:i + BATCH_SIZE] for i in range(0, len(candidates), BATCH_SIZE)]
    tasks = [(annotate_batch, (batch,), {}) for batch in batches]
    batch_results_list = run_concurrent(tasks, max_workers=int(os.environ.get("AUDIT_AI_MAX_WORKERS", "8")))
    all_annotations = []
    for batch_results in batch_results_list:
        if batch_results:
            all_annotations.extend(batch_results)

    # 转换为统一格式
    results = []
    for annotation in all_annotations:
        table_id = annotation.get("table_id")
        table_data = annotation.get("annotation")

        if not table_id or not table_data:
            continue

        # 验证返回结构
        if not isinstance(table_data, dict):
            continue
        if "rows" not in table_data or "columns" not in table_data:
            continue

        # 找到对应的 page
        page = None
        for table in candidates:
            if table.get("id") == table_id:
                page = table.get("page")
                break

        results.append({
            "table_id": table_id,
            "page": page,
            "annotation": table_data,
        })

    return results


# ──────────────────────────────────────────────────────────────────
# 审计术语白名单
# ──────────────────────────────────────────────────────────────────
AUDIT_TERM_WHITELIST = [
    # 会计科目
    "金融机构", "所得税", "决策", "所有者权益", "非经常性损益", "关联方", "合并报表",
    "资产负债", "现金流量", "营业收入", "营业成本", "应收账款", "固定资产", "无形资产",
    "长期股权投资", "应付账款", "存货", "资本公积", "盈余公积", "未分配利润",
    "递延所得税", "公允价值", "减值准备", "坏账准备", "累计折旧", "累计摊销",
    "应付职工薪酬", "应收票据", "其他应收款", "其他应付款", "预付账款", "预收账款",
    "投资性房地产", "在建工程", "无形资产", "长期待摊费用", "短期借款", "长期借款",
    "应付债券", "股本", "实收资本", "专项储备", "其他综合收益", "盈余公积",
    "未分配利润", "营业利润", "利润总额", "净利润", "每股收益", "基本每股收益",
    "稀释每股收益", "归属于母公司所有者的净利润", "少数股东损益", "经营活动",
    "投资活动", "筹资活动", "现金及现金等价物净增加额", "期初现金及现金等价物余额",
    "期末现金及现金等价物余额", "流动资产", "非流动资产", "流动负债", "非流动负债",
    "资产总计", "负债合计", "所有者权益合计", "负债和所有者权益总计", "一年内到期的",
    "长期应收款", "长期应付款", "预计负债", "递延收益", "递延所得税资产", "递延所得税负债",
    "应付股利", "应交税费", "其他应付款", "其他流动资产", "其他非流动资产",
    "交易性金融资产", "债权投资", "其他债权投资", "长期股权投资", "投资收益",
    "资产处置收益", "其他收益", "营业外收入", "营业外支出", "所得税费用",
    "持续经营净利润", "终止经营净利润", "综合收益总额", "归属于母公司",
    "少数股东权益", "财务费用", "销售费用", "管理费用", "研发费用", "税金及附加",
    "信用减值损失", "资产减值损失", "净敞口套期收益", "其他权益工具投资",
    "合同资产", "合同负债", "合同取得成本", "合同履约成本", "应收款项融资",
]


# ──────────────────────────────────────────────────────────────────
# PDF空格预处理
# ──────────────────────────────────────────────────────────────────
def _preprocess_pdf_spaces(text: str) -> str:
    """预处理文本，合并中文间的空格。

    检测"中文+空格+中文"模式，合并断词，避免 PDF 提取的空格被当错别字。
    保留中文与数字/英文间的空格（正常分隔）。

    例子：
        "金融机 构" → "金融机构"
        "所得 税" → "所得税"
        "企 业 管 理" → "企业管理"
        "中国 2023年" → "中国 2023年"（保留）

    注意：需要循环替换直到没有中文-空格-中文模式，因为一次替换可能产生新的匹配。
    例如："企 业 管 理" → "企业 管理" → "企业管理"
    """
    pattern = r"([\u4e00-\u9fa5])\s+([\u4e00-\u9fa5])"
    while re.search(pattern, text):
        text = re.sub(pattern, r"\1\2", text)
    return text


# 场景2：文本错别字检查
# ──────────────────────────────────────────────────────────────────
# 错别字+病句检查 prompt
_TYPO_AND_GRAMMAR_CHECK_PROMPT = """你是注册会计师，审阅审计报告文本，找出错别字和病句。

## 错别字定义：
- 明显的拼写错误（如"资负债表"→"资产负债表"）
- 只报告**原文中确实出现的词**，不要联想/推测/编造原文没有的词
- 不是专业术语（如"减值准备""先进先出""月末一次加权平均"都是正确的专业术语，不报告）
- 不是换行导致的拆词（如"利\n润表"不视为错别字）

错别字过滤规则：
- 错别字长度 ≥ 2
- **必须从输入文本中逐字引用原文存在的词**（禁止编造文本里没有的词作为"错别字"）
- 不报告专业术语变体（如"先进先出""个别计价法"是会计术语，正确）
- 不报告因换行导致的拆词
- 原文中同时包含正确的完整词时，不报告（如"利润表"正确时，不报告"润表"）

## 病句定义：
- 语序不当（如"过去的学习方法，对于现在不适用"→"现在的学习方法，对于过去不适用"）
- 搭配不当（如"提高水平"→"提升水平"、"解决问题"→"处理问题"）
- 成分残缺（主语、谓语、宾语缺失，如"通过这次会议，使我们认识到..."缺少主语）
- 句式杂糅（如"原因是...造成的"→"原因是..."或"...是...造成的"）
- 重复冗余（如"大约1000元左右"→"约1000元"或"1000元左右"）
- 不合逻辑（前后矛盾、概念不清）

病句排除规则（严格执行，宁可漏报不可误报）：
- **会计政策与会计估计段落整段不查**：以"本公司""按照""根据"开头，描述金融工具分类/收入确认/减值政策/租赁/套期/公允价值计量等准则应用的段落，是 CAS 准则原文的标准措辞，不要作为病句报告
- 专业会计术语（资产负债表、应收账款坏账准备、预期信用损失、摊余成本等）
- 法规/准则名称全文（如《企业会计准则第14号——收入》是标准全称，不重复、不病句）
- 列表式短句、表格数据行、含大量数字的技术性语句
- **只报告明确无疑的语法错误**（成分严重残缺导致无法理解、明确的语序颠倒），不报告：标点建议、长句难读、措辞偏好、术语用词、风格问题
- 不确定是否病句时 confidence 必须 < 0.5（即不报告）

confidence 字段说明：0-1 的置信度，表示你对这个判断的确信程度。
- confidence >= 0.85：高置信度，确认为错别字/病句
- 0.5 <= confidence < 0.85：中等置信度，可能是错别字/病句
- confidence < 0.5：低置信度，不确定，不要报告

输入文本：
{text}

返回 JSON 格式：
{
  "typos": [
    {"original": "错误词", "suggestion": "正确词", "reason": "原因", "confidence": 0.0-1.0}
  ],
  "grammar_errors": [
    {"original": "原句", "suggestion": "修改建议", "reason": "原因", "confidence": 0.0-1.0}
  ]
}

无错别字则 typos 为空数组，无病句则 grammar_errors 为空数组。直接返回 JSON，不要加 markdown 代码块标记。
"""


def _filter_audit_terms(typos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """过滤审计术语白名单。

    如果错别字的 original 或 suggestion（经预处理后）完全等于白名单术语，则跳过（是术语不是错别字）。

    注意：
    1. 需要先对 original/suggestion 进行 PDF 空格预处理（去除中文间空格）
    2. 使用精确匹配（完全等于），而不是部分匹配（包含）
    3. 例如："金融机 构" preprocessed 后是"金融机构"（在白名单），所以过滤
    4. 但"现金流理表" preprocessed 后不是"现金流量"（不完全等于），所以保留

    Args:
        typos: 错别字列表，每项包含 original/suggestion

    Returns:
        过滤后的错别字列表
    """
    filtered = []
    for typo in typos:
        original = typo.get("original", "")
        suggestion = typo.get("suggestion", "")

        # 先进行 PDF 空格预处理（去除中文间空格）
        original_preprocessed = _preprocess_pdf_spaces(original)
        suggestion_preprocessed = _preprocess_pdf_spaces(suggestion)

        # 检查 preprocessed 后的字符串是否完全等于白名单术语（精确匹配）
        is_term = False
        if original_preprocessed in AUDIT_TERM_WHITELIST:
            # original preprocessed 后是术语，说明是 PDF 断词，不是真错别字
            is_term = True
        elif suggestion_preprocessed in AUDIT_TERM_WHITELIST:
            # suggestion preprocessed 后是术语，说明 DeepSeek 建议的是术语，不是错别字
            is_term = True

        if not is_term:
            filtered.append(typo)

    return filtered


def check_typos_ai(report_chunks: List[str], client: DeepSeekClient) -> Dict[str, List[Dict[str, Any]]]:
    """并发检查文本错别字和病句（场景2）。

    改进（D2.2 + D3）：
    1. PDF空格预处理：合并中文间空格，避免断词被当错别字
    2. 审计术语白名单：过滤术语，避免专业术语被报错
    3. 置信度过滤：confidence < 0.5 丢弃，0.5-0.85 标 warning，>=0.85 标 error
    4. Think Off：temperature=0.1 降低过度报告
    5. 批量化：一次 prompt 检查 5 段文本
    6. 新增病句检查：同时检测语序不当、搭配不当、成分残缺、句式杂糅、重复冗余、不合逻辑

    Args:
        report_chunks: 文本块列表
        client: DeepSeek 客户端

    Returns:
        {"typos": [...], "grammar_errors": [...]}
        typos: 错别字列表，每项包含 original/suggestion/reason/confidence/severity
        grammar_errors: 病句列表，每项包含 original(原句)/suggestion(修改建议)/reason/confidence/severity
    """
    if not report_chunks:
        return {"typos": [], "grammar_errors": []}

    # 批量大小（每次检查 5 段文本）
    BATCH_SIZE = 5

    # 预处理：合并中文间空格
    preprocessed_chunks = [_preprocess_pdf_spaces(chunk) for chunk in report_chunks]

    def check_batch(chunks: List[str]) -> Dict[str, Any]:
        """检查一批文本块（最多5段）。"""
        # 合并多段文本，用分隔符区分
        separator = "\n\n=== 分隔符 ===\n\n"
        combined_text = separator.join(chunks)

        # 构建 prompt（批量化版本）
        prompt = _TYPO_AND_GRAMMAR_CHECK_PROMPT.replace("{text}", combined_text)

        try:
            result = client.chat_json(
                [{"role": "user", "content": prompt}],
                temperature=0.1,  # Think Off：降低温度减少过度报告
            )
            return result
        except Exception as e:
            print(f"[WARN] 文本错别字/病句检查失败: {e}", file=sys.stderr)
            return {"typos": [], "grammar_errors": []}

    # 按批次并发检查（原串行 N 批会累积超时；改为并发，max_workers 限流避免 DeepSeek 503）
    batches = [preprocessed_chunks[i:i + BATCH_SIZE] for i in range(0, len(preprocessed_chunks), BATCH_SIZE)]
    tasks = [(check_batch, (batch,), {}) for batch in batches]
    batch_results_list = run_concurrent(tasks, max_workers=int(os.environ.get("AUDIT_AI_MAX_WORKERS", "8")))

    # 合并所有批次结果
    all_typos = []
    all_grammar_errors = []
    for batch_result in batch_results_list:
        if batch_result:
            all_typos.extend(batch_result.get("typos", []))
            all_grammar_errors.extend(batch_result.get("grammar_errors", []))

    # 过滤审计术语白名单（只对错别字）
    filtered_typos = _filter_audit_terms(all_typos)

    # 根据置信度设置 severity（错别字）
    final_typos = []
    for typo in filtered_typos:
        confidence = typo.get("confidence", 0.0)

        # 过滤低置信度（< 0.5）
        if confidence < 0.5:
            continue

        # 设置 severity
        if confidence >= 0.85:
            typo["severity"] = "error"
        else:
            typo["severity"] = "warning"  # 0.5-0.85 标存疑

        final_typos.append(typo)

    # 根据置信度设置 severity（病句主观性强，统一为 warning）
    final_grammar_errors = []
    for err in all_grammar_errors:
        confidence = err.get("confidence", 0.0)

        # 过滤低置信度（病句误报率高，要求 >= 0.85 高置信度才报告）
        if confidence < 0.85:
            continue

        # 病句统一为 warning（主观性强）
        err["severity"] = "warning"
        final_grammar_errors.append(err)

    # 全局去重：同 (original, suggestion) 的错别字/病句合并为一条，记录出现次数
    # （批处理多 chunk 可能让 DeepSeek 对同一错误重复报告，去重避免噪音）
    def _dedup(items):
        deduped = {}
        for it in items:
            key = (it.get("original", ""), it.get("suggestion", ""))
            if key not in deduped:
                d = it.copy()
                d["count"] = 1
                deduped[key] = d
            else:
                deduped[key]["count"] = deduped[key].get("count", 1) + 1
        return list(deduped.values())

    final_typos = _dedup(final_typos)
    final_grammar_errors = _dedup(final_grammar_errors)

    return {
        "typos": final_typos,
        "grammar_errors": final_grammar_errors
    }


# ──────────────────────────────────────────────────────────────────
# 场景3：warning/error 二次复核
# ──────────────────────────────────────────────────────────────────
# 二次复核 prompt
_REVIEW_PROMPT = """你是注册会计师，复核审计报告检查结果。

检查规则：{rule_name}
规则描述：{description}

源值：{expected}
目标值：{actual}
差异：{difference}

原文摘录：
{context}

**表结构上下文**：
{table_context}

判断此差异是以下哪种情况，返回 JSON：
- confirm: 确认真错（计算错误、加总不平、明显不一致）
- downgrade: 特殊结构（百分比表、子项重复列示、减项处理、不同口径），降为 info
  * 百分比列被当金额加（列名含"%"、"比例"等但参与了加减验算）
  * "其中"子项被重复加（sub_detail 行参与了竖加）
  * 合计行空值被当0验算（合计行无有效数字）
  * 不同时点混合（年末列与年初列混合验算）
- needs_review: 存疑（口径差异、披露不完整），保持 warning

返回格式：
{{"decision": "confirm|downgrade|needs_review", "reason": "原因说明"}}

直接返回 JSON，不要加 markdown 代码块标记。
"""


def review_results(warnings: List[Dict[str, Any]], client: DeepSeekClient) -> List[Dict[str, Any]]:
    """并发复核 warning/error（场景3）。

    改进（D3）：
    1. 批量化：一次 prompt 复核 5-10 条 warning/error

    Args:
        warnings: warning/error 列表
        client: DeepSeek 客户端

    Returns:
        复核结果列表，每项包含 original_index, decision, reason
    """
    if not warnings:
        return []

    # 批量大小（每次复核 10 条）
    BATCH_SIZE = 10

    def build_table_context(warning: Dict[str, Any]) -> str:
        """构建表结构上下文。"""
        table_context_parts = []
        description = warning.get("description", "")
        context = warning.get("context", "")
        evidence = warning.get("evidence", "")

        # 提取列名信息（从 description 或 context）
        if "列" in description or "金额列" in description:
            table_context_parts.append("列类型信息：需判断验算列是否为百分比列")

        # 提取行结构信息（是否含"其中"子项）
        if "其中" in str(context) or "sub_detail" in str(evidence):
            table_context_parts.append("行结构信息：表格包含'其中'子项行，需判断是否重复加总")

        # 提取合计行信息（是否为空值）
        actual = warning.get("actual", "")
        if "合计" in str(context) and (not actual or actual == "0.00"):
            table_context_parts.append("合计行信息：合计行可能为空值，无法验算")

        # 提取百分比表特征（列名含%、比例等）
        if any(kw in str(description).lower() for kw in ["%", "比例", "百分比", "rate"]):
            table_context_parts.append("表类型信息：可能是百分比表，数值应在0-100范围")

        return "\n".join(table_context_parts) if table_context_parts else "无特殊结构信息"

    def review_batch(items: List[Tuple[Dict[str, Any], int]]) -> List[Dict[str, Any]]:
        """复核一批 warning/error（最多10条）。"""
        # 构建多项目文本，用分隔符区分
        separator = "\n\n=== 项目分隔符 ===\n\n"
        item_texts = []

        for warning, idx in items:
            rule_name = warning.get("rule_name", "")
            description = warning.get("description", "")
            expected = warning.get("expected", "")
            actual = warning.get("actual", "")
            difference = warning.get("difference", "")
            context = warning.get("context", "")
            table_context = build_table_context(warning)

            item_text = f"""项目索引: {idx}
检查规则：{rule_name}
规则描述：{description}
源值：{expected}
目标值：{actual}
差异：{difference}
原文摘录：
{context}

**表结构上下文**：
{table_context}
"""
            item_texts.append(item_text)

        combined_text = separator.join(item_texts)

        # 构建批量化 prompt
        batch_prompt = f"""你是注册会计师，复核审计报告检查结果。

**重要**：当前输入包含多个检查项（用 "=== 项目分隔符 ===" 分隔），请为每项返回独立的复核结果。

判断规则：
- confirm: 确认真错（计算错误、加总不平、明显不一致）
- downgrade: 特殊结构（百分比表、子项重复列示、减项处理、不同口径），降为 info
  * 百分比列被当金额加（列名含"%"、"比例"等但参与了加减验算）
  * "其中"子项被重复加（sub_detail 行参与了竖加）
  * 合计行空值被当0验算（合计行无有效数字）
  * 不同时点混合（年末列与年初列混合验算）
- needs_review: 存疑（口径差异、披露不完整），保持 warning

返回格式：
{{
  "reviews": [
    {{"index": 项目索引1, "decision": "confirm|downgrade|needs_review", "reason": "原因说明"}},
    {{"index": 项目索引2, "decision": "confirm|downgrade|needs_review", "reason": "原因说明"}}
  ]
}}

检查项：
{combined_text}
"""

        try:
            result = client.chat_json(
                [{"role": "user", "content": batch_prompt}],
                temperature=0.0,
            )
            return result.get("reviews", [])
        except Exception as e:
            print(f"[WARN] 批量复核失败: {e}", file=sys.stderr)
            return []

    # 按批次并发复核（原串行 N 批会累积超时；改为并发，max_workers 限流避免 DeepSeek 503）
    batches = []
    for i in range(0, len(warnings), BATCH_SIZE):
        batch = [(warnings[idx], idx) for idx in range(i, min(i + BATCH_SIZE, len(warnings)))]
        batches.append(batch)
    tasks = [(review_batch, (batch,), {}) for batch in batches]
    batch_results_list = run_concurrent(tasks, max_workers=int(os.environ.get("AUDIT_AI_MAX_WORKERS", "8")))
    all_reviews = []
    for batch_results in batch_results_list:
        if batch_results:
            all_reviews.extend(batch_results)

    # 转换为统一格式并过滤无效结果
    results = []
    for review in all_reviews:
        idx = review.get("index")
        decision = review.get("decision")
        reason = review.get("reason")

        if idx is None or decision not in ["confirm", "downgrade", "needs_review"]:
            continue

        results.append({
            "original_index": idx,
            "decision": decision,
            "reason": reason,
        })

    return results


# ──────────────────────────────────────────────────────────────────
# 场景4：表注勾稽数值定位
# ──────────────────────────────────────────────────────────────────
# 表注定位 prompt
_LOCATE_NOTE_VALUES_PROMPT = """你是注册会计师，从附注表格中提取科目的汇总数。

科目：{subject}
表格：
{table_text}

**重要：提取的是该科目的合计/总计金额（"合计""小计""总额"行），不是某一账龄段/类别/明细行的金额。若表含多行明细 + 合计行，一律取合计行数值。**

请提取以下数值（如果存在）：
- 账面余额
- 坏账准备/减值准备
- 跌价准备（存货用）
- 账面价值
- 期初余额
- 期末余额
- 本期增加
- 本期减少
- 本期发生额（收入/费用类科目本年累计发生额，如营业收入/营业成本/各项费用）

返回 JSON 格式：
{{"values": {{"账面余额": 数值或null, "坏账准备": 数值或null, "跌价准备": 数值或null, "账面价值": 数值或null, "期初余额": 数值或null, "期末余额": 数值或null, "本期增加": 数值或null, "本期减少": 数值或null, "本期发生额": 数值或null}}}}

直接返回 JSON，不要加 markdown 代码块标记。
"""


def locate_note_values(
    note_map: Dict[str, Dict[str, Any]],
    tables: List[Dict[str, Any]],
    client: DeepSeekClient,
) -> Dict[str, List[Dict[str, Any]]]:
    """并发定位表注勾稽数值（场景4）。

    按 note_map（Claude 语义定位生成的「科目→附注明细表id」映射）精确取数，
    只发该科目的候选附注表（省 token），避免旧式按表名匹配把报表行当附注。

    定位由 Claude 一次性语义完成（note_map），取数由 DeepSeek 并发完成，
    算术由 calculator 比较——三层分工（准确定位 + 高效取数 + 确定性算术）。

    Args:
        note_map: {科目: {"table_ids": [int], "pages": [int], "field": str, "formula": str, ...}}
            由 Claude 在 Step 2 理解附注结构时生成（定位到表级别，非页级别）
        tables: 全部表格列表（按 id 查找候选表）
        client: DeepSeek 客户端

    Returns:
        {科目: [{table_id, page, values}]}，values 含账面余额/坏账准备/账面价值/期初/期末等
    """
    if not note_map or not tables:
        return {}

    tbl_by_id = {t.get("id"): t for t in tables}

    def locate_single(subject: str, table: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """定位单个科目在某张附注表的数值。"""
        table_text = _table_to_text(table)
        prompt = _LOCATE_NOTE_VALUES_PROMPT.format(
            subject=subject, table_text=table_text
        )

        try:
            result = client.chat_json(
                [{"role": "user", "content": prompt}],
                temperature=0.0,
            )

            values = result.get("values", {})
            return {
                "subject": subject,
                "table_id": table.get("id"),
                "page": table.get("page"),
                "source_file": table.get("source_file"),
                "values": values,
            }
        except Exception as e:
            print(f"[WARN] 定位科目 {subject} 失败: {e}", file=sys.stderr)
            return None

    # 按 note_map 精确定位（只发映射到的候选附注表，省 token；不再按表名匹配避免误抓报表行）
    tasks = []
    for subject, info in note_map.items():
        for tid in info.get("table_ids", []):
            t = tbl_by_id.get(tid)
            if t:
                tasks.append((locate_single, (subject, t), {}))

    results = run_concurrent(tasks, max_workers=int(os.environ.get("AUDIT_AI_MAX_WORKERS", "8")))

    # 合并结果（按科目聚合）
    result_map = {}
    for r in results:
        if r is None:
            continue
        subject = r["subject"]
        result_map.setdefault(subject, []).append(r)

    return result_map


# ──────────────────────────────────────────────────────────────────
# 场景5：detect_layout（manifest 自动生成）
# ──────────────────────────────────────────────────────────────────
_DETECT_LAYOUT_PROMPT = """你是审计报告结构识别专家。判断这个Excel sheet的结构：
sheet名: {name}
表头: {headers}
前3行数据: {rows}
返回JSON: {{"role":"BS合并|BS母公司|IS合并|IS母公司|CF合并|CF母公司|权益合并|权益母公司|附注|其他","item_col":0,"current_col":2,"prior_col":3,"kind":"linear|matrix","confidence":0.9}}
- item_col: 项目名所在列(含中文)
- current_col/prior_col: 本期/上期金额列(数字)
- BS续表(负债段) role 标 "BS合并续"/"BS母公司续"
- 权益变动表 kind="matrix"，其他表 kind="linear"
直接返回 JSON，不要加 markdown 代码块标记。"""


def _detect_single_sheet(sheet_name: str, headers: List[str], rows: List[List[Any]], client: DeepSeekClient) -> Optional[Dict[str, Any]]:
    """检测单个 sheet 的结构。"""
    # 取前3行（避免过长）
    sample_rows = rows[:3]
    rows_text = "\n".join("|".join(str(cell) for cell in row) for row in sample_rows)
    headers_text = "|".join(str(h) for h in headers)

    prompt = _DETECT_LAYOUT_PROMPT.format(
        name=sheet_name,
        headers=headers_text,
        rows=rows_text,
    )

    try:
        result = client.chat_json(
            [{"role": "user", "content": prompt}],
            temperature=0.0,
        )
        return result
    except Exception as e:
        print(f"[WARN] sheet '{sheet_name}' 检测失败: {e}", file=sys.stderr)
        return None


def _validate_sheet_detection(sheet_name: str, detection: Dict[str, Any], worksheet: Any) -> Tuple[bool, str]:
    """代码结构性校验：item_col 含中文，amount 列多为数字。"""
    item_col = detection.get("item_col")
    current_col = detection.get("current_col")
    prior_col = detection.get("prior_col")

    if item_col is None or current_col is None or prior_col is None:
        return False, "缺少列索引"

    # 检查 item_col 列是否含中文
    chinese_count = 0
    total_rows = 0
    for row in worksheet.iter_rows(values_only=True):
        if len(row) > item_col:
            cell_val = str(row[item_col]) if row[item_col] else ""
            if cell_val and re.search(r"[\u4e00-\u9fa5]", cell_val):
                chinese_count += 1
            total_rows += 1

    if total_rows > 0 and chinese_count < total_rows * 0.3:
        return False, f"item_col({item_col})中文行不足30%"

    # 检查 current_col/prior_col 列是否多为数字
    def check_digit_col(col_idx, col_name):
        digit_count = 0
        total = 0
        for row in worksheet.iter_rows(values_only=True):
            if len(row) > col_idx:
                cell_val = str(row[col_idx]) if row[col_idx] else ""
                # 检查是否是数字（含千分位、小数、负号等）
                if cell_val and re.search(r"-?\d[\d,.]*\d|-?\d", cell_val):
                    digit_count += 1
                total += 1
        if total > 0 and digit_count < total * 0.3:
            return False, f"{col_name}列数字行不足30%"
        return True, ""

    digit_ok1, digit_msg1 = check_digit_col(current_col, "current_col")
    if not digit_ok1:
        return False, digit_msg1

    digit_ok2, digit_msg2 = check_digit_col(prior_col, "prior_col")
    if not digit_ok2:
        return False, digit_msg2

    return True, ""


def detect_layout(xlsx_path: str, client: DeepSeekClient) -> Dict[str, Any]:
    """自动生成 manifest.json（DeepSeek 并发检测 xlsx sheet 结构）。

    Args:
        xlsx_path: xlsx 文件路径
        client: DeepSeek 客户端

    Returns:
        manifest 草稿：{report_format, files, statements_map}
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX 文件不存在: {xlsx_path}")

    # 读 xlsx
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    sheet_names = wb.sheetnames

    # 并发检测每个 sheet
    tasks = []
    for sn in sheet_names:
        ws = wb[sn]
        # 读取表头（第一行非空）和前3行数据
        headers = []
        first_data_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell else "" for cell in row]
            elif first_data_row is None:
                first_data_row = i
                break

        # 读取前3行数据
        rows_data = []
        if first_data_row is not None:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < first_data_row:
                    continue
                if i >= first_data_row + 3:
                    break
                rows_data.append([cell for cell in row])

        tasks.append((_detect_single_sheet, (sn, headers, rows_data, client), {}))

    sheet_results = run_concurrent(tasks, max_workers=int(os.environ.get("AUDIT_AI_MAX_WORKERS", "8")))

    # 代码校验 + 合并结果
    statements_map = {}
    pending_bs_continuation = None  # BS续表待合并

    for sn, result in zip(sheet_names, sheet_results):
        if not result:
            continue

        ws = wb[sn]
        role = result.get("role", "")
        item_col = result.get("item_col")
        current_col = result.get("current_col")
        prior_col = result.get("prior_col")
        kind = result.get("kind", "linear")
        confidence = result.get("confidence", 0.0)

        # 代码校验
        validate_ok, validate_msg = _validate_sheet_detection(sn, result, ws)
        if not validate_ok:
            print(f"[WARN] sheet '{sn}' 校验失败: {validate_msg}，标记存疑", file=sys.stderr)
            confidence = max(0.0, confidence - 0.3)

        # 跳过低置信度和非报表表
        if confidence < 0.6 or role in ("附注", "其他"):
            continue

        # 映射到 statements_map
        stmt_name = None
        cfg = {
            "source": xlsx_path.name,
            "kind": kind,
            "item_col": item_col,
            "current_col": current_col,
            "prior_col": prior_col,
            "field_aliases": {},
        }

        if role == "BS合并":
            stmt_name = "合并资产负债表"
            cfg["sheets"] = [sn]
        elif role == "BS合并续":
            # BS续表需要合并到主表
            pending_bs_continuation = {
                "sheet": sn,
                "cfg": cfg,
            }
            continue
        elif role == "BS母公司":
            stmt_name = "母公司资产负债表"
            cfg["sheets"] = [sn]
        elif role == "BS母公司续":
            # 母公司BS续表
            if "母公司资产负债表" in statements_map:
                statements_map["母公司资产负债表"]["sheets"].append(sn)
            continue
        elif role == "IS合并":
            stmt_name = "合并利润表"
            cfg["sheets"] = [sn]
        elif role == "IS母公司":
            stmt_name = "母公司利润表"
            cfg["sheets"] = [sn]
        elif role == "CF合并":
            stmt_name = "合并现金流量表"
            cfg["sheets"] = [sn]
        elif role == "CF母公司":
            stmt_name = "母公司现金流量表"
            cfg["sheets"] = [sn]
        elif role == "权益合并":
            stmt_name = "合并所有者权益变动表"
            # 权益变动表是矩阵布局，特殊处理
            cfg["kind"] = "matrix"
            cfg["sheets_main"] = [sn]
            cfg["sheets_prior"] = []  # 待检测续表
            cfg["value_col"] = 2
            cfg.pop("item_col", None)
            cfg.pop("current_col", None)
            cfg.pop("prior_col", None)
        elif role == "权益母公司":
            stmt_name = "母公司所有者权益变动表"
            cfg["kind"] = "matrix"
            cfg["sheets_main"] = [sn]
            cfg["sheets_prior"] = []
            cfg["value_col"] = 2
            cfg.pop("item_col", None)
            cfg.pop("current_col", None)
            cfg.pop("prior_col", None)

        if stmt_name:
            statements_map[stmt_name] = cfg

    # 合并 BS 续表
    if pending_bs_continuation:
        main_stmt = statements_map.get("合并资产负债表")
        if main_stmt:
            main_stmt["sheets"].append(pending_bs_continuation["sheet"])

    # 生成 manifest 草稿
    manifest = {
        "report_format": "split_xlsx",
        "files": {
            xlsx_path.name: {"role": "statements", "loader": "xlsx"},
        },
        "statements_map": statements_map,
    }

    return manifest


# ──────────────────────────────────────────────────────────────────
# 场景5.4：detect_docx_layout_code（合并式 Word/PDF 四表代码识别）
# ──────────────────────────────────────────────────────────────────
def _table_text(table: Dict[str, Any], max_rows: int = 9999) -> str:
    rows = table.get("rows", [])
    return " ".join(" ".join(str(c) for c in r) for r in rows[:max_rows])


def _has_any(text: str, keywords: List[str]) -> bool:
    return any(kw in text for kw in keywords)


def _has_all_like(text: str, keyword_groups: List[List[str]]) -> bool:
    """每组关键字只要命中一个即可（处理括弧变体）。"""
    for group in keyword_groups:
        if not _has_any(text, group):
            return False
    return True


def _headers_year_columns(headers: List[str]) -> Tuple[Optional[int], Optional[int]]:
    """根据表头文本推断 current/prior 列索引。"""
    current_col = None
    prior_col = None
    for idx, h in enumerate(headers):
        s = str(h)
        if current_col is None and any(k in s for k in ("期末", "本期", "本年", "当年", "2025")):
            current_col = idx
        if prior_col is None and any(k in s for k in ("期初", "上期", "上年", "年初", "2024")):
            prior_col = idx
    # 如果只有年份，优先把较大年份当 current
    if current_col is None and prior_col is None:
        years = {}
        for idx, h in enumerate(headers):
            m = re.search(r"20(\d{2})", str(h))
            if m:
                years[idx] = int(m.group(0))
        if years:
            sorted_years = sorted(years.items(), key=lambda x: x[1])
            current_col = sorted_years[-1][0]
            if len(sorted_years) > 1:
                prior_col = sorted_years[-2][0]
    return current_col, prior_col


def _numeric_rich_columns(rows: List[List[Any]], headers: List[str]) -> List[int]:
    """返回按数字含量排序的列索引（排除第一列）。"""
    counts = []
    for col in range(1, len(headers)):
        digit = 0
        total = 0
        for row in rows:
            if col < len(row):
                total += 1
                val = str(row[col]).strip()
                if val and re.search(r"-?\d[\d,.]*", val):
                    digit += 1
        if total > 0:
            counts.append((col, digit / total))
    counts.sort(key=lambda x: x[1], reverse=True)
    return [c[0] for c in counts]


def _infer_linear_cols(table: Dict[str, Any]) -> Tuple[int, int, int]:
    headers = [str(h) for h in table.get("headers", [])]
    rows = table.get("rows", [])
    item_col = 0
    current_col, prior_col = _headers_year_columns(headers)
    if current_col is None or prior_col is None:
        numeric_cols = _numeric_rich_columns(rows, headers)
        if current_col is None and numeric_cols:
            current_col = numeric_cols[0]
        if prior_col is None and len(numeric_cols) > 1:
            prior_col = numeric_cols[1]
    # 保底
    if current_col is None:
        current_col = 1 if len(headers) > 1 else 0
    if prior_col is None:
        prior_col = current_col + 1 if current_col + 1 < len(headers) else current_col
    return item_col, current_col, prior_col


def _score_table(table: Dict[str, Any]) -> Dict[str, float]:
    text = _table_text(table)
    headers_text = " ".join(str(h) for h in table.get("headers", []))
    scores = {}
    # BS：必须同时有资产总计 + 负债/权益总计（支持括弧变体）
    bs_groups = [["资产总计"], ["负债合计", "负债和", "负债及"], ["所有者权益合计", "股东权益合计", "所有者权益（或股东权益）合计", "股东权益（或所有者权益）合计"]]
    if _has_all_like(text, bs_groups):
        scores["BS"] = 1.0
    elif _is_complete_bs_table(table):
        scores["BS"] = 0.8
    elif _is_assets_table(table) or _is_liabilities_equity_table(table):
        scores["BS"] = 0.5
    else:
        scores["BS"] = 0.0

    # PL
    pl_kws = ["营业收入", "营业成本", "营业利润", "利润总额", "净利润"]
    pl_count = sum(1 for k in pl_kws if k in text)
    scores["PL"] = pl_count / len(pl_kws)

    # CF
    cf_kws = ["经营活动产生的现金流量", "投资活动产生的现金流量", "筹资活动产生的现金流量", "现金及现金等价物净增加额"]
    cf_count = sum(1 for k in cf_kws if k in text)
    scores["CF"] = cf_count / len(cf_kws)

    # 权益 matrix
    equity_kws = ["股本", "资本公积", "盈余公积", "未分配利润", "所有者权益合计", "股东权益合计"]
    equity_count = sum(1 for k in equity_kws if k in text)
    equity_score = equity_count / len(equity_kws)
    # 矩阵表头特征：年份/本期/上期重复多次
    year_repeat = len(re.findall(r"20\d{2}|本期金额|上期金额|本年金额|上年金额", headers_text))
    if year_repeat >= 4 and equity_score >= 0.6:
        scores["EQUITY_MATRIX"] = equity_score
    else:
        scores["EQUITY_MATRIX"] = 0.0

    return scores


def _is_assets_table(table: Dict[str, Any]) -> bool:
    text = _table_text(table)
    return "资产总计" in text and not _has_any(text, ["负债合计", "所有者权益合计", "股东权益合计", "负债和所有者权益总计", "负债及所有者权益总计"])


def _is_liabilities_equity_table(table: Dict[str, Any]) -> bool:
    text = _table_text(table)
    return _has_any(text, ["负债合计", "所有者权益合计", "股东权益合计", "负债和所有者权益总计", "负债及所有者权益总计"]) and "资产总计" not in text


def _is_complete_bs_table(table: Dict[str, Any]) -> bool:
    text = _table_text(table)
    return "资产总计" in text and _has_any(text, ["负债合计", "所有者权益合计", "股东权益合计", "负债和所有者权益总计", "负债及所有者权益总计"])


def _matrix_value_col(table: Dict[str, Any]) -> int:
    headers = [str(h) for h in table.get("headers", [])]
    rows = table.get("rows", [])
    # 找最后一个含"合计"的列
    for idx in range(len(headers) - 1, -1, -1):
        if "合计" in headers[idx]:
            return idx
    # fallback：从右往左找第一个有数字的列
    for idx in range(len(headers) - 1, 0, -1):
        for row in rows:
            if idx < len(row) and re.search(r"-?\d[\d,.]*", str(row[idx])):
                return idx
    return len(headers) - 1 if len(headers) > 1 else 0


def _matrix_year(table: Dict[str, Any]) -> str:
    headers = [str(h) for h in table.get("headers", [])]
    for h in headers:
        m = re.search(r"20\d{2}", h)
        if m:
            return m.group(0)
        if "本期" in h or "本年" in h:
            return "current"
        if "上期" in h or "上年" in h:
            return "prior"
    return ""


def detect_docx_layout_code(extracted_tables_path: str) -> Dict[str, Any]:
    """用代码启发式识别合并式 Word/PDF 报告中的四表。

    返回 manifest 草稿；识别失败返回空 dict。
    """
    extracted_path = Path(extracted_tables_path)
    if not extracted_path.exists():
        return {}
    with extracted_path.open(encoding="utf-8") as f:
        data = json.load(f)
    tables = data.get("tables", [])
    candidates = [t for t in tables if len(t.get("rows", [])) >= 10 and t.get("headers")]
    if not candidates:
        return {}

    source_file = candidates[0].get("source_file", "")

    # 1. 打分
    scored = []
    for t in candidates:
        scores = _score_table(t)
        scored.append((t, scores))

    # 2. 识别 BS（支持左右分表：资产表 + 负债权益表）
    bs_tables = [(t, s) for t, s in scored if s.get("BS", 0) > 0]
    bs_groups = []  # [(merged?, [table_ids])]
    i = 0
    used_ids = set()
    while i < len(bs_tables):
        t, s = bs_tables[i]
        tid = t["id"]
        if tid in used_ids:
            i += 1
            continue
        if _is_complete_bs_table(t):
            bs_groups.append(([tid]))
            used_ids.add(tid)
            i += 1
        elif _is_assets_table(t) and i + 1 < len(bs_tables):
            next_t, next_s = bs_tables[i + 1]
            if _is_liabilities_equity_table(next_t):
                # 检查表头是否一致
                h1 = [str(h) for h in t.get("headers", [])]
                h2 = [str(h) for h in next_t.get("headers", [])]
                if len(h1) == len(h2):
                    bs_groups.append(([tid, next_t["id"]]))
                    used_ids.add(tid)
                    used_ids.add(next_t["id"])
                    i += 2
                    continue
            i += 1
        else:
            i += 1

    # 3. 识别 PL / CF（取最高分的两张表）
    def pick_top2(role: str) -> List[List[int]]:
        items = [(t, s.get(role, 0)) for t, s in scored if s.get(role, 0) >= 0.5]
        items.sort(key=lambda x: x[1], reverse=True)
        # 去重并保留顺序
        chosen = []
        seen = set()
        for t, sc in items:
            tid = t["id"]
            if tid in seen:
                continue
            seen.add(tid)
            chosen.append([tid])
            if len(chosen) >= 2:
                break
        return chosen

    pl_groups = pick_top2("PL")
    cf_groups = pick_top2("CF")

    # 4. 识别权益 matrix 并配对 current/prior
    matrix_tables = [(t, s) for t, s in scored if s.get("EQUITY_MATRIX", 0) > 0]
    # 按列数分组，组内按年份分 current/prior
    by_cols: Dict[int, List[Dict[str, Any]]] = {}
    for t, s in matrix_tables:
        cols = len(t.get("headers", []))
        by_cols.setdefault(cols, []).append(t)

    equity_pairs = []  # [(main_tid, prior_tid), ...]
    for cols, group in by_cols.items():
        if len(group) < 2:
            continue
        currents = [t for t in group if _matrix_year(t) in ("current", "", "2025")]
        priors = [t for t in group if _matrix_year(t) == "prior" or _matrix_year(t) == "2024"]
        # 按 id 排序，假设 current 在 prior 前
        if not currents and not priors:
            sorted_group = sorted(group, key=lambda x: x["id"])
            mid = len(sorted_group) // 2
            currents = sorted_group[:mid]
            priors = sorted_group[mid:]
        # 配对：current 与 prior 数量相同时按 id 排序一一对应
        currents_sorted = sorted(currents, key=lambda x: x["id"])
        priors_sorted = sorted(priors, key=lambda x: x["id"])
        for cur, pri in zip(currents_sorted, priors_sorted):
            equity_pairs.append((cur["id"], pri["id"]))

    # 5. 区分合并/母公司：默认按出现顺序，BS 含少数股东权益为合并
    def split_merged_parent(groups: List[List[int]], merged_indicator: callable = None) -> Tuple[Optional[List[int]], Optional[List[int]]]:
        if not groups:
            return None, None
        if len(groups) == 1:
            return groups[0], None
        # 两张时，按 merged_indicator 判断；否则第一张为合并
        if merged_indicator:
            merged_idx = 0 if merged_indicator(groups[0]) else 1
            parent_idx = 1 - merged_idx
            return groups[merged_idx], groups[parent_idx]
        return groups[0], groups[1]

    def bs_has_minority(ids: List[int]) -> bool:
        for tid in ids:
            t = next((x for x in candidates if x["id"] == tid), None)
            if t and "少数股东权益" in _table_text(t):
                return True
        return False

    bs_merged, bs_parent = split_merged_parent(bs_groups, merged_indicator=bs_has_minority)
    pl_merged, pl_parent = split_merged_parent(pl_groups)
    cf_merged, cf_parent = split_merged_parent(cf_groups)
    equity_merged, equity_parent = (equity_pairs[0] if equity_pairs else (None, None), equity_pairs[1] if len(equity_pairs) > 1 else (None, None))

    statements_map: Dict[str, Any] = {}

    def add_linear(name: str, ids: Optional[List[int]]):
        if not ids:
            return
        t0 = next((x for x in candidates if x["id"] == ids[0]), None)
        if not t0:
            return
        item_col, current_col, prior_col = _infer_linear_cols(t0)
        statements_map[name] = {
            "source": source_file,
            "loader": "docx",
            "sheets": ids,
            "kind": "linear",
            "item_col": item_col,
            "current_col": current_col,
            "prior_col": prior_col,
            "field_aliases": {},
        }

    def add_matrix(name: str, pair: Tuple[Optional[int], Optional[int]]):
        main_id, prior_id = pair
        if not main_id:
            return
        main_t = next((x for x in candidates if x["id"] == main_id), None)
        if not main_t:
            return
        value_col = _matrix_value_col(main_t)
        statements_map[name] = {
            "source": source_file,
            "loader": "docx",
            "kind": "matrix",
            "sheets_main": [main_id],
            "sheets_prior": [prior_id] if prior_id else [],
            "value_col": value_col,
            "field_aliases": {},
        }

    add_linear("合并资产负债表", bs_merged)
    add_linear("母公司资产负债表", bs_parent)
    add_linear("合并利润表", pl_merged)
    add_linear("母公司利润表", pl_parent)
    add_linear("合并现金流量表", cf_merged)
    add_linear("母公司现金流量表", cf_parent)
    add_matrix("合并所有者权益变动表", equity_merged)
    add_matrix("母公司所有者权益变动表", equity_parent)

    if not statements_map:
        return {}

    return {
        "report_format": "merged_docx",
        "files": {source_file: {"role": "statements", "loader": "docx"}},
        "statements_map": statements_map,
    }


# ──────────────────────────────────────────────────────────────────
# 场景5.5：detect_docx_layout（合并式 Word/PDF 报告自动识别四表，DeepSeek 兜底）
# ──────────────────────────────────────────────────────────────────
def _extract_year_from_headers(headers: List[str]) -> Optional[int]:
    """从表头文本中提取年份（如 2025/2024）。"""
    for h in headers:
        m = re.search(r"20(\d{2})", str(h))
        if m:
            return int(m.group(0))
    return None


def detect_docx_layout(extracted_tables_path: str, client: DeepSeekClient) -> Dict[str, Any]:
    """从 extracted_tables.json 自动识别合并/母公司四表，生成 docx/pdf 版 manifest。

    适用于四表内嵌在 Word/PDF 正文中的合并式报告（无单独 xlsx）。
    """
    extracted_path = Path(extracted_tables_path)
    if not extracted_path.exists():
        raise FileNotFoundError(f"extracted_tables.json 不存在: {extracted_tables_path}")

    with extracted_path.open(encoding="utf-8") as f:
        data = json.load(f)
    tables = data.get("tables", [])
    # 候选表：至少 10 行且有表头
    candidates = [t for t in tables if len(t.get("rows", [])) >= 10 and t.get("headers")]
    if not candidates:
        return {}

    source_file = candidates[0].get("source_file", "")

    # 并发检测每个候选表
    tasks = []
    for t in candidates:
        headers = [str(h) for h in t["headers"]]
        rows = t.get("rows", [])[:3]
        tasks.append((_detect_single_sheet, (f"表ID_{t['id']}", headers, rows, client), {}))
    results = run_concurrent(tasks, max_workers=int(os.environ.get("AUDIT_AI_MAX_WORKERS", "8")))

    statements_map: Dict[str, Any] = {}
    pending_bs_cont: Dict[str, List[int]] = {}

    for t, res in zip(candidates, results):
        if not res:
            continue
        role = res.get("role", "")
        confidence = res.get("confidence", 0.0)
        if confidence < 0.6 or role in ("附注", "其他"):
            continue

        tid = t["id"]
        cfg: Dict[str, Any] = {
            "source": t.get("source_file", source_file),
            "loader": "docx",
            "sheets": [tid],
            "kind": res.get("kind", "linear"),
            "item_col": res.get("item_col"),
            "current_col": res.get("current_col"),
            "prior_col": res.get("prior_col"),
            "field_aliases": {},
        }

        stmt_name = None
        if role == "BS合并":
            stmt_name = "合并资产负债表"
        elif role == "BS合并续":
            pending_bs_cont.setdefault("合并资产负债表", []).append(tid)
            continue
        elif role == "BS母公司":
            stmt_name = "母公司资产负债表"
        elif role == "BS母公司续":
            pending_bs_cont.setdefault("母公司资产负债表", []).append(tid)
            continue
        elif role == "IS合并":
            stmt_name = "合并利润表"
        elif role == "IS母公司":
            stmt_name = "母公司利润表"
        elif role == "CF合并":
            stmt_name = "合并现金流量表"
        elif role == "CF母公司":
            stmt_name = "母公司现金流量表"
        elif role == "权益合并":
            stmt_name = "合并所有者权益变动表"
            cfg["kind"] = "matrix"
            cfg["sheets_main"] = [tid]
            cfg["sheets_prior"] = []
            cfg["value_col"] = 2
            cfg.pop("item_col", None)
            cfg.pop("current_col", None)
            cfg.pop("prior_col", None)
        elif role == "权益母公司":
            stmt_name = "母公司所有者权益变动表"
            cfg["kind"] = "matrix"
            cfg["sheets_main"] = [tid]
            cfg["sheets_prior"] = []
            cfg["value_col"] = 2
            cfg.pop("item_col", None)
            cfg.pop("current_col", None)
            cfg.pop("prior_col", None)

        if stmt_name:
            statements_map[stmt_name] = cfg

    # 合并 BS 续表：把 sheets 改成 [主表id, 续表id, ...]
    for stmt_name, cont_ids in pending_bs_cont.items():
        if stmt_name in statements_map:
            main_id = statements_map[stmt_name]["sheets"][0]
            statements_map[stmt_name]["sheets"] = [main_id] + cont_ids

    # 权益变动表 matrix 配对：按表头年份找上年续表
    for stmt_name, cfg in statements_map.items():
        if cfg.get("kind") != "matrix" or not cfg.get("sheets_main"):
            continue
        main_tid = cfg["sheets_main"][0]
        main_table = next((t for t in candidates if t.get("id") == main_tid), None)
        if not main_table:
            continue
        main_year = _extract_year_from_headers(main_table.get("headers", []))
        main_cols = len(main_table.get("headers", []))
        for t in candidates:
            if t.get("id") == main_tid:
                continue
            headers = t.get("headers", [])
            t_year = _extract_year_from_headers(headers)
            if not t_year or (main_year and t_year >= main_year):
                continue
            if abs(len(headers) - main_cols) > 2:
                continue
            text = " ".join(str(c) for r in t.get("rows", []) for c in r)
            if "股本" in text and ("所有者权益合计" in text or "股东权益合计" in text):
                cfg["sheets_prior"] = [t["id"]]
                break

    manifest = {
        "report_format": "merged_docx",
        "files": {source_file: {"role": "statements", "loader": "docx"}},
        "statements_map": statements_map,
    }
    return manifest


# ──────────────────────────────────────────────────────────────────
# 场景6：build_note_map（note_map 自动生成 + 回原文校验）
# ──────────────────────────────────────────────────────────────────
_BUILD_NOTE_MAP_PROMPT = """你是注册会计师，判断这张附注表归属哪些报表科目。

附注表内容（表头+前5行）:
{table_content}

候选报表科目: {candidates}

返回JSON: {{"mappings":[{{"科目":"科目名","col":1,"field":"年末余额|期末余额|账面余额|账面价值|本年发生额|合计","formula":"","confidence":0.9}}, ...]}}
- 科目: 从候选科目中选择最匹配的（必须是候选列表中的科目名）
- col: 该科目取值所在的列索引（0-based）。普通表（项目在col0/金额在col1）返回 col=1；matrix 表（多金额列）每个科目返回各自 col。
- field: 取值字段（年末余额/期末余额/账面余额/账面价值/本年发生额/合计等）
- formula: 运算公式（如"账面余额-坏账准备"），无需运算则留空
- confidence: 置信度 0-1
**重要**：如果表有多个金额列分别对应不同科目（如营业收入/营业成本分列），为每个科目返回一个 mapping，col 填该科目金额所在列索引。普通明细表返回 1 个 mapping，matrix 表返回多个。
直接返回 JSON，不要加 markdown 代码块标记。"""


def _parse_number(value: Any) -> Optional[float]:
    """解析数字（支持千分位、括号负数等）。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s or s in ("-", "—", "–", "－", "_"):
        return None

    # 括号负数
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
        negative = True
    else:
        negative = False

    # 移除千分位逗号
    s = s.replace(",", "").replace(" ", "")
    # 全角转半角
    s = s.replace("，", ",").replace("（", "(").replace("）", ")")

    # 提取数字
    match = re.search(r"-?\d+\.?\d*", s)
    if not match:
        return None

    try:
        result = float(match.group())
        return -result if negative else result
    except ValueError:
        return None


def _get_table_total(table: Dict[str, Any], col: Optional[int] = None) -> Optional[float]:
    """获取表的合计行值。

    Args:
        table: 表格数据
        col: 指定列索引（0-based）。如果提供，取该列的合计值；否则取第一个数值列。

    Returns:
        合计行的值，失败返回 None
    """
    rows = table.get("rows", [])
    headers = table.get("headers", [])

    # 找合计行（包含"合计"/"总计"/"小计"的第一列）
    total_row = None
    for row in rows:
        if not row:
            continue
        first_col = str(row[0]).strip()
        if any(kw in first_col for kw in ["合计", "总计", "小计"]):
            total_row = row
            break

    if total_row and col is not None:
        # 有 col：找合计行，取该 col 的值
        if col < len(total_row):
            return _parse_number(total_row[col])
        return None

    if total_row:
        # 无 col：从合计行取值（找第一个数值列）
        for col_idx, cell in enumerate(total_row):
            val = _parse_number(cell)
            if val is not None:
                return val
        return None

    # 没有合计行，尝试求和所有数值列
    if col is not None:
        # 求和指定列
        values = []
        for row in rows:
            if len(row) > col:
                val = _parse_number(row[col])
                if val is not None:
                    values.append(val)
        return sum(values) if values else None
    else:
        # 求和第一个数值列
        for col_idx in range(len(headers)):
            values = []
            for row in rows:
                if len(row) > col_idx:
                    val = _parse_number(row[col_idx])
                    if val is not None:
                        values.append(val)
            if values:
                return sum(values)
        return None


def _get_statement_value(statements: Dict[str, Any], subject: str) -> Optional[float]:
    """从 statements 获取科目值（优先合并报表的 current 值）。

    支持常见变体："应收账款净值" → 先找精确名，找不到再找 "应收账款"。
    """
    if not statements:
        return None

    def _try_keys(stmt: dict, keys: list) -> Optional[float]:
        for k in keys:
            val = stmt.get(k)
            if val and isinstance(val, dict):
                cur = val.get("current")
                if cur:
                    return _parse_number(cur)
        return None

    # 构造候选键：精确名 + 去掉后缀的裸名
    keys = [subject]
    for suffix in ["净值", "净额", "账面价值", "账面余额", "原值"]:
        if subject.endswith(suffix):
            keys.append(subject[:-len(suffix)])
            break

    # 优先合并报表
    for stmt_name, stmt in statements.items():
        if "合并" not in stmt_name:
            continue
        if not isinstance(stmt, dict):
            continue
        v = _try_keys(stmt, keys)
        if v is not None:
            return v

    # 降级母公司报表
    for stmt_name, stmt in statements.items():
        if "合并" in stmt_name:
            continue
        if not isinstance(stmt, dict):
            continue
        v = _try_keys(stmt, keys)
        if v is not None:
            return v

    return None


def _check_table_contains_multiple_statements(table: Dict[str, Any], statement_subjects: Set[str]) -> bool:
    """检查表是否包含多个报表科目（排除风险汇总表）。"""
    if not statement_subjects:
        return False

    rows = table.get("rows", [])
    # 检查前10行（避免太长）
    sample_rows = rows[:10]

    match_count = 0
    for row in sample_rows:
        if not row:
            continue
        row_text = " ".join(str(cell) for cell in row)
        # 检查是否包含任一科目名
        for subject in statement_subjects:
            if subject in row_text:
                match_count += 1
                break

    # 如果匹配超过2个科目，视为汇总表
    return match_count > 2


def _build_note_mappings(
    table: Dict[str, Any],
    candidates: List[str],
    client: DeepSeekClient,
) -> Optional[List[Dict[str, Any]]]:
    """构建单个表的 note_map 条目（支持多科目）。

    Returns:
        mappings 列表，每个 mapping 含 {科目, col, field, formula, confidence}
    """
    # 构建表内容文本
    headers = table.get("headers", [])
    rows = table.get("rows", [])
    sample_rows = rows[:5]  # 前5行

    table_text = "表头: " + "|".join(str(h) for h in headers) + "\n"
    table_text += "\n".join("|".join(str(cell) for cell in row) for row in sample_rows)

    prompt = _BUILD_NOTE_MAP_PROMPT.format(
        table_content=table_text,
        candidates="|".join(candidates),
    )

    try:
        result = client.chat_json(
            [{"role": "user", "content": prompt}],
            temperature=0.0,
        )
        return result.get("mappings", [])
    except Exception as e:
        print(f"[WARN] 表ID={table.get('id')} 检测失败: {e}", file=sys.stderr)
        return None


def build_note_map(
    tables: List[Dict[str, Any]],
    statements: Dict[str, Any],
    client: DeepSeekClient,
) -> Dict[str, Dict[str, Any]]:
    """自动生成 note_map.json（DeepSeek 并发 + 回原文校验降错）。

    Args:
        tables: 附注表列表（extracted_tables.json 中的 tables）
        statements: 报表数据（statements.json）
        client: DeepSeek 客户端

    Returns:
        note_map: {科目: {"table_ids": [], "field": "", "formula": "", "confidence": 0.0}}
    """
    if not tables:
        return {}

    # 提取报表科目列表（候选），并扩展常见变体
    statement_subjects = set()
    for stmt_name, stmt in statements.items():
        if not isinstance(stmt, dict):
            continue
        for item in stmt.keys():
            # 过滤掉合计/总计类项目
            if not any(kw in item for kw in ["合计", "总计", "小计", "行次"]):
                statement_subjects.add(item)

    # 扩展科目别名："应收账款净值" → 同时尝试 "应收账款"
    _SUBJECT_SUFFIXES = ["净值", "净额", "账面价值", "账面余额", "原值"]
    expanded_subjects = set(statement_subjects)
    for subj in list(statement_subjects):
        for suffix in _SUBJECT_SUFFIXES:
            if subj.endswith(suffix):
                expanded_subjects.add(subj[:-len(suffix)])
                break
    candidates = sorted(expanded_subjects)
    print(f"[INFO] 报表科目候选: {len(candidates)} 个", file=sys.stderr)

    # 过滤四表（只处理附注明细表）
    # 四表特征：headers 含"项目"/"期末余额"/"年初余额"等关键字
    note_tables = []
    for t in tables:
        headers = t.get("headers", [])
        headers_text = " ".join(str(h) for h in headers)
        # 排除四表（含"期末余额"/"年初余额"且多行的）
        if ("期末余额" in headers_text or "年初余额" in headers_text) and len(t.get("rows", [])) > 20:
            continue
        # 排除无金额列的表（允许"合计"/"小计"/"总计"作为金额列标识，支持 matrix 表如营业收入+营业成本）
        if not any(kw in headers_text for kw in ["金额", "余额", "价值", "元", "万元", "合计", "小计", "总计"]):
            continue
        note_tables.append(t)

    print(f"[INFO] 候选附注表: {len(note_tables)} 张", file=sys.stderr)

    # 并发检测每张附注表
    tasks = [(_build_note_mappings, (t, candidates, client), {}) for t in note_tables]
    table_results = run_concurrent(tasks, max_workers=int(os.environ.get("AUDIT_AI_MAX_WORKERS", "8")))

    # 构建初步 note_map（支持一表多科目）
    preliminary_map: Dict[str, List[Dict[str, Any]]] = {}
    for table, mappings in zip(note_tables, table_results):
        if not mappings:
            continue

        # mappings 是列表，每个元素包含 {科目, col, field, formula, confidence}
        for mapping in mappings:
            subject = mapping.get("科目")
            col = mapping.get("col")
            confidence = mapping.get("confidence", 0.0)
            field = mapping.get("field", "")
            formula = mapping.get("formula", "")

            if not subject or confidence < 0.5:
                continue

            preliminary_map.setdefault(subject, []).append({
                "table_id": table.get("id"),
                "table": table,
                "col": col,
                "field": field,
                "formula": formula,
                "confidence": confidence,
            })

    print(f"[INFO] 初步映射: {len(preliminary_map)} 个科目", file=sys.stderr)

    # 回原文校验（关键降错，支持 fallback 到下一候选）
    final_note_map = {}
    for subject, candidates_list in preliminary_map.items():
        # 按置信度排序
        candidates_list.sort(key=lambda x: x["confidence"], reverse=True)

        chosen = None
        for cand in candidates_list:
            table = cand["table"]
            table_id = cand["table_id"]
            col = cand.get("col")  # 获取列索引

            # 1. 排除汇总表 → 试下一个候选
            is_summary = _check_table_contains_multiple_statements(table, statement_subjects)
            if is_summary:
                print(f"[WARN] {subject} → ID{table_id} 排除（汇总表），尝试下一个候选", file=sys.stderr)
                continue

            # 2. 值校验（最可靠，优先）
            table_total = _get_table_total(table, col=col)
            statement_value = _get_statement_value(statements, subject)

            value_match = False
            if table_total is not None and statement_value is not None:
                # 允许 5% 差异
                diff_ratio = abs(table_total - statement_value) / max(abs(statement_value), 1.0)
                if diff_ratio <= 0.05:
                    value_match = True
                    col_info = f"col={col}" if col is not None else ""
                    print(f"[INFO] {subject} → ID{table_id} {col_info} 值校验通过（{table_total:.2f} ≈ {statement_value:.2f}）", file=sys.stderr)

            # 3. 名称校验（辅助）
            name_match = False
            table_text = " ".join(" ".join(str(cell) for cell in row) for row in table.get("rows", []))
            if subject in table_text:
                name_match = True

            # 4. 综合判断
            final_confidence = cand["confidence"]
            if value_match:
                # 值匹配，高置信度
                final_confidence = min(0.95, final_confidence + 0.2)
            elif name_match:
                # 名称匹配，中等置信度
                final_confidence = min(0.8, final_confidence + 0.1)
            else:
                # 值不匹配且名称不匹配，低置信度（标记存疑）
                final_confidence = max(0.4, final_confidence - 0.3)
                print(f"[WARN] {subject} → ID{table_id} 校验失败（值不匹配且名称不匹配），标记存疑", file=sys.stderr)

            # 5. 检查置信度阈值
            if final_confidence >= 0.6:
                chosen = (cand, final_confidence)
                break  # 找到通过校验的，停止
            else:
                print(f"[WARN] {subject} → ID{table_id} 校验失败(conf={final_confidence:.2f})，尝试下一个候选", file=sys.stderr)
                continue  # 试下一个候选

        # 生成 note_map 条目
        if chosen:
            cand, fc = chosen
            table_id = cand["table_id"]
            final_note_map[subject] = {
                "table_ids": [table_id],
                "field": cand["field"],
                "formula": cand["formula"],
                "confidence": round(fc, 2),
            }

    print(f"[INFO] 最终 note_map: {len(final_note_map)} 个科目", file=sys.stderr)

    # 核心科目覆盖校验
    CORE_SUBJECTS = [
        "货币资金", "应收账款", "存货", "固定资产", "短期借款",
        "应付账款", "营业收入", "营业成本", "净利润"
    ]
    covered = sum(1 for s in CORE_SUBJECTS if s in final_note_map)
    if covered < 8:
        print(f"[WARN] 核心科目覆盖不足（{covered}/{len(CORE_SUBJECTS)}），建议人工复核", file=sys.stderr)

    return final_note_map


# ──────────────────────────────────────────────────────────────────
# 测试入口
# ──────────────────────────────────────────────────────────────────
def main():
    """测试入口。"""
    import argparse

    parser = argparse.ArgumentParser(description="DeepSeek 并发加速层测试")
    parser.add_argument("--test-check-key", action="store_true", help="测试 check_api_key")
    parser.add_argument("--test-list-models", action="store_true", help="测试 list_models")
    parser.add_argument("--api-key", help="API key（临时，不持久化）")

    args = parser.parse_args()

    # 测试 check_api_key
    if args.test_check_key:
        has_key, config = check_api_key(args.api_key)
        print(f"check_api_key 结果: has_key={has_key}")
        if config:
            print(f"配置: {config}")
        sys.exit(0 if has_key else 1)

    # 测试 list_models
    if args.test_list_models:
        has_key, config = check_api_key(args.api_key)
        if not has_key:
            print("错误: 无有效 API key", file=sys.stderr)
            sys.exit(1)

        client = DeepSeekClient(
            api_key=config["api_key"],
            model=config["model"],
            base_url=config["base_url"],
        )

        models = list_models(client)
        print("可用模型列表:")
        for m in models:
            print(f"  - {m}")
        sys.exit(0)


if __name__ == "__main__":
    main()