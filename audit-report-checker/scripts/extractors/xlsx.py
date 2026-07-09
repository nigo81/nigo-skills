"""extractors/xlsx.py — xlsx/xlsm 源 → 统一行迭代器。"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


def load_xlsx_sheets(xlsx_path: Path) -> dict:
    """加载 xlsx/xlsm 所有 sheet，返回 {sheet名: worksheet}。"""
    if openpyxl is None:
        raise RuntimeError("需要 openpyxl")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    return {sn: wb[sn] for sn in wb.sheetnames}


def iter_xlsx_rows(sheets: dict, sheet_names: list):
    """按 sheet_names 顺序产生行（每行是单元格值列表）。"""
    for sn in sheet_names:
        ws = sheets.get(sn)
        if ws is None:
            print(f"[WARN] sheet '{sn}' 不存在", file=sys.stderr)
            continue
        for row in ws.iter_rows(values_only=True):
            yield list(row) if row else []
